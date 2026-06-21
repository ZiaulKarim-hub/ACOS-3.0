#!/usr/bin/env python3
"""
Thurston Lux II & Haystack — acos-dataroom 'lender_package' include/exclude analysis.

Advisory only. Applies the acos-dataroom skill's Phase 3/5/7b classification logic
to the SOURCE folders for a takeout-lender document request (email items a-e), and
reviews the EXISTING dataroom against the same standard.

Classification is filename/folder-based (not full text extraction) — see the
Methodology note in the workbook. Genuinely ambiguous files are marked REVIEW.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STAGING = "/Users/zee/Thurston Staging"
SRC_ROOTS = ["Thurston Lux 2", "Thurston_Haystack"]
DR_ROOT = "Lux II & Haystack Data Room"
OUT = "/Users/zee/Documents/Vibe Coding/ACOS 3.0/Thurston_LuxII_Haystack_DataRoom_Include-Exclude_Analysis_2026-05-18.xlsx"

SYSTEM_EXT = {".ppinfocache", ".ptn", ".ptn2", ".ini2", ".webloc", ".log"}
MEDIA_EXT = {".png", ".jpg", ".jpeg", ".gif", ".heic", ".bmp", ".mov", ".mp4", ".tif", ".tiff"}

# ---------------------------------------------------------------- request item
def request_item(nl):
    """Map an included file to the email's request scope (a)-(e) or supporting set."""
    if any(k in nl for k in ["allonge"]):
        return "(a) Note / allonge"
    if "note" in nl and "footnote" not in nl:
        return "(a) Promissory note"
    if "loan agreement" in nl or "term loan agreement" in nl:
        return "(a) Loan agreement"
    if any(k in nl for k in ["modification", "extension", "amendment"]):
        return "(a) Amendment / modification"
    if any(k in nl for k in ["mortgage", "deed of trust", "trust deed", " dot ", "deed of trust"]):
        return "(b) Mortgage / deed of trust"
    if "ucc" in nl:
        return "(c) UCC-1 financing statement"
    if "guarant" in nl:
        return "(d) Guaranty"
    if "pledge" in nl:
        return "(d) Pledge agreement"
    if "security agreement" in nl:
        return "(d) Security agreement"
    if "environmental indemnity" in nl:
        return "(d) Environmental indemnity"
    if any(k in nl for k in ["intercreditor", "subordination", "standstill", "participation", "side letter"]):
        return "(e) Intercreditor / subordination"
    if any(k in nl for k in ["nod", "notice of default", "rfn", "request for notice",
                             "sot", "substitution of trustee"]):
        return "Recorded loan-status instrument"
    if any(k in nl for k in ["written consent", "authorization", "resolution", "omnibus", "ominibus"]):
        return "Closing set — entity authority"
    if "title policy" in nl or "alta loan policy" in nl or "loan policy" in nl or "lenders policy" in nl:
        return "Closing set — title policy"
    if "settlement statement" in nl or "closing statement" in nl:
        return "Closing set — settlement"
    if "firpta" in nl or "foreign affidavit" in nl or "nonforeign" in nl or "non foreign" in nl or "non-foreign" in nl:
        return "Closing set — FIRPTA / affidavit"
    if "escrow" in nl:
        return "Closing set — escrow"
    if "opinion letter" in nl:
        return "Closing set — legal opinion"
    if "signed docs" in nl or "loan docs" in nl:
        return "Closing set — executed loan package"
    return "Closing set — supporting"

# ---------------------------------------------------------------- classifier
def classify(relpath):
    """Return (verdict, item, category, rationale, sensitivity, confidence)."""
    p = relpath.lower()
    name = os.path.basename(relpath)
    nl = name.lower()
    ext = os.path.splitext(nl)[1]

    # --- system / media / archive --------------------------------------------
    if ext in SYSTEM_EXT or nl in (".ds_store",) or nl.startswith("~wrd") or nl == "maxdesk.ini2":
        return ("EXCLUDE", "—", "System file",
                "Operating-system / application cache file — not a document.", "Low", "High")
    if ext in MEDIA_EXT:
        return ("EXCLUDE", "—", "Image / media",
                "Image, photo or video file — not a loan instrument.", "Low", "High")
    if ext == ".zip":
        return ("EXCLUDE", "—", "Archive",
                "Zip archive — out of scope; loose contents handled individually.", "Low", "High")

    # --- broad out-of-scope diligence folders --------------------------------
    if "construction - deer run" in p:
        return ("EXCLUDE", "—", "Construction draws",
                "Construction draw package (invoices, inspection reports, budgets) — "
                "project diligence, not a loan instrument; outside the takeout lender's request.",
                "Low", "High")
    if "/appraisals/" in p or "thurston hawaii appraisals" in p or "property value refresh" in p:
        return ("EXCLUDE", "—", "Appraisal / valuation",
                "Appraisal or valuation refresh — collateral diligence, outside the loan-document request.",
                "Low", "High")
    if "/background checks/" in p or nl.startswith("background check"):
        return ("EXCLUDE", "—", "Background check",
                "Borrower background check — sensitive personal information; not a loan instrument; withhold.",
                "High", "High")
    if "/deal sheet & credit memos/" in p:
        return ("EXCLUDE", "—", "Internal credit memo",
                "Internal credit memo / deal sheet — OKOA work product; strategic; never shared with a counterparty.",
                "Critical", "High")
    if "/property insurance/" in p or "insurance non-renewal" in p or "insurance reinststement" in p \
            or "miami eoi" in nl:
        return ("EXCLUDE", "—", "Insurance",
                "Property insurance certificate / notice — collateral diligence, outside the loan-document request.",
                "Low", "High")
    if "/property lists/" in p or "master property list" in nl or "property tax" in nl \
            or "loan offer comparis" in nl:
        return ("EXCLUDE", "—", "Internal list / analysis",
                "Internal portfolio list or comparison sheet — OKOA work product, outside the request.",
                "Medium", "High")
    if "/property sales/" in p:
        return ("EXCLUDE", "—", "Property-sale workout",
                "Property-sale / partial-release workout file (REPC, payoff, lien recon) — "
                "a separate transaction; outside the existing-loan document request.",
                "Medium", "High")
    if "/fairview and okoa payoffs/" in p or "fairview" in nl or "payoff letter" in nl \
            or "payoff" in nl and ext in (".pdf", ".docx") and "payoffs" in p:
        return ("EXCLUDE", "—", "Prior-loan payoff",
                "Prior-lender (Fairview) payoff / reconveyance file — a different loan; outside the request.",
                "Medium", "Medium")
    if "/variant counsel - dropbox (shared)/" in p:
        return ("EXCLUDE", "—", "Participant-counsel duplicate",
                "Borrower-org / title material duplicated into the participant's (Variant) counsel folder — "
                "duplicate and/or non-loan-instrument; outside the request.",
                "Medium", "High")
    if "/communications/" in p:
        return ("EXCLUDE", "—", "Correspondence",
                "Internal correspondence — not a loan instrument; outside the request.", "Medium", "High")
    if "variant distribution & waterfall" in p or "variant buyout" in p:
        return ("EXCLUDE", "—", "Participant economics",
                "Variant participant economics / waterfall model — internal capital-stack detail; "
                "reveals OKOA economics; withhold.", "Critical", "High")
    if "variant legal reimbursementn" in p:
        return ("EXCLUDE", "—", "Legal-fee file",
                "Legal-fee reimbursement file — administrative; outside the request.", "Medium", "High")
    if "loi archived drafts" in p:
        return ("EXCLUDE", "—", "Superseded draft",
                "Archived letter-of-intent draft — superseded; outside the request.", "Low", "High")
    if "lux 2 final docs unsigned" in p:
        return ("EXCLUDE", "—", "Unsigned final-form",
                "Unsigned final-form document — superseded by the executed copies; withhold the unsigned version.",
                "Low", "High")
    if "/closing docs/final loan docs/" in p:
        return ("EXCLUDE", "—", "Unsigned final-form (Word)",
                "Final-form Word working file — superseded by the executed PDFs under Signed Loan Docs.",
                "Low", "High")

    # --- participation agreements (judgment call — surface for boss) ---------
    if "investor participation agreement" in p or ("/variant/" in p and "participation" in nl) \
            or "participation agreement" in nl or "side letter with variant" in nl \
            or ("variant-okoa" in nl and "participation" in nl) or "participation blackline" in nl:
        return ("REVIEW", "(e) Intercreditor / subordination", "Loan participation",
                "Loan participation agreement (Variant / investors). NOT a document the takeout lender "
                "ordinarily needs to retire the note, and it reveals OKOA's internal economics — default "
                "is WITHHOLD. BUT if it contains intercreditor / subordination / standstill terms it becomes "
                "responsive to request item (e). Boss decision required.",
                "Critical", "Medium")

    # --- org docs (Lux 2) ----------------------------------------------------
    if "/org docs/" in p:
        if "opinion letter" in nl:
            return ("INCLUDE", "Closing set — legal opinion", "Legal opinion",
                    "Borrower's counsel closing opinion letter — a deliverable of the loan closing; "
                    "executed copy.", "Low", "Medium")
        return ("EXCLUDE", "—", "Borrower org document",
                "Borrower entity / organizational document (operating agreement, certificate, resolution) — "
                "not a loan instrument; outside the strict loan-document request (the new lender can request "
                "entity docs separately).", "Medium", "Medium")
    if "/borrower & entity info/" in p:
        return ("EXCLUDE", "—", "Borrower org document",
                "Borrower entity / authorization document — not a loan instrument; outside the request.",
                "Medium", "Medium")
    # existing-dataroom folder vocabulary: Borrower Documents subtree
    if "/borrower documents/" in p:
        if "/w-9s/" in p or nl.endswith("w9.pdf") or " w9" in nl:
            return ("EXCLUDE", "—", "Borrower tax form (PII)",
                    "Borrower W-9 — taxpayer-identification PII; not a loan instrument; withhold.",
                    "High", "High")
        return ("EXCLUDE", "—", "Borrower org document",
                "Borrower entity / organizational document (articles, partnership agreement, good-standing "
                "certificate, charter, living trust) — not a loan instrument; outside the loan-document "
                "request.", "Medium", "High")
    # existing-dataroom folder vocabulary: Loan Extension subtree
    if "/loan extension/" in p:
        if "signed loan extension docs" in p:
            return ("INCLUDE", "(a) Amendment / modification", "Executed amendment",
                    "Executed loan extension / modification — an amendment to the existing loan "
                    "(request item a).", "Low", "High")
        return ("EXCLUDE", "—", "Extension diligence",
                "Loan-extension supporting file (valuation refresh, draft) — outside the request.",
                "Low", "Medium")
    # existing-dataroom folder vocabulary: Haystack NOD folder
    if "/nod/" in p:
        return ("INCLUDE", "Recorded loan-status instrument", "Recorded foreclosure instrument",
                "Recorded notice-of-default / request-for-notice / substitution-of-trustee — a recorded "
                "instrument evidencing the current status of OKOA's lien.", "High", "Medium")

    # --- Haystack foreclosure folder ----------------------------------------
    if "thurston_haystack/foreclosure/" in p:
        if "lien clarification" in p:
            return ("EXCLUDE", "—", "Duplicate",
                    "Copy of a title / signed-docs file duplicated into the lien-clarification working "
                    "subfolder — duplicate.", "Low", "Medium")
        if any(k in nl for k in ["nod recorded", "rfn recorded", "sot recorded", "signed sot"]):
            return ("INCLUDE", "Recorded loan-status instrument", "Recorded foreclosure instrument",
                    "Recorded notice-of-default / request-for-notice / substitution-of-trustee — a recorded "
                    "instrument evidencing the current status of OKOA's lien; relevant to a takeout lender.",
                    "High", "Medium")
        if "default letter" in nl or "foreclosure report" in nl or "transmit" in nl or "sot.family farmers" in nl:
            return ("EXCLUDE", "—", "Internal foreclosure work product",
                    "Internal default letter / foreclosure report / transmittal — OKOA work product; "
                    "sensitive; withhold.", "Critical", "Medium")
        if "alta loan policy" in nl or "signed docs" in nl or "772 n haystack" in nl:
            return ("EXCLUDE", "—", "Duplicate",
                    "Copy of a closing document duplicated into the Foreclosure folder — duplicate "
                    "(primary copy classified under Closing Docs).", "Low", "Medium")
        return ("REVIEW", "—", "Foreclosure folder item",
                "File in the Haystack foreclosure folder — review individually.", "High", "Low")
    if "/foreclosure/" in p:  # Lux 2 foreclosure
        return ("EXCLUDE", "—", "Internal foreclosure work product",
                "Internal default-letter draft — OKOA work product; sensitive; withhold.",
                "Critical", "High")

    # --- Haystack draft docs -------------------------------------------------
    if "thurston_haystack/draft docs/" in p:
        return ("EXCLUDE", "—", "Superseded draft",
                "Draft Haystack closing document / settlement statement — superseded by the executed "
                "'FAMILY FARMERS SIGNED DOCS' set.", "Low", "Medium")

    # --- Haystack closing docs ----------------------------------------------
    if "thurston_haystack/closing docs/" in p:
        if "wire" in nl:
            return ("EXCLUDE", "—", "Wire confirmation",
                    "Wire / funding confirmation — banking detail, not a loan instrument; withhold.",
                    "High", "Medium")
        if "qcd" in nl:
            return ("REVIEW", "(b) Mortgage / deed of trust", "Deed",
                    "Quitclaim deed (qcd) — a recorded deed; review whether it pertains to the collateral "
                    "title chain before release.", "Medium", "Low")
        if "alta loan policy" in nl:
            return ("INCLUDE", "Closing set — title policy", "Title policy",
                    "ALTA loan policy of title insurance for the Haystack collateral — executed/issued.",
                    "Low", "Medium")
        if "loan agreement" in nl:
            return ("INCLUDE", "(a) Loan agreement", "Loan agreement",
                    "Haystack (Family Farmers) loan agreement — core loan instrument.", "Low", "Medium")
        if "signed docs" in nl:
            return ("INCLUDE", "Closing set — executed loan package", "Executed loan package",
                    "Executed Haystack (Family Farmers) signed-documents package — core loan instruments.",
                    "Medium", "Medium")
        if "772 n haystack" in nl:
            return ("INCLUDE", "(b) Mortgage / deed of trust", "Recorded security instrument",
                    "Recorded Puerto Rico mortgage on 772 N Haystack — security instrument.", "Low", "Medium")
        if nl.startswith("20240911"):
            return ("INCLUDE", "Closing set — executed loan package", "Recorded loan instrument",
                    "Date-stamped recorded Haystack closing instrument (Sept 2024 closing).", "Low", "Low")
        return ("REVIEW", "—", "Haystack closing item",
                "Haystack closing-docs file — review individually.", "Medium", "Low")

    # --- Dorado Properties (Lux 2) ------------------------------------------
    if "/dorado properties/" in p:
        if "pr mortgages" in nl:
            return ("INCLUDE", "(b) Mortgage / deed of trust", "Recorded security instrument",
                    "Recorded Puerto Rico mortgages — security instruments encumbering the PR collateral.",
                    "Low", "Medium")
        return ("EXCLUDE", "—", "Property-level diligence",
                "Property-level diligence (BPO, property deeds, screen recordings) — outside the "
                "loan-document request.", "Medium", "Medium")

    # --- Title folder (Lux 2) -----------------------------------------------
    if "thurston lux 2/title/" in p:
        if "lien waivers on ut properties" in p:
            if "indemnity agreement construction" in nl:
                return ("REVIEW", "(d) Credit support", "Construction indemnity",
                        "Construction title-indemnity agreement (STG) — affects title/lien; review whether "
                        "responsive to the request.", "Medium", "Low")
            return ("EXCLUDE", "—", "Title clearance work product",
                    "Lien waiver / owner's affidavit / filing-status working file — title-clearance work "
                    "product, not a loan instrument.", "Low", "Medium")
        if "deed of trust (recorded)" in nl:
            return ("INCLUDE", "(b) Mortgage / deed of trust", "Recorded security instrument",
                    "Recorded deed of trust — security instrument encumbering the collateral.",
                    "Low", "High")
        if "ucc financing statement (recorded)" in nl:
            return ("INCLUDE", "(c) UCC-1 financing statement", "Recorded UCC",
                    "Recorded UCC financing statement — perfects OKOA's security interest.", "Low", "High")
        if "lien & financing statement" in nl:
            return ("REVIEW", "(c) UCC-1 financing statement", "Lien / financing statement",
                    "Lien & financing statement — review whether this is a recorded UCC pertaining to "
                    "the facility.", "Low", "Low")
        if "master property title insurance" in nl:
            return ("EXCLUDE", "—", "Title work product",
                    "Master title-insurance summary — internal title tracking, not an issued policy.",
                    "Low", "Medium")
        return ("EXCLUDE", "—", "Title production work product",
                "Title-production artifact (commitment, pro-forma, CPL, estimated invoice, owner's "
                "affidavit, legal opinion draft) — superseded by the issued title policies; not a loan "
                "instrument.", "Low", "Medium")

    # --- Lux 2 Closing Folder (curated executed set) ------------------------
    if "lux 2 closing folder/" in p:
        if "title policies/" in p:
            return ("INCLUDE", "Closing set — title policy", "Title policy",
                    "Issued loan title-insurance policy for a collateral property.", "Low", "High")
        if "ucc filings (recorded)/" in p:
            return ("INCLUDE", "(c) UCC-1 financing statement", "Recorded UCC",
                    "Recorded UCC-1 financing statement — perfects OKOA's security interest.", "Low", "High")
        if "written consents/" in p:
            return ("INCLUDE", "Closing set — entity authority", "Written consent",
                    "Borrower written consent authorizing the loan — closing-set authority document.",
                    "Low", "High")
        # root of Lux 2 Closing Folder = curated executed instruments
        return ("INCLUDE", request_item(nl), "Executed loan instrument",
                "Executed copy of a core loan instrument, filed in the curated Lux II closing folder.",
                "Medium" if ("settlement" in nl or "escrow" in nl) else "Low", "High")

    # --- Signed Loan Docs (Lux 2) -------------------------------------------
    if "/signed loan docs/" in p:
        if "buyer" in nl and "closing statement" in nl:
            return ("INCLUDE", "Closing set — settlement", "Settlement statement",
                    "Buyer's / lender's closing statement — closing-set document; contains figures and "
                    "may contain wire detail (consider redaction).", "High", "Medium")
        if "miami property originals" in nl:
            return ("REVIEW", "(b) Mortgage / deed of trust", "Original instruments",
                    "'Lux II Miami Property Originals' — likely original recorded instruments for the "
                    "Miami collateral; review contents.", "Low", "Low")
        if "limited liability limited partnership agreement affidavit" in nl:
            return ("INCLUDE", "Closing set — FIRPTA / affidavit", "Affidavit",
                    "Partnership-agreement affidavit — executed closing-set affidavit.", "Low", "Medium")
        return ("INCLUDE", request_item(nl), "Executed loan instrument",
                "Executed loan document from the signed-loan-docs set — directly responsive to the "
                "takeout lender's request.",
                "High" if ("settlement" in nl or "escrow" in nl) else "Low", "High")

    # --- Puerto Rico Notes (Lux 2) ------------------------------------------
    if "/puerto rico notes/" in p:
        return ("INCLUDE", request_item(nl), "Executed loan instrument",
                "Executed Puerto Rico note / mortgage-note pledge — core loan instrument for the PR "
                "collateral.", "Low", "High")

    # --- Lux 2 Extension -----------------------------------------------------
    if "lux 2 extension/" in p:
        if "signed extension" in p or "signed lux ii loan extension" in nl:
            return ("INCLUDE", "(a) Amendment / modification", "Executed amendment",
                    "Executed loan extension / modification — an amendment to the existing loan "
                    "(request item a).", "Low", "High")
        if "promissory note" in nl:
            return ("REVIEW", "(a) Amendment / modification", "Extension note draft",
                    "Extension-period promissory note — confirm whether this is the executed extension "
                    "note before release (drafts present).", "Low", "Low")
        return ("EXCLUDE", "—", "Superseded extension draft",
                "Unsigned loan-modification / authorization draft — superseded by the executed extension "
                "documents.", "Low", "Medium")

    # --- Closing Docs root (Lux 2) — mostly working drafts ------------------
    if "thurston lux 2/closing docs/" in p:
        # written consents (existing-dataroom 'Closing Docs/Written Consents')
        if "/written consents/" in p:
            return ("INCLUDE", "Closing set — entity authority", "Written consent",
                    "Borrower written consent authorizing the loan — closing-set authority document.",
                    "Low", "High")
        # the UCCs subfolder
        if "/closing docs/uccs/" in p:
            if "firpta" in nl:
                return ("EXCLUDE", "—", "Misfiled / superseded",
                        "FIRPTA Word draft filed in the UCCs folder — superseded; FIRPTA is not a UCC.",
                        "Low", "Medium")
            if "transaction checklist" in nl:
                return ("EXCLUDE", "—", "Internal checklist",
                        "Internal transaction checklist — OKOA work product.", "Medium", "High")
            return ("EXCLUDE", "—", "Pre-recording UCC version",
                    "Prepared (pre-recording) UCC-1 / addendum — superseded by the file-stamped recorded "
                    "UCCs under 'UCC Filings (recorded)'.", "Low", "Medium")
        # recorded instruments loose in Closing Docs root
        if "recorded" in nl or "ucc1pr" in nl or nl.startswith("pr ucc-1") \
                or nl.startswith("ucc1-") or nl.startswith("ucc1 -") or "ucc - d-001" in nl \
                or "20250129 mortgage" in nl or "20250129 ucc-1" in nl:
            it = "(c) UCC-1 financing statement" if "ucc" in nl else "(b) Mortgage / deed of trust"
            return ("INCLUDE", it, "Recorded security instrument",
                    "Recorded security instrument (mortgage / deed of trust / UCC) loose in the "
                    "closing-docs folder — perfects or evidences OKOA's lien.", "Low", "Medium")
        if "deed of trust- utah county, ut.pdf" == nl:
            return ("INCLUDE", "(b) Mortgage / deed of trust", "Security instrument",
                    "Utah County deed of trust — security instrument (executed copy also under Signed "
                    "Loan Docs).", "Low", "Medium")
        if "alta loan policy" in nl and "proforma" not in nl:
            return ("INCLUDE", "Closing set — title policy", "Title policy",
                    "Issued ALTA loan policy of title insurance.", "Low", "Medium")
        if nl.startswith("title policy"):
            return ("INCLUDE", "Closing set — title policy", "Title policy",
                    "Issued loan title-insurance policy.", "Low", "Medium")
        if "lenders policy" in nl:
            return ("INCLUDE", "Closing set — title policy", "Title policy",
                    "Lender's title-insurance policy.", "Low", "Medium")
        if "ucc fixture filing" in nl:
            return ("REVIEW", "(c) UCC-1 financing statement", "Fixture filing",
                    "UCC fixture filing (Dorado property) — review whether recorded and pertaining to "
                    "the facility.", "Low", "Low")
        if "mortgage note pledge" in nl and ext == ".doc":
            return ("EXCLUDE", "—", "Superseded draft",
                    "Mortgage-note-pledge Word draft — superseded by the executed copies under "
                    "'Puerto Rico Notes'.", "Low", "Medium")
        if "okoa_servicing_agreement" in nl:
            return ("REVIEW", "—", "Servicing agreement",
                    "OKOA servicing agreement — review: relevant to who services the loan, but may be "
                    "internal; boss decision.", "Medium", "Low")
        if "litera redline" in nl or "redline" in nl or "blackline" in nl:
            return ("EXCLUDE", "—", "Drafting artifact",
                    "Redline / blackline comparison — a drafting artifact, not an executed instrument.",
                    "Low", "High")
        if "central policy distribution" in nl:
            return ("REVIEW", "Closing set — title policy", "Title-related",
                    "'Central Policy Distribution LP' — review; appears title-policy-distribution related.",
                    "Low", "Low")
        if "invoice" in nl or "image00" in nl:
            return ("EXCLUDE", "—", "Administrative / image",
                    "Invoice or embedded image — not a loan instrument.", "Low", "High")
        if "lux loan docs" in nl:
            return ("REVIEW", "Closing set — executed loan package", "Loan-docs compilation",
                    "'Lux Loan Docs' compilation — review whether executed before release.", "Low", "Low")
        # default: working draft / superseded version
        return ("EXCLUDE", "—", "Working draft / superseded version",
                "Working draft or superseded version (Word draft, '- Version N', dated draft, "
                "'Not Sent') — the executed copy is filed under Signed Loan Docs / Lux 2 Closing Folder.",
                "Low", "Medium")

    # --- Lux 2 root-level loose files ---------------------------------------
    if p.count("/") == 1 and p.startswith("thurston lux 2/"):
        if "scoping letter" in nl:
            return ("EXCLUDE", "—", "Pre-closing scoping letter",
                    "Loan scoping letter — pre-closing engagement document, not a loan instrument.",
                    "Low", "Medium")
        if "strategic assessment" in nl:
            return ("EXCLUDE", "—", "Internal analysis",
                    "Internal strategic assessment of the portfolio loan — OKOA work product; withhold.",
                    "Critical", "High")
        if "title commitment" in nl:
            return ("EXCLUDE", "—", "Title commitment",
                    "Title commitment — superseded by the issued title policies; title-production "
                    "artifact.", "Low", "Medium")
        if "genesis okoa loan split" in nl:
            return ("REVIEW", "—", "Loan-split / facility structure",
                    "'Genesis Okoa Loan Split' — appears to describe how the facility is structured / "
                    "split; review for the note-reconciliation question (single $55M note vs. two loans).",
                    "High", "Low")
        if "payoff" in nl:
            return ("EXCLUDE", "—", "Internal payoff calculation",
                    "Internal payoff calculation worksheet — OKOA work product.", "High", "Medium")
        if "participation agreement" in nl:
            return ("REVIEW", "(e) Intercreditor / subordination", "Loan participation",
                    "Loan participation agreement — see participation note; boss decision required.",
                    "Critical", "Medium")
        if "term sheet" in nl or "draft term sheet" in nl:
            return ("EXCLUDE", "—", "Term sheet",
                    "Draft term sheet — pre-closing; superseded by the executed loan documents.",
                    "Medium", "Medium")
        if "legal opinion" in nl:
            return ("REVIEW", "Closing set — legal opinion", "Legal opinion",
                    "Legal opinion (title / homestead) — review whether final and responsive.",
                    "Low", "Low")
        if "escrow instruction" in nl:
            return ("EXCLUDE", "—", "Superseded draft",
                    "Escrow-instruction-letter draft — superseded by the executed escrow letter.",
                    "Low", "Medium")
        if "title requirements" in nl or "comp25659" in nl or "updated numbers" in nl \
                or "magleby" in nl or "sundance property insurance" in nl:
            return ("EXCLUDE", "—", "Diligence / administrative",
                    "Title-requirements, comparable, numbers or administrative file — outside the request.",
                    "Medium", "Medium")
        return ("REVIEW", "—", "Root-level loose file",
                "Loose file at the Lux II folder root — review individually.", "Medium", "Low")

    # --- Haystack root-level loose files ------------------------------------
    if p.count("/") == 1 and p.startswith("thurston_haystack/"):
        if "payoff" in nl:
            return ("EXCLUDE", "—", "Internal payoff calculation",
                    "Internal Haystack payoff calculation worksheet — OKOA work product.", "High", "Medium")
        return ("REVIEW", "—", "Root-level loose file",
                "Loose file at the Haystack folder root — review individually.", "Medium", "Low")

    # --- fallback ------------------------------------------------------------
    return ("REVIEW", "—", "Unclassified",
            "Not matched by a rule — review individually.", "Medium", "Low")

# ---------------------------------------------------------------- walk
def walk(root):
    rows = []
    base = os.path.join(STAGING, root)
    for dirpath, _, files in os.walk(base):
        for f in files:
            if f == ".DS_Store":
                continue
            full = os.path.join(dirpath, f)
            rel = os.path.relpath(full, base)
            rows.append((root, rel, f))
    return sorted(rows, key=lambda r: r[1].lower())

src_rows = []
for r in SRC_ROOTS:
    src_rows += walk(r)
dr_rows = walk(DR_ROOT)

# dataroom basenames for presence check
dr_names = {f.lower() for (_, _, f) in dr_rows}

# ---------------------------------------------------------------- workbook
wb = Workbook()
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3A5F")
SUB_FONT = Font(bold=True, size=11, color="1F3A5F")
WRAP = Alignment(wrap_text=True, vertical="top")
VERDICT_FILL = {
    "INCLUDE": PatternFill("solid", fgColor="C6EFCE"),
    "EXCLUDE": PatternFill("solid", fgColor="F2F2F2"),
    "REVIEW":  PatternFill("solid", fgColor="FFEB9C"),
}

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER

# ============================================================ Sheet 1: Summary
ws = wb.active
ws.title = "Summary & Findings"
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", (3, 30, 16, 16, 16, 60)):
    ws.column_dimensions[col].width = w

def put(row, col, val, font=None, fill=None, wrap=False, span=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font: cell.font = font
    if fill: cell.fill = fill
    if wrap: cell.alignment = WRAP
    if span:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    return cell

r = 2
put(r, 2, "Thurston Lux II & Haystack — Data Room Include / Exclude Analysis", TITLE_FONT, span=5); r += 1
put(r, 2, "Advisory output of the acos-dataroom skill (deal type: lender_package). "
          "No data room was created or modified.", Font(italic=True, size=9, color="595959"), span=5); r += 2

# counts
src_class = [(root, rel, name, classify(os.path.join(root, rel) if False else rel if False else
                                        (root + "/" + rel))) for (root, rel, name) in src_rows]
# fix: classify on path including root for foreclosure rules etc.
src_class = []
for (root, rel, name) in src_rows:
    relpath = root + "/" + rel
    src_class.append((root, rel, name, classify(relpath)))
dr_class = []
for (root, rel, name) in dr_rows:
    # reclassify dataroom files: strip the DR_ROOT, re-root under loan label so rules fire
    sub = rel
    if sub.startswith("Lux II Loan/"):
        relpath = "Thurston Lux 2/" + sub[len("Lux II Loan/"):]
    elif sub.startswith("Haystack Loan/"):
        relpath = "Thurston_Haystack/" + sub[len("Haystack Loan/"):]
    else:
        relpath = "Thurston Lux 2/" + sub
    dr_class.append((root, rel, name, classify(relpath)))

def tally(rows):
    t = {"INCLUDE": 0, "EXCLUDE": 0, "REVIEW": 0}
    for *_, c in rows:
        t[c[0]] += 1
    return t

st = tally(src_class)
dt = tally(dr_class)

put(r, 2, "Scope of this analysis", SUB_FONT, span=5); r += 1
for line in [
    "Objective: a takeout lender refinancing the borrower will pay off OKOA's existing loan at "
    "closing and has requested OKOA's complete set of existing loan documents.",
    "Source folders analysed: 'Thurston Lux 2' and 'Thurston_Haystack' (846 files).",
    "Existing data room reviewed (read-only, NOT modified): 'Lux II & Haystack Data Room' (261 files).",
    "Request scope (email): a complete set of loan documents evidencing the note, including "
    "(a) the master promissory note and amendments, (b) recorded mortgages / deeds of trust on each "
    "collateral property, (c) all UCC-1 financing statements, (d) guarantees / pledges / credit-support "
    "documents, and (e) intercreditor / subordination / standstill agreements.",
]:
    put(r, 2, "•", Font(size=9)); put(r, 3, line, Font(size=9), wrap=True, span=4)
    ws.row_dimensions[r].height = 30; r += 1
r += 1

put(r, 2, "Result — source files", SUB_FONT, span=5); r += 1
put(r, 2, "Verdict", HEAD_FONT, HEAD_FILL); put(r, 3, "Source", HEAD_FONT, HEAD_FILL)
put(r, 4, "Existing room", HEAD_FONT, HEAD_FILL)
ws.cell(row=r, column=2).alignment = ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
r += 1
for v in ("INCLUDE", "REVIEW", "EXCLUDE"):
    put(r, 2, v, Font(bold=True, size=9), VERDICT_FILL[v])
    put(r, 3, st[v], Font(size=9))
    put(r, 4, dt[v], Font(size=9))
    r += 1
put(r, 2, "TOTAL", Font(bold=True, size=9)); put(r, 3, sum(st.values()), Font(bold=True, size=9))
put(r, 4, sum(dt.values()), Font(bold=True, size=9)); r += 2

put(r, 2, "Key findings", SUB_FONT, span=5); r += 1
findings = [
    ("Headline", "The existing data room is a broad diligence room (appraisals, ~73 construction-draw "
     "files, 27 property photos/videos, insurance, background checks, borrower org docs). The takeout "
     "lender's request is narrow — loan instruments only. Run for THIS request, the skill would include "
     "only the executed loan documents and recorded security instruments and exclude the diligence "
     "material as out of scope for this request."),
    ("GAP — executed Guaranty", "The email expressly requests guarantees (item d). The source folders "
     "contain only UNSIGNED drafts (Guarantee - Version 1, Guarantee - 2025.05.26, Guarantee - Redline). "
     "No executed guaranty was located in Signed Loan Docs. Confirm whether an executed guaranty exists "
     "before responding to the new lender."),
    ("Note reconciliation", "The request describes a single ~$55,000,000 master note over twelve "
     "properties. The source material spans TWO loans — Lux II (one file calls it a '$48.16M Luxury "
     "Portfolio Loan') and a separate Haystack / Family Farmers loan. Confirm the exact note(s), "
     "amount and collateral list before release. See 'Genesis Okoa Loan Split'."),
    ("Haystack in foreclosure", "The Haystack / Family Farmers loan shows a recorded Notice of Default, "
     "Request for Notice and Substitution of Trustee. A takeout lender must be told the loan status. "
     "Recorded NOD/RFN/SOT are included; internal default letters and the foreclosure report are "
     "excluded as OKOA work product."),
    ("Participation agreements", "Variant and investor (Adam Schanz, Kyle Demourdant, Sopris Trust) "
     "participation agreements are marked REVIEW. They are not ordinarily given to a takeout lender and "
     "reveal OKOA's economics — default is WITHHOLD — but if any contains intercreditor / subordination "
     "/ standstill terms it becomes responsive to item (e). Boss decision required."),
    ("Drafts vs. executed", "The folders hold many superseded drafts, redlines and '- Version N' files. "
     "Only executed / recorded copies are included; drafts are excluded as superseded."),
    ("Sensitivity", "Background checks, internal credit memos, the strategic assessment, payoff "
     "worksheets and participant waterfall models are excluded as internal / strategic. Settlement and "
     "escrow documents that are included may contain wire detail — consider redaction before release."),
]
for tag, txt in findings:
    put(r, 2, tag, Font(bold=True, size=9, color="1F3A5F"), wrap=True)
    put(r, 3, txt, Font(size=9), wrap=True, span=4)
    ws.row_dimensions[r].height = max(45, 12 * (len(txt) // 70 + 1))
    r += 1
r += 1

put(r, 2, "Methodology & limitations", SUB_FONT, span=5); r += 1
for line in [
    "This analysis applies the acos-dataroom skill's Phase 3 (inventory), Phase 5 (classification) and "
    "Phase 7b (three-state status) logic for deal type 'lender_package'.",
    "Classification is filename- and folder-based. Unlike a full skill run it does not extract and read "
    "the text inside each document. Files that cannot be adjudicated from name and location are marked "
    "REVIEW rather than guessed.",
    "Verdicts: INCLUDE = responsive executed/recorded loan document; EXCLUDE = out of scope, superseded, "
    "duplicate, internal or sensitive; REVIEW = open the file and decide.",
    "'In existing room' on the Source sheet is an approximate filename match against the existing data "
    "room and is indicative only (the room renamed some files).",
    "No data room was created, modified, renamed or shared. The existing data room was read only.",
    "Not legal advice. Confirm the final release set with counsel before sending anything to the new "
    "lender.",
]:
    put(r, 2, "•", Font(size=9)); put(r, 3, line, Font(size=9), wrap=True, span=4)
    ws.row_dimensions[r].height = 28; r += 1

# ============================================================ Sheet 2: Source
ws = wb.create_sheet("Source File Analysis")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"
headers = ["Loan", "Folder", "File name", "Verdict", "Request item",
           "Category", "Rationale", "Sensitivity", "Confidence", "In existing room?"]
widths = [14, 40, 38, 11, 24, 24, 60, 12, 11, 16]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, 1, len(headers))
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(src_class)+1}"

rownum = 2
for (root, rel, name, c) in src_class:
    verdict, item, cat, rationale, sens, conf = c
    folder = os.path.dirname(rel) or "(root)"
    loan = "Lux II" if root == "Thurston Lux 2" else "Haystack"
    indr = "Yes" if name.lower() in dr_names else "No"
    vals = [loan, folder, name, verdict, item, cat, rationale, sens, conf, indr]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=rownum, column=i, value=v)
        cell.font = Font(size=9)
        cell.border = BORDER
        cell.alignment = WRAP
    ws.cell(row=rownum, column=4).fill = VERDICT_FILL[verdict]
    ws.cell(row=rownum, column=4).font = Font(size=9, bold=True)
    ws.row_dimensions[rownum].height = 42
    rownum += 1

# ============================================================ Sheet 3: DR review
ws = wb.create_sheet("Existing Dataroom Review")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"
headers = ["Folder (in existing room)", "File name", "Verdict for THIS request",
           "Category", "Rationale", "Sensitivity"]
widths = [46, 40, 20, 24, 64, 12]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, 1, len(headers))
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(dr_class)+1}"

rownum = 2
for (root, rel, name, c) in dr_class:
    verdict, item, cat, rationale, sens, conf = c
    folder = os.path.dirname(rel) or "(root)"
    label = {"INCLUDE": "Keep (in scope)",
             "EXCLUDE": "Out of scope for this request",
             "REVIEW": "Review"}[verdict]
    vals = [folder, name, label, cat, rationale, sens]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=rownum, column=i, value=v)
        cell.font = Font(size=9)
        cell.border = BORDER
        cell.alignment = WRAP
    ws.cell(row=rownum, column=3).fill = VERDICT_FILL[verdict]
    ws.cell(row=rownum, column=3).font = Font(size=9, bold=True)
    ws.row_dimensions[rownum].height = 42
    rownum += 1

wb.save(OUT)
print("WROTE", OUT)
print("source:", st, "total", sum(st.values()))
print("dataroom:", dt, "total", sum(dt.values()))
