#!/usr/bin/env python3
"""acos-dataroom-v2 scoped run — main-thread text extraction for the 211 scoped files."""
import os, json, subprocess

RUN_DIR = open("/Users/zee/Documents/Vibe Coding/ACOS 3.0/.thurston_v2_run_dir").read().strip()
manifest = json.load(open(os.path.join(RUN_DIR, "intermediate", "scoped_manifest.json")))
EXTRACT_DIR = os.path.join(RUN_DIR, "extraction")
CHAR_CAP = 9000

def run(cmd, timeout=60):
    try:
        r = subprocess.run(cmd, capture_output=True, timeout=timeout)
        return r.stdout.decode("utf-8", "ignore")
    except Exception as e:
        return ""

index = {}
for rec in manifest:
    fid, path, name = rec["file_id"], rec["abspath"], rec["name"]
    ext = os.path.splitext(name)[1].lower()
    text, method, image_only = "", "none", False
    if not os.path.exists(path):
        method = "missing"
    elif ext == ".pdf":
        text = run(["pdftotext", "-f", "1", "-l", "8", "-layout", path, "-"])
        method = "pdftotext"
        if len(text.strip()) < 120:
            image_only = True
            method = "pdftotext(image-only)"
    elif ext in (".docx", ".doc", ".rtf"):
        text = run(["textutil", "-convert", "txt", "-stdout", path])
        method = "textutil"
    elif ext in (".txt", ".md"):
        try:
            text = open(path, encoding="utf-8", errors="ignore").read()
        except Exception:
            text = ""
        method = "read"
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets[:3]:
                parts.append(f"[sheet: {ws.title}]")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i > 40:
                        break
                    parts.append(" | ".join("" if c is None else str(c) for c in row))
            text = "\n".join(parts)
        except Exception:
            text = ""
        method = "openpyxl"
    elif ext in (".pptx",):
        method = "pptx(skipped)"
        text = f"[PowerPoint file: {name} — not text-extracted]"
    else:
        method = f"unsupported({ext})"

    text = text.strip()[:CHAR_CAP]
    header = (f"FILE: {name}\nLOAN: {rec['loan']}\nFOLDER: {rec['folder']}\n"
              f"EXTRACTION METHOD: {method}\n"
              f"{'NOTE: little/no extractable text — likely scanned/image-only PDF.' if image_only else ''}\n"
              f"{'-'*60}\n")
    with open(os.path.join(EXTRACT_DIR, fid + ".txt"), "w", encoding="utf-8") as f:
        f.write(header + (text if text else "[no extractable text]"))
    index[fid] = {"name": name, "method": method, "chars": len(text), "image_only": image_only}

with open(os.path.join(RUN_DIR, "intermediate", "extraction_index.json"), "w") as f:
    json.dump(index, f, indent=2)

img = sum(1 for v in index.values() if v["image_only"])
miss = sum(1 for v in index.values() if v["method"] == "missing")
print(f"extracted {len(index)} files | image-only/low-text: {img} | missing: {miss}")
