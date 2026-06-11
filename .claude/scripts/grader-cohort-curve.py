#!/usr/bin/env python3
"""
grader-cohort-curve.py — Load acos-grader per-paper result artifacts, apply
z-score curve, build final XLSX workbook.

Input: a session directory produced by acos-grader.
Output: an XLSX workbook with per-paper grade sheets, cohort summary, audit log.

Usage:
    python3 grader-cohort-curve.py --session-dir .acos/state/grader-sessions/grader-20261018T143000 --output grades.xlsx
"""

from __future__ import annotations

import argparse
import math
import re
import statistics
import sys
from pathlib import Path
from typing import Any


# ---------- Tiny YAML loader (stdlib only) ---------------------------------
# We avoid a PyYAML dependency by calling into python3 -c "import yaml; ..."
# if PyYAML is present, falling back to a minimal parser otherwise.


def _load_yaml(path: Path) -> dict[str, Any]:
    try:
        import yaml  # type: ignore
        return yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except ImportError:
        sys.exit(
            "ERROR: PyYAML not installed. Run: pip install pyyaml\n"
            "(Required for grader-cohort-curve.py — session artifacts are YAML.)"
        )


# ---------- Data collection -------------------------------------------------


def collect_papers(session_dir: Path) -> tuple[dict, list[dict], list[dict]]:
    """Return (manifest, paper_results, audit_entries).

    Per-file YAML loads are wrapped in try/except so one malformed or empty
    artifact doesn't kill the entire cohort build. Skipped files are logged
    to stderr and recorded in the manifest under a 'load_failures' key the
    caller can surface in the cohort summary.
    """
    manifest = _load_yaml(session_dir / "manifest.yaml")
    results_dir = session_dir / "results"
    audit_dir = session_dir / "audit"

    load_failures: list[dict] = []

    def _safe_load(path: Path) -> dict | None:
        if path.stat().st_size == 0:
            load_failures.append({"path": str(path), "reason": "empty file"})
            print(f"WARN: empty artifact skipped: {path}", file=sys.stderr)
            return None
        try:
            import yaml
            data = yaml.safe_load(path.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                load_failures.append({"path": str(path), "reason": f"expected dict, got {type(data).__name__}"})
                print(f"WARN: malformed artifact skipped (not a dict): {path}", file=sys.stderr)
                return None
            return data
        except Exception as e:
            load_failures.append({"path": str(path), "reason": f"{type(e).__name__}: {e}"})
            print(f"WARN: YAML load failed for {path}: {e}", file=sys.stderr)
            return None

    paper_results = []
    for result_file in sorted(results_dir.glob("*.yaml")):
        data = _safe_load(result_file)
        if data is not None:
            paper_results.append(data)

    audit_entries = []
    for audit_file in sorted(audit_dir.glob("*-audit.yaml")):
        data = _safe_load(audit_file)
        if data is not None:
            audit_entries.append(data)

    manifest.setdefault("load_failures", []).extend(load_failures)
    return manifest, paper_results, audit_entries


def compute_raw_totals(paper_results: list[dict]) -> list[dict]:
    """Annotate each paper with raw_total and disputed_count."""
    for paper in paper_results:
        criteria = paper.get("criteria", {}) or {}
        disputed = paper.get("disputed_ids", []) or []

        # criteria may be a dict keyed by criterion_id or a list
        if isinstance(criteria, dict):
            values = criteria.values()
        else:
            values = criteria

        raw_total = 0.0
        for c in values:
            if isinstance(c, dict):
                # Accept both legacy "points" and grader-paper's "points_awarded"
                pts = c.get("points_awarded", c.get("points", 0))
                raw_total += float(pts or 0)
        paper["raw_total"] = round(raw_total, 2)
        paper["disputed_count"] = len(disputed)

    return paper_results


def cohort_stats(paper_results: list[dict], field: str = "raw_total") -> dict[str, float]:
    values = [p[field] for p in paper_results if field in p]
    if not values:
        return {"n": 0, "mean": 0.0, "median": 0.0, "min": 0.0, "max": 0.0, "stdev": 0.0}
    mean = statistics.fmean(values)
    return {
        "n": len(values),
        "mean": mean,
        "median": statistics.median(values),
        "min": min(values),
        "max": max(values),
        "stdev": statistics.pstdev(values) if len(values) > 1 else 0.0,
    }


def apply_z_curve(
    paper_results: list[dict],
    raw_mean: float,
    raw_stdev: float,
    target_mean: float,
    target_stdev: float,
    floor: float,
    ceiling: float,
) -> None:
    """In-place: set paper['curved_total'] for every paper."""
    for paper in paper_results:
        raw = paper.get("raw_total", 0.0)
        if raw_stdev > 0:
            z = (raw - raw_mean) / raw_stdev
            curved = target_mean + z * target_stdev
        else:
            curved = target_mean
        curved = max(floor, min(ceiling, curved))
        paper["curved_total"] = round(curved, 2)


# ---------- Distribution table ---------------------------------------------


def distribution(values: list[float], buckets: list[tuple[float, float, str]]) -> list[tuple[str, int]]:
    out = []
    for lo, hi, label in buckets:
        count = sum(1 for v in values if lo <= v < hi)
        out.append((label, count))
    return out


def make_buckets(floor: float, ceiling: float, n_bins: int = 7) -> list[tuple[float, float, str]]:
    """Generate evenly-spaced distribution buckets across the actual curve range
    [floor, ceiling] taken from the manifest, instead of a hardcoded 0–100 grade
    scale (which produced a wrong distribution table for non-100 rubrics).

    Bins are half-open [lo, hi) except the final bin, whose upper edge is nudged
    just past the ceiling so a grade exactly equal to the ceiling is counted
    rather than dropped.
    """
    if ceiling <= floor:
        # Degenerate range — single catch-all bucket (defensive; main() already
        # aborts on a non-positive range, so this is belt-and-suspenders).
        label = f"{floor:g}+"
        return [(float(floor), float(ceiling) + 1.0, label)]

    width = (ceiling - floor) / n_bins
    buckets: list[tuple[float, float, str]] = []
    for i in range(n_bins):
        lo = floor + i * width
        hi = floor + (i + 1) * width
        if i == n_bins - 1:
            # Make the top bucket inclusive of the ceiling.
            hi = ceiling + 1e-9
            label = f"{lo:g}–{ceiling:g}"
        else:
            label = f"{lo:g}–{hi:g}"
        buckets.append((lo, hi, label))
    return buckets


# ---------- XLSX build ------------------------------------------------------


def build_xlsx(
    manifest: dict,
    paper_results: list[dict],
    audit_entries: list[dict],
    output_path: Path,
    raw: dict[str, float],
    curved: dict[str, float],
    target_mean: float,
    target_stdev: float,
    session_dir: Path | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()

    # Styling primitives
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    disputed_fill = PatternFill("solid", fgColor="FFF3CD")
    ungraded_fill = PatternFill("solid", fgColor="F8D7DA")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    # First sheet is Cohort Summary
    summary_ws = wb.active
    summary_ws.title = "Cohort Summary"

    summary_ws["A1"] = "acos-grader — Cohort Summary"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A2"] = f"Session: {manifest.get('session_id', '?')}"
    summary_ws["A3"] = f"Subject: {manifest.get('subject_subtype', '?')}"
    summary_ws["A4"] = f"Range: {manifest.get('total_floor', '?')}–{manifest.get('total_ceiling', '?')}"
    summary_ws["A5"] = f"Target curve: mean={target_mean}, stdev={target_stdev} (z-score rescale)"

    # Per-paper rows
    header_row = 7
    headers = ["Paper ID", "Raw Total", "Curved Grade", "Disputed Criteria", "Iterations"]
    for i, h in enumerate(headers):
        cell = summary_ws.cell(row=header_row, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, paper in enumerate(sorted(paper_results, key=lambda p: p.get("paper_id", ""))):
        r = header_row + 1 + idx
        summary_ws.cell(row=r, column=1, value=paper.get("paper_id", ""))
        summary_ws.cell(row=r, column=2, value=paper.get("raw_total", ""))
        summary_ws.cell(row=r, column=3, value=paper.get("curved_total", ""))
        summary_ws.cell(row=r, column=4, value=paper.get("disputed_count", 0))
        summary_ws.cell(row=r, column=5, value=paper.get("iterations_used", ""))

        if paper.get("disputed_count", 0) > 0:
            for c in range(1, 6):
                summary_ws.cell(row=r, column=c).fill = disputed_fill

    # Cohort stats block
    stats_start = header_row + 2 + len(paper_results)
    summary_ws.cell(row=stats_start, column=1, value="Cohort stats").font = Font(bold=True)
    for i, (label, val_raw, val_curved) in enumerate([
        ("n",       raw.get("n"),       curved.get("n")),
        ("mean",    raw.get("mean"),    curved.get("mean")),
        ("median",  raw.get("median"),  curved.get("median")),
        ("min",     raw.get("min"),     curved.get("min")),
        ("max",     raw.get("max"),     curved.get("max")),
        ("stdev",   raw.get("stdev"),   curved.get("stdev")),
    ]):
        summary_ws.cell(row=stats_start + 1 + i, column=1, value=label)
        summary_ws.cell(row=stats_start + 1 + i, column=2, value=round(val_raw, 2) if isinstance(val_raw, float) else val_raw)
        summary_ws.cell(row=stats_start + 1 + i, column=3, value=round(val_curved, 2) if isinstance(val_curved, float) else val_curved)
    summary_ws.cell(row=stats_start, column=2, value="Raw").font = Font(bold=True)
    summary_ws.cell(row=stats_start, column=3, value="Curved").font = Font(bold=True)

    # Distribution table (curved grades)
    dist_start = stats_start + 9
    summary_ws.cell(row=dist_start, column=1, value="Distribution (curved)").font = Font(bold=True)
    curved_values = [p["curved_total"] for p in paper_results if "curved_total" in p]
    dist_floor = float(manifest.get("total_floor", 0))
    dist_ceiling = float(manifest.get("total_ceiling", 100))
    dist = distribution(curved_values, make_buckets(dist_floor, dist_ceiling))
    for i, (label, count) in enumerate(dist):
        summary_ws.cell(row=dist_start + 1 + i, column=1, value=label)
        summary_ws.cell(row=dist_start + 1 + i, column=2, value=count)

    # Column widths
    for col, width in enumerate([20, 14, 14, 18, 14], start=1):
        summary_ws.column_dimensions[get_column_letter(col)].width = width

    # Per-paper sheets — dedup names to prevent openpyxl collision errors
    used_sheet_names: set[str] = {"Cohort Summary"}
    for paper in paper_results:
        paper_id = str(paper.get("paper_id", ""))
        sheet_name = _sanitize_sheet_name(paper_id, used_sheet_names)
        used_sheet_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)

        ws["A1"] = f"Paper: {paper_id}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Raw total: {paper.get('raw_total', 0)}"
        ws["A3"] = f"Curved grade: {paper.get('curved_total', 0)}"
        ws["A4"] = f"Iterations: {paper.get('iterations_used', 0)}"
        if paper.get("disputed_count", 0) > 0:
            ws["A5"] = f"DISPUTED criteria: {paper.get('disputed_count')}"
            ws["A5"].font = Font(bold=True, color="9A6700")

        # Grading sheet
        tbl_start = 7
        for i, h in enumerate(["Criterion", "Points Awarded", "Points Total", "Reasoning", "Status"]):
            cell = ws.cell(row=tbl_start, column=i + 1, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        criteria = paper.get("criteria", {}) or {}
        if isinstance(criteria, dict):
            rows_iter = list(criteria.items())
        else:
            rows_iter = [(c.get("id", f"crit_{i}"), c) for i, c in enumerate(criteria)]

        disputed_ids = set(paper.get("disputed_ids", []) or [])
        for i, (cid, cdata) in enumerate(rows_iter):
            if not isinstance(cdata, dict):
                continue
            r = tbl_start + 1 + i
            ws.cell(row=r, column=1, value=cdata.get("name", cid))
            ws.cell(row=r, column=2, value=cdata.get("points_awarded", cdata.get("points", "")))
            ws.cell(row=r, column=3, value=cdata.get("points_total", ""))
            reasoning_cell = ws.cell(row=r, column=4, value=cdata.get("final_reasoning", cdata.get("reasoning", "")))
            reasoning_cell.alignment = wrap
            status = "DISPUTED (max_iters)" if cid in disputed_ids else "CONVERGED"
            ws.cell(row=r, column=5, value=status)
            if cid in disputed_ids:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = disputed_fill

        # Column widths
        widths = [32, 14, 12, 80, 22]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    # Audit log sheet
    audit_ws = wb.create_sheet("Audit Log")
    audit_headers = ["Paper ID", "Criterion", "Iteration", "Grader A Points", "Grader B Points", "Grader C Points", "QA Verdict"]
    for i, h in enumerate(audit_headers):
        cell = audit_ws.cell(row=1, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for entry in audit_entries:
        paper_id = entry.get("paper_id", "")
        history = entry.get("iteration_history", []) or []
        for it in history:
            iter_num = it.get("iteration", "")
            for (cid, a_pts, b_pts, c_pts, verdict) in _normalize_iteration_audit(it, session_dir, paper_id):
                audit_ws.cell(row=row, column=1, value=paper_id)
                audit_ws.cell(row=row, column=2, value=cid)
                audit_ws.cell(row=row, column=3, value=iter_num)
                audit_ws.cell(row=row, column=4, value=a_pts)
                audit_ws.cell(row=row, column=5, value=b_pts)
                audit_ws.cell(row=row, column=6, value=c_pts)
                audit_ws.cell(row=row, column=7, value=verdict)
                row += 1

    for col, width in enumerate([20, 24, 12, 16, 16, 16, 16], start=1):
        audit_ws.column_dimensions[get_column_letter(col)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _sanitize_sheet_name(name: str, used: set[str] | None = None) -> str:
    """Sanitize to XLSX sheet-name rules (max 31 chars, no [ ] : * ? / \\)
    and ensure uniqueness within the workbook by appending a short suffix
    when the truncated-and-sanitized form collides with an already-used name.
    """
    sanitized = re.sub(r"[\[\]:*?/\\]", "_", name)[:31] or "Paper"
    if used is None or sanitized not in used:
        return sanitized
    # Collision — append _2, _3, ... until unique, shrinking the base to fit
    for suffix_n in range(2, 1000):
        suffix = f"_{suffix_n}"
        base_len = 31 - len(suffix)
        candidate = sanitized[:base_len] + suffix
        if candidate not in used:
            return candidate
    # Extreme fallback — should be unreachable with <1000 papers
    return f"Paper_{len(used)}"


def _load_grader_sheets_from_disk(session_dir: Path, paper_id: str, iter_num: int) -> list[tuple] | None:
    """Canonical disk-based audit loader.

    The grader-paper agent always writes three individual grader sheets and
    one QA verdict file at predictable paths under
    ``grading/<paper_id>/iter-<N>-*.yaml``. Those files have a consistent
    schema (unlike the audit YAML, which each agent invents fresh). When
    inline audit shapes fail to parse, fall back to reading these files.
    """
    try:
        import yaml
    except ImportError:
        return None

    grading_dir = session_dir / "grading" / paper_id
    if not grading_dir.is_dir():
        return None

    def _load(name: str) -> dict:
        p = grading_dir / name
        if not p.exists():
            return {}
        try:
            return yaml.safe_load(p.read_text(encoding="utf-8")) or {}
        except Exception:
            return {}

    # Filenames may use either "grader-opus-A" or "grader-opus-a"; check both.
    def _load_any(*names: str) -> dict:
        for n in names:
            d = _load(n)
            if d:
                return d
        return {}

    a_sheet = _load_any(f"iter-{iter_num}-grader-opus-A.yaml", f"iter-{iter_num}-grader-opus-a.yaml")
    b_sheet = _load_any(f"iter-{iter_num}-grader-opus-B.yaml", f"iter-{iter_num}-grader-opus-b.yaml")
    c_sheet = _load_any(f"iter-{iter_num}-grader-sonnet.yaml")
    qa_sheet = _load_any(f"iter-{iter_num}-qa-verdict.yaml")

    if not (a_sheet or b_sheet or c_sheet):
        return None

    def _sheet_points_by_cid(sheet: dict) -> dict:
        if not isinstance(sheet, dict):
            return {}
        out = {}
        crits = sheet.get("criteria", []) or []
        if isinstance(crits, dict):
            for cid, c in crits.items():
                if isinstance(c, dict):
                    out[cid] = c.get("points_awarded", c.get("points", ""))
        else:
            for c in crits:
                if isinstance(c, dict):
                    cid = c.get("id") or c.get("criterion_id")
                    if cid:
                        out[cid] = c.get("points_awarded", c.get("points", ""))
        return out

    a_pts = _sheet_points_by_cid(a_sheet)
    b_pts = _sheet_points_by_cid(b_sheet)
    c_pts = _sheet_points_by_cid(c_sheet)

    # QA verdict — probe a few common shapes
    qa_map: dict[str, str] = {}
    if isinstance(qa_sheet, dict):
        for field in ("verdicts", "per_criterion", "per_criterion_detail"):
            payload = qa_sheet.get(field)
            if isinstance(payload, list):
                for v in payload:
                    if isinstance(v, dict):
                        cid = v.get("criterion_id") or v.get("id")
                        verdict = v.get("final_verdict") or v.get("verdict")
                        if cid and verdict:
                            qa_map[cid] = verdict
            elif isinstance(payload, dict):
                for cid, v in payload.items():
                    if isinstance(v, dict):
                        qa_map[cid] = v.get("final_verdict") or v.get("verdict") or ""
                    else:
                        qa_map[cid] = v
            if qa_map:
                break
        if not qa_map:
            passed_ids = qa_sheet.get("passed_ids") or []
            failed_ids = qa_sheet.get("failed_ids") or []
            for cid in passed_ids:
                qa_map[cid] = "PASS"
            for cid in failed_ids:
                qa_map[cid] = "FAIL"

    crit_ids = set(a_pts) | set(b_pts) | set(c_pts) | set(qa_map)
    if not crit_ids:
        return None
    return [(cid, a_pts.get(cid, ""), b_pts.get(cid, ""), c_pts.get(cid, ""), qa_map.get(cid, ""))
            for cid in sorted(crit_ids)]


def _load_qa_verdicts_from_disk(session_dir: Path, paper_id: str, iter_num: int) -> dict[str, str]:
    """Load the canonical QA verdict YAML and return {criterion_id: verdict}.

    Probes several field shapes (verdicts list, per_criterion list/dict,
    passed_ids/failed_ids lists) because the qa-verdict.yaml schema also
    varies across agent instances.
    """
    try:
        import yaml
    except ImportError:
        return {}
    p = session_dir / "grading" / paper_id / f"iter-{iter_num}-qa-verdict.yaml"
    if not p.exists():
        return {}
    try:
        data = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    out: dict[str, str] = {}
    # passed_ids / failed_ids lists (most common explicit shape)
    for cid in (data.get("passed_ids") or []):
        out[cid] = "PASS"
    for cid in (data.get("failed_ids") or []):
        out[cid] = "FAIL"
    # verdicts field — may be list of objects OR dict keyed by criterion_id
    verdicts = data.get("verdicts")
    if isinstance(verdicts, list):
        for v in verdicts:
            if isinstance(v, dict):
                cid = v.get("criterion_id") or v.get("id")
                verdict = v.get("final_verdict") or v.get("verdict")
                if cid and verdict and cid not in out:
                    out[cid] = verdict
    elif isinstance(verdicts, dict):
        for cid, v in verdicts.items():
            if cid in out:
                continue
            if isinstance(v, dict):
                verdict = v.get("final_verdict") or v.get("verdict")
                if verdict:
                    out[cid] = verdict
            elif isinstance(v, str):
                out[cid] = v
    # per_criterion / per_criterion_detail (list or dict)
    for field in ("per_criterion", "per_criterion_detail"):
        payload = data.get(field)
        if isinstance(payload, list):
            for v in payload:
                if isinstance(v, dict):
                    cid = v.get("criterion_id") or v.get("id")
                    verdict = v.get("final_verdict") or v.get("verdict")
                    if cid and verdict and cid not in out:
                        out[cid] = verdict
        elif isinstance(payload, dict):
            for cid, v in payload.items():
                if cid in out:
                    continue
                if isinstance(v, dict):
                    verdict = v.get("final_verdict") or v.get("verdict")
                    if verdict:
                        out[cid] = verdict
                elif isinstance(v, str):
                    out[cid] = v
    return out


def _backfill_verdicts(rows: list[tuple], session_dir: Path | None, paper_id: str | None, iter_num: int) -> list[tuple]:
    """If any row has an empty verdict, backfill from the canonical on-disk
    qa-verdict.yaml for that iteration. Leaves non-empty verdicts untouched."""
    if not rows:
        return rows
    if session_dir is None or not paper_id:
        return rows
    if all(v for (_, _, _, _, v) in rows):
        return rows
    try:
        iter_num_int = int(iter_num)
    except (TypeError, ValueError):
        iter_num_int = 1
    qa = _load_qa_verdicts_from_disk(session_dir, paper_id, iter_num_int)
    if not qa:
        return rows
    return [(cid, a, b, c, v or qa.get(cid, "")) for (cid, a, b, c, v) in rows]


def _normalize_iteration_audit(iteration: dict, session_dir: Path | None = None, paper_id: str | None = None) -> list[tuple]:
    """Normalize one iteration's audit data into a list of
    (criterion_id, grader_a_pts, grader_b_pts, grader_c_pts, verdict) tuples.

    The grader-paper agent's audit YAML schema drifts between agent instances
    (15+ observed variants). We try each known inline shape, and fall back to
    loading the canonical grader sheet files from disk when nothing matches.
    Any missing QA verdicts are backfilled from the canonical
    ``iter-N-qa-verdict.yaml`` file when session_dir+paper_id are available.
    """
    iter_num = iteration.get("iteration") or 1
    # Audit schemas drift across 15+ agent variants — an iteration label may be
    # a non-numeric string (e.g. 'final', '1a', 'I'). Coerce defensively so the
    # whole cohort build doesn't crash on a single odd label. Original `iter_num`
    # is retained for display; `iter_num_int` is used for on-disk file lookups.
    try:
        iter_num_int = int(iter_num)
    except (TypeError, ValueError):
        iter_num_int = 1

    # Shape 1: list of summary objects (per_criterion_summary / per_criterion /
    # per_criterion_grades / grades_by_criterion / grader_scores / grades /
    # grader_raw_points — various synonyms).
    for field in ("per_criterion_summary", "per_criterion", "per_criterion_detail",
                  "per_criterion_grades", "grades_by_criterion", "grader_scores",
                  "grader_raw_points", "grades", "criterion_summaries"):
        summary = iteration.get(field)
        if isinstance(summary, list) and summary:
            out = []
            for cs in summary:
                if not isinstance(cs, dict):
                    continue
                vals = cs.get("values") or cs.get("points") or []
                a = vals[0] if len(vals) > 0 else cs.get("A", cs.get("grader_a", cs.get("grader-opus-A", "")))
                b = vals[1] if len(vals) > 1 else cs.get("B", cs.get("grader_b", cs.get("grader-opus-B", "")))
                c = vals[2] if len(vals) > 2 else cs.get("C", cs.get("grader_c", cs.get("grader-sonnet", "")))
                verdict = cs.get("verdict") or cs.get("final_verdict") or cs.get("qa_verdict") or ""
                if isinstance(verdict, dict):
                    verdict = verdict.get("verdict") or verdict.get("final_verdict") or ""
                out.append((cs.get("criterion_id") or cs.get("id", ""), a, b, c, verdict))
            if out:
                return _backfill_verdicts(out, session_dir, paper_id, iter_num_int)
        if isinstance(summary, dict) and summary:
            # dict keyed by criterion_id → {A, B, C} or values list
            out = []
            for cid, cdata in summary.items():
                if not isinstance(cdata, dict):
                    continue
                vals = cdata.get("values") or cdata.get("points") or []
                a = vals[0] if len(vals) > 0 else cdata.get("A", cdata.get("grader_a", cdata.get("grader-opus-A", "")))
                b = vals[1] if len(vals) > 1 else cdata.get("B", cdata.get("grader_b", cdata.get("grader-opus-B", "")))
                c = vals[2] if len(vals) > 2 else cdata.get("C", cdata.get("grader_c", cdata.get("grader-sonnet", "")))
                verdict = cdata.get("verdict") or cdata.get("final_verdict") or ""
                if isinstance(verdict, dict):
                    verdict = verdict.get("verdict") or verdict.get("final_verdict") or ""
                out.append((cid, a, b, c, verdict))
            if out:
                return _backfill_verdicts(out, session_dir, paper_id, iter_num_int)

    # Shape 2: separate grader_outputs / grader_sheets (dict of role → sheet) and
    # qa_verdicts / qa_verdict dict of crit_id → verdict
    for go_field in ("grader_outputs", "grader_sheets", "graders", "grader_grades"):
        grader_outputs = iteration.get(go_field)
        if isinstance(grader_outputs, dict) and grader_outputs:
            roles = {k.lower().replace("-", "_"): k for k in grader_outputs.keys()}

            def role_lookup(*candidates: str):
                for c_ in candidates:
                    if c_ in roles:
                        return grader_outputs[roles[c_]]
                return {}

            a_sheet = role_lookup("grader_opus_a", "grader_a", "opus_a", "a")
            b_sheet = role_lookup("grader_opus_b", "grader_b", "opus_b", "b")
            c_sheet = role_lookup("grader_sonnet", "grader_c", "sonnet", "c")
            qa_src = iteration.get("qa_verdicts") or iteration.get("qa_verdict") or iteration.get("qa_summary") or iteration.get("qa_verdicts_summary") or {}

            def _flatten(sheet) -> dict:
                if not isinstance(sheet, dict):
                    return {}
                out = {}
                # Direct `criteria` list or dict
                if "criteria" in sheet:
                    crits = sheet["criteria"] or []
                    if isinstance(crits, list):
                        for c_ in crits:
                            if isinstance(c_, dict):
                                cid = c_.get("id") or c_.get("criterion_id")
                                if cid:
                                    out[cid] = c_.get("points_awarded", c_.get("points", ""))
                    elif isinstance(crits, dict):
                        for cid, c_ in crits.items():
                            if isinstance(c_, dict):
                                out[cid] = c_.get("points_awarded", c_.get("points", ""))
                    return out
                # Nested `per_criterion` dict (common when grader_sheets carries
                # both a path reference and inline points)
                for nested_key in ("per_criterion", "per_criterion_grades", "criterion_grades", "grades", "scores"):
                    nested = sheet.get(nested_key)
                    if isinstance(nested, dict) and nested:
                        for cid, v in nested.items():
                            if isinstance(v, dict):
                                out[cid] = v.get("points_awarded", v.get("points", ""))
                            else:
                                out[cid] = v
                        return out
                # Fallback — treat sheet itself as {cid: points}, skipping meta keys
                META_KEYS = {"total", "raw_total", "total_awarded", "path", "role",
                             "grader_role", "output_path", "iteration", "nonce",
                             "timestamp", "paper_id"}
                for k, v in sheet.items():
                    if k in META_KEYS:
                        continue
                    if isinstance(v, dict):
                        # Only treat nested dict as a criterion if it looks like
                        # a points record
                        pv = v.get("points_awarded", v.get("points", None))
                        if pv is not None:
                            out[k] = pv
                    elif isinstance(v, (int, float)):
                        out[k] = v
                return out

            a_pts = _flatten(a_sheet)
            b_pts = _flatten(b_sheet)
            c_pts = _flatten(c_sheet)
            crit_ids = set(a_pts) | set(b_pts) | set(c_pts)
            if crit_ids:
                # Build verdict map
                qa_map: dict[str, str] = {}
                QA_META = {"output_path", "path", "passed", "failed", "passed_ids",
                           "failed_ids", "iteration", "per_criterion_detail",
                           "verdicts", "summary"}
                if isinstance(qa_src, dict):
                    # Prefer explicit passed_ids/failed_ids lists (canonical summary shape)
                    for cid in (qa_src.get("passed_ids") or []):
                        qa_map[cid] = "PASS"
                    for cid in (qa_src.get("failed_ids") or []):
                        qa_map[cid] = "FAIL"
                    # Fall through: fill any missing verdicts from crit-keyed entries
                    for cid, v in qa_src.items():
                        if cid in QA_META or cid in qa_map:
                            continue
                        if isinstance(v, dict):
                            vv = v.get("verdict") or v.get("final_verdict") or ""
                            if vv:
                                qa_map[cid] = vv
                        elif isinstance(v, str):
                            qa_map[cid] = v
                rows = [(cid, a_pts.get(cid, ""), b_pts.get(cid, ""), c_pts.get(cid, ""), qa_map.get(cid, ""))
                        for cid in sorted(crit_ids)]
                return _backfill_verdicts(rows, session_dir, paper_id, iter_num_int)

    # Shape 3 (fallback): load canonical grader sheet files from disk.
    if session_dir is not None and paper_id:
        disk = _load_grader_sheets_from_disk(session_dir, paper_id, iter_num_int)
        if disk:
            return _backfill_verdicts(disk, session_dir, paper_id, iter_num_int)

    return []


def build_student_xlsx(
    paper: dict,
    audit_entry: dict | None,
    manifest: dict,
    raw: dict[str, float],
    curved: dict[str, float],
    target_mean: float,
    target_stdev: float,
    output_path: Path,
    session_dir: Path | None = None,
) -> None:
    """Write a single-student grade report XLSX: header + criterion table +
    mini audit table (per-criterion grader points and QA verdicts).

    Designed for handback to the student — contains only that student's
    grades, their criterion-by-criterion reasoning, and the audit trail that
    justifies each score.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    disputed_fill = PatternFill("solid", fgColor="FFF3CD")
    section_font = Font(bold=True, size=12, color="2C3E50")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Report"

    paper_id = str(paper.get("paper_id", ""))

    # --- Header block --------------------------------------------------------
    ws["A1"] = f"Grade Report — {paper_id}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Session: {manifest.get('session_id', '?')}"
    ws["A3"] = f"Subject: {manifest.get('subject_subtype', '?')}"
    ws["A4"] = f"Graded range: {manifest.get('total_floor', '?')}–{manifest.get('total_ceiling', '?')}"

    ws["A6"] = "Raw total"
    ws["B6"] = paper.get("raw_total", 0)
    ws["A6"].font = Font(bold=True)
    ws["A7"] = "Curved grade"
    ws["B7"] = paper.get("curved_total", 0)
    ws["A7"].font = Font(bold=True)
    ws["B7"].font = Font(bold=True, size=12, color="2C3E50")
    ws["A8"] = "Iterations used"
    ws["B8"] = paper.get("iterations_used", 1)
    ws["A9"] = "Disputed criteria"
    ws["B9"] = paper.get("disputed_count", 0)

    ws["A11"] = f"Cohort context: n={raw.get('n', 0)}, raw mean={raw.get('mean', 0):.2f}, curved mean={curved.get('mean', 0):.2f}"
    ws["A11"].font = Font(italic=True, size=10, color="555555")

    # --- Criterion table -----------------------------------------------------
    ws["A13"] = "Per-criterion grades"
    ws["A13"].font = section_font

    tbl_start = 14
    headers = ["Criterion", "Points Awarded", "Points Total", "Reasoning", "Status"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=tbl_start, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    criteria = paper.get("criteria", {}) or {}
    if isinstance(criteria, dict):
        rows_iter = list(criteria.items())
    else:
        rows_iter = [(c.get("id", f"crit_{i}"), c) for i, c in enumerate(criteria)]

    disputed_ids = set(paper.get("disputed_ids", []) or [])
    last_row = tbl_start
    for i, (cid, cdata) in enumerate(rows_iter):
        if not isinstance(cdata, dict):
            continue
        r = tbl_start + 1 + i
        ws.cell(row=r, column=1, value=cdata.get("name", cid))
        ws.cell(row=r, column=2, value=cdata.get("points_awarded", cdata.get("points", "")))
        ws.cell(row=r, column=3, value=cdata.get("points_total", ""))
        reasoning_cell = ws.cell(row=r, column=4, value=cdata.get("final_reasoning", cdata.get("reasoning", "")))
        reasoning_cell.alignment = wrap
        status = "DISPUTED (max_iters)" if cid in disputed_ids else "CONVERGED"
        ws.cell(row=r, column=5, value=status)
        if cid in disputed_ids:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = disputed_fill
        last_row = r

    # Tall rows for wrapped reasoning
    for r in range(tbl_start + 1, last_row + 1):
        ws.row_dimensions[r].height = 110

    # --- Mini audit table ----------------------------------------------------
    audit_start = last_row + 3
    ws.cell(row=audit_start, column=1, value="Grading audit trail").font = section_font
    ws.cell(row=audit_start + 1, column=1, value=(
        "Each criterion was graded independently by three finance-expert agents "
        "(two Opus, one Sonnet). QA gated consensus on both points (±5% spread) "
        "and reasoning (≥90% similarity). Failed criteria were re-dispatched blind."
    )).font = Font(italic=True, size=9, color="555555")
    ws.cell(row=audit_start + 1, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=audit_start + 1, start_column=1, end_row=audit_start + 1, end_column=5)
    ws.row_dimensions[audit_start + 1].height = 45

    audit_hdr_row = audit_start + 3
    audit_headers = ["Criterion", "Iteration", "Grader A", "Grader B", "Grader C", "QA Verdict"]
    for i, h in enumerate(audit_headers):
        cell = ws.cell(row=audit_hdr_row, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    audit_row = audit_hdr_row + 1
    paper_id_for_audit = str(paper.get("paper_id", ""))
    history = (audit_entry or {}).get("iteration_history", []) or [{"iteration": 1}]
    for it in history:
        iter_num = it.get("iteration", 1)
        for (cid, a_pts, b_pts, c_pts, verdict) in _normalize_iteration_audit(it, session_dir, paper_id_for_audit):
            ws.cell(row=audit_row, column=1, value=cid)
            ws.cell(row=audit_row, column=2, value=iter_num)
            ws.cell(row=audit_row, column=3, value=a_pts)
            ws.cell(row=audit_row, column=4, value=b_pts)
            ws.cell(row=audit_row, column=5, value=c_pts)
            ws.cell(row=audit_row, column=6, value=verdict)
            audit_row += 1

    # Column widths
    widths = [34, 14, 12, 80, 22]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    # Audit cols 2..6 narrower (overrides first-block widths partially;
    # openpyxl uses the last-set width per column, so re-set here)
    for col, width in [(2, 11), (3, 11), (4, 11), (5, 11), (6, 18)]:
        ws.column_dimensions[get_column_letter(col)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _all_criteria_in_iteration(sheets: dict) -> list[str]:
    crits: set[str] = set()
    for s in sheets.values():
        if not isinstance(s, dict):
            continue
        cs = s.get("criteria", []) or []
        for c in cs:
            if isinstance(c, dict) and "id" in c:
                crits.add(c["id"])
    return sorted(crits)


def _lookup_points(sheet: dict, criterion_id: str) -> Any:
    if not isinstance(sheet, dict):
        return ""
    for c in (sheet.get("criteria", []) or []):
        if isinstance(c, dict) and c.get("id") == criterion_id:
            return c.get("points_awarded", "")
    return ""


# ---------- Main -----------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--session-dir", required=True, help="Path to .acos/state/grader-sessions/<id>/")
    ap.add_argument("--output", required=True, help="Path to write cohort XLSX")
    ap.add_argument(
        "--per-student-dir",
        default=None,
        help=(
            "Optional directory for per-student XLSX files (one per paper). "
            "If unset, defaults to '<cohort-xlsx-parent>/per-student/'. "
            "Pass 'NONE' to disable per-student output entirely."
        ),
    )
    args = ap.parse_args()

    session_dir = Path(args.session_dir).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not session_dir.is_dir():
        sys.exit(f"ERROR: Session directory not found: {session_dir}")
    if not (session_dir / "manifest.yaml").exists():
        sys.exit(f"ERROR: No manifest.yaml in {session_dir}")

    manifest, paper_results, audit_entries = collect_papers(session_dir)
    if not paper_results:
        sys.exit(f"ERROR: No per-paper result artifacts found in {session_dir / 'results'}")

    compute_raw_totals(paper_results)

    raw = cohort_stats(paper_results, "raw_total")

    total_floor = float(manifest.get("total_floor", 0))
    total_ceiling = float(manifest.get("total_ceiling", 100))
    if total_floor >= total_ceiling:
        sys.exit(
            f"ERROR: manifest total_floor ({total_floor}) must be strictly less than "
            f"total_ceiling ({total_ceiling}). Aborting — a zero/negative range would "
            f"collapse every curved grade to the midpoint."
        )
    target_mean = (total_floor + total_ceiling) / 2
    target_stdev = (total_ceiling - total_floor) / 4

    apply_z_curve(
        paper_results,
        raw_mean=raw["mean"],
        raw_stdev=raw["stdev"],
        target_mean=target_mean,
        target_stdev=target_stdev,
        floor=total_floor,
        ceiling=total_ceiling,
    )

    curved = cohort_stats(paper_results, "curved_total")

    build_xlsx(
        manifest=manifest,
        paper_results=paper_results,
        audit_entries=audit_entries,
        output_path=output_path,
        raw=raw,
        curved=curved,
        target_mean=target_mean,
        target_stdev=target_stdev,
        session_dir=session_dir,
    )

    print(f"Wrote {output_path}")
    print(f"  Papers: {raw['n']}")
    print(f"  Raw mean:    {raw['mean']:.2f}   stdev: {raw['stdev']:.2f}")
    print(f"  Curved mean: {curved['mean']:.2f}   stdev: {curved['stdev']:.2f}")
    print(f"  Target: mean={target_mean}, stdev={target_stdev}")

    # --- Per-student XLSX files ---------------------------------------------
    if args.per_student_dir and args.per_student_dir.upper() == "NONE":
        per_student_dir = None
    elif args.per_student_dir:
        per_student_dir = Path(args.per_student_dir).expanduser().resolve()
    else:
        per_student_dir = output_path.parent / "per-student"

    if per_student_dir is not None:
        per_student_dir.mkdir(parents=True, exist_ok=True)
        audit_by_pid = {a.get("paper_id"): a for a in audit_entries if isinstance(a, dict)}
        written = 0
        for paper in paper_results:
            pid = str(paper.get("paper_id", "")).strip()
            if not pid:
                continue
            safe_pid = re.sub(r"[^A-Za-z0-9._-]", "_", pid)
            student_path = per_student_dir / f"{safe_pid}.xlsx"
            build_student_xlsx(
                paper=paper,
                audit_entry=audit_by_pid.get(pid),
                manifest=manifest,
                raw=raw,
                curved=curved,
                target_mean=target_mean,
                target_stdev=target_stdev,
                output_path=student_path,
                session_dir=session_dir,
            )
            written += 1
        print(f"Wrote {written} per-student file(s) to {per_student_dir}/")
    return 0


if __name__ == "__main__":
    sys.exit(main())
