#!/usr/bin/env python3
"""
ACOS Loan Document Generator — XLSX Cell-Level Extraction Utility

Extracts structured data from .xlsx files with cell-level provenance.
Outputs YAML suitable for LLM consumption in Phase 2 analyzer agents.

Usage:
  python3 xlsx-extract.py <file.xlsx> [--output <output.yaml>] [--sheet <name>]

Features:
  - Cell-level extraction with addresses (e.g., Sheet1!B7)
  - Formula detection and tracking (stores both formula and computed value)
  - Merged cell handling (attributes value to the merge anchor)
  - Named range resolution
  - Structured table detection (Excel ListObjects)
  - Hidden sheet detection (flagged but still extracted)
  - Number formatting preservation (currency, percentage, date)
  - Empty row/column trimming
"""

import sys
import os
import json
import re
from datetime import datetime, date, time
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def classify_value(value, number_format):
    """Classify a cell value for provenance metadata."""
    if value is None:
        return "empty", None
    if isinstance(value, bool):
        return "boolean", str(value).lower()
    if isinstance(value, (int, float)):
        fmt = (number_format or "").lower()
        if any(c in fmt for c in ["$", "currency", "#,##0"]):
            return "currency", format_number(value, number_format)
        if "%" in fmt:
            # Excel stores percentages as the underlying fraction (0.25 == 25%),
            # so always scale by 100 — uniformly, for any magnitude. The old
            # branch only multiplied for |value|<1 (so 1.5 -> "1.5%" instead of
            # "150%") and guarded on a dead `'%' not in str(value)` check.
            return "percentage", f"{value * 100:g}%"
        if any(d in fmt for d in ["yy", "mm", "dd"]):
            return "date", str(value)
        return "number", value
    if isinstance(value, datetime):
        return "datetime", value.isoformat()
    if isinstance(value, date):
        return "date", value.isoformat()
    if isinstance(value, time):
        return "time", value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return "empty", None
        return "text", value
    return "unknown", str(value)


def format_number(value, fmt):
    """Format a number using its Excel format string for human-readable output."""
    if fmt and "$" in fmt:
        if value >= 0:
            return f"${value:,.2f}"
        return f"-${abs(value):,.2f}"
    if fmt and "%" in fmt:
        # Excel stores percentages as fractions; scale by 100 uniformly for all
        # magnitudes. The old abs<10 vs >=10 split was contradictory: `:.2%`
        # multiplies by 100 while `:.1f}%` did not, so 0.25 -> "25.00%" but
        # 15 -> "15.0%" (should be "1500%").
        return f"{value * 100:g}%"
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def extract_sheet(ws, wb, sheet_name, file_path):
    """Extract all data from a single worksheet."""
    sheet_data = {
        "sheet_name": sheet_name,
        "dimensions": ws.dimensions,
        "is_hidden": ws.sheet_state != "visible",
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        "tables": [],
        "rows": [],
        "summary": {
            "total_cells_with_data": 0,
            "formula_cells": 0,
            "numeric_cells": 0,
            "currency_cells": 0,
            "percentage_cells": 0,
            "text_cells": 0,
        }
    }

    # Extract Excel tables (ListObjects)
    if hasattr(ws, 'tables'):
        for table_name, table in ws.tables.items():
            sheet_data["tables"].append({
                "name": table_name,
                "range": table.ref,
                "header_row": True,
                "columns": [col.name for col in table.tableColumns] if hasattr(table, 'tableColumns') else []
            })

    # Build merged cell map: for any cell in a merged range, point to the anchor
    merge_map = {}
    for merge_range in ws.merged_cells.ranges:
        anchor = merge_range.min_row, merge_range.min_col
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                if (row, col) != anchor:
                    merge_map[(row, col)] = f"{sheet_name}!{get_column_letter(anchor[1])}{anchor[0]}"

    # Extract rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
        row_data = {"row": row_idx, "cells": []}
        row_has_data = False

        for cell in row:
            if cell.value is None and (cell.row, cell.column) not in merge_map:
                continue

            col_letter = get_column_letter(cell.column)
            cell_ref = f"{sheet_name}!{col_letter}{cell.row}"
            value_type, display_value = classify_value(cell.value, cell.number_format)

            if value_type == "empty" and (cell.row, cell.column) in merge_map:
                continue  # Skip non-anchor cells of merged ranges

            if value_type == "empty":
                continue

            row_has_data = True
            cell_entry = {
                "ref": cell_ref,
                "col": col_letter,
                "value": cell.value if not isinstance(cell.value, (datetime, date, time)) else str(cell.value),
                "display": display_value,
                "type": value_type,
            }

            # Track formulas
            if cell.data_type == "f" or (hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith("=")):
                cell_entry["formula"] = str(cell.value)
                cell_entry["is_calculated"] = True
                sheet_data["summary"]["formula_cells"] += 1

            # Track number format
            if cell.number_format and cell.number_format != "General":
                cell_entry["format"] = cell.number_format

            # Track if in merged range
            if (cell.row, cell.column) in merge_map:
                cell_entry["merged_from"] = merge_map[(cell.row, cell.column)]

            # Provenance
            cell_entry["source"] = {
                "file": os.path.basename(file_path),
                "sheet": sheet_name,
                "cell": f"{col_letter}{cell.row}",
                "full_ref": cell_ref,
            }

            # Update summary
            sheet_data["summary"]["total_cells_with_data"] += 1
            if value_type == "currency":
                sheet_data["summary"]["currency_cells"] += 1
            elif value_type == "percentage":
                sheet_data["summary"]["percentage_cells"] += 1
            elif value_type == "number":
                sheet_data["summary"]["numeric_cells"] += 1
            elif value_type == "text":
                sheet_data["summary"]["text_cells"] += 1

            row_data["cells"].append(cell_entry)

        if row_has_data:
            sheet_data["rows"].append(row_data)

    return sheet_data


def extract_named_ranges(wb):
    """Extract all named ranges from the workbook."""
    named_ranges = []
    try:
        # openpyxl 3.1+: DefinedNameDict supports iteration over values
        names_iter = wb.defined_names.values() if hasattr(wb.defined_names, 'values') else []
        for name in names_iter:
            try:
                destinations = list(name.destinations)
                for sheet_title, cell_range in destinations:
                    named_ranges.append({
                        "name": name.name,
                        "sheet": sheet_title,
                        "range": cell_range,
                        "scope": "workbook" if getattr(name, 'localSheetId', None) is None else f"sheet:{sheet_title}",
                    })
            except Exception:
                named_ranges.append({
                    "name": name.name,
                    "value": str(getattr(name, 'value', '')),
                    "scope": "workbook",
                    "error": "Could not resolve destinations"
                })
    except Exception:
        pass  # No named ranges or incompatible API version
    return named_ranges


def extract_workbook(file_path, target_sheet=None):
    """Extract all data from an xlsx workbook."""
    file_path = os.path.abspath(file_path)

    # Pass 1: Load with data_only=False to capture formulas
    # Then close before Pass 2 to halve peak memory
    formula_overlay = {}
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    sheet_names = wb_formulas.sheetnames
    named_ranges = extract_named_ranges(wb_formulas)

    for sheet_name in sheet_names:
        if target_sheet and sheet_name != target_sheet:
            continue
        ws_formulas = wb_formulas[sheet_name]
        for row in ws_formulas.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value and cell.value.startswith("=")):
                    col_letter = get_column_letter(cell.column)
                    ref = f"{sheet_name}!{col_letter}{cell.row}"
                    formula_overlay[ref] = cell.value

    wb_formulas.close()  # Free memory before loading computed values

    # Pass 2: Load with data_only=True for computed values
    wb = openpyxl.load_workbook(file_path, data_only=True)

    result = {
        "file": os.path.basename(file_path),
        "file_path": file_path,
        "extracted_at": datetime.now().isoformat(),
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
        "named_ranges": named_ranges,
        "sheets": [],
    }

    for sheet_name in sheet_names:
        if target_sheet and sheet_name != target_sheet:
            continue

        ws = wb[sheet_name]
        sheet_data = extract_sheet(ws, wb, sheet_name, file_path)

        # Apply formula overlay from Pass 1
        for row_data in sheet_data["rows"]:
            for cell_entry in row_data["cells"]:
                ref = cell_entry["ref"]
                if ref in formula_overlay:
                    cell_entry["formula"] = str(formula_overlay[ref])
                    cell_entry["is_calculated"] = True
                    sheet_data["summary"]["formula_cells"] = sheet_data["summary"].get("formula_cells", 0) + 1

        result["sheets"].append(sheet_data)

    wb.close()

    # Compute overall summary
    result["summary"] = {
        "total_sheets": len(result["sheets"]),
        "total_cells": sum(s["summary"]["total_cells_with_data"] for s in result["sheets"]),
        "total_formulas": sum(s["summary"]["formula_cells"] for s in result["sheets"]),
        "total_currency": sum(s["summary"]["currency_cells"] for s in result["sheets"]),
        "total_numeric": sum(s["summary"]["numeric_cells"] for s in result["sheets"]),
        "total_percentages": sum(s["summary"]["percentage_cells"] for s in result["sheets"]),
        "total_text": sum(s["summary"]["text_cells"] for s in result["sheets"]),
        "named_ranges": len(result["named_ranges"]),
    }

    return result


def to_yaml_safe(obj, indent=0):
    """Convert extracted data to YAML format without requiring PyYAML."""
    prefix = "  " * indent
    if obj is None:
        return "null"
    if isinstance(obj, bool):
        return "true" if obj else "false"
    if isinstance(obj, (int, float)):
        return str(obj)
    if isinstance(obj, str):
        if any(c in obj for c in [":", "#", "'", '"', "\n", "{", "}", "[", "]", ",", "&", "*", "!", "|", ">", "%", "@", "`"]):
            escaped = obj.replace("\\", "\\\\").replace('"', '\\"')
            return f'"{escaped}"'
        if not obj:
            return '""'
        return f'"{obj}"'
    if isinstance(obj, list):
        if not obj:
            return "[]"
        lines = []
        for item in obj:
            if isinstance(item, dict):
                dict_yaml = to_yaml_dict(item, indent + 1)
                lines.append(f"{prefix}- {dict_yaml}")
            else:
                lines.append(f"{prefix}- {to_yaml_safe(item)}")
        return "\n".join(lines)
    if isinstance(obj, dict):
        return to_yaml_dict(obj, indent)
    return str(obj)


def to_yaml_dict(d, indent=0):
    """Convert a dict to YAML lines."""
    prefix = "  " * indent
    lines = []
    first = True
    for key, value in d.items():
        if isinstance(value, (dict, list)) and value:
            if isinstance(value, dict):
                lines.append(f"{'' if first else prefix}{key}:")
                lines.append(to_yaml_dict(value, indent + 1))
            elif isinstance(value, list):
                lines.append(f"{'' if first else prefix}{key}:")
                lines.append(to_yaml_safe(value, indent + 1))
        else:
            lines.append(f"{'' if first else prefix}{key}: {to_yaml_safe(value)}")
        first = False
    return "\n".join(lines)


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Extract structured data from XLSX files")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--output", "-o", help="Output YAML file path (default: stdout)")
    parser.add_argument("--sheet", "-s", help="Extract only this sheet name")
    parser.add_argument("--json", action="store_true", help="Output as JSON instead of YAML")
    parser.add_argument("--summary-only", action="store_true", help="Only output summary, not cell data")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"ERROR: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    if not args.file.lower().endswith((".xlsx", ".xlsm")):
        print(f"ERROR: Not an xlsx file: {args.file}", file=sys.stderr)
        sys.exit(1)

    result = extract_workbook(args.file, target_sheet=args.sheet)

    if args.summary_only:
        result = {
            "file": result["file"],
            "file_path": result["file_path"],
            "extracted_at": result["extracted_at"],
            "summary": result["summary"],
            "sheet_names": result["sheet_names"],
            "named_ranges": result["named_ranges"],
        }

    if args.json:
        output = json.dumps(result, indent=2, default=str)
    else:
        output = f"# XLSX Extraction: {result['file']}\n"
        output += f"# Extracted: {result.get('extracted_at', 'unknown')}\n"
        output += f"# Cells: {result.get('summary', {}).get('total_cells', 0)}\n"
        output += f"# Formulas: {result.get('summary', {}).get('total_formulas', 0)}\n\n"
        output += to_yaml_safe(result)

    if args.output:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        with open(args.output, "w") as f:
            f.write(output)
        print(f"Extracted {result['summary']['total_cells']} cells from {result['file']} -> {args.output}", file=sys.stderr)
    else:
        print(output)


if __name__ == "__main__":
    main()
