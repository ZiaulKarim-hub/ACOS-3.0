#!/usr/bin/env python3
"""plan-ordinals.py — renumber MANY rows from a spreadsheet (Zee, 2026-08-24).

  plan-ordinals.py export [--out PATH] [--format xlsx|csv]
  plan-ordinals.py apply --file PATH [--apply]

WHY THIS EXISTS. `manage-ordinals.py renumber` moves ONE row and asks for a dry
run then an apply, so ten rows is ten rounds. Worse, a sequence of single moves
cannot express a reshuffle at all: to swap 5 and 7 you must move one of them to
a number that is already taken, and `renumber` REFUSES that — correctly, because
the refusal is what stops a row being silently overwritten. Zee's answer
(2026-08-24): hand him a sheet with every row and its number, plus an empty
column; he types the numbers he wants; this reads the sheet back and performs
the whole rearrangement in one planned pass.

THE SHEET
  name            the project, as the book shows it
  current_number  what it holds today
  new_number      EMPTY — the column Zee fills in:
                    blank  leave this row exactly where it is
                    a number  move it there
                    0        DELETE this row (Zee, 2026-08-25)
  status / tier / root / project_uuid   context, and the real key

WHY 0 MEANS DELETE. It is the one value that cannot mean a position: the
renderer never issues 0, and no row can hold it. (acos-safe-close uses 0 for
"new project", but that is a different question asked in a different place —
nothing reads both.) So 0 is free here, and a delete typed into the row it
deletes is a per-row deliberate act, not a blanket yes.

DELETES RUN FIRST, then the moves. Deleting frees a number, so a row may move
INTO a number a deleted row is vacating in the same sheet — put 0 on row 5 and
5 on row 9, and row 9 ends up at 5. Doing it the other way round would collide.

`project_uuid` is the identity that survives the round trip. Names repeat on
this machine (`FruitSync`, `To Do Tree` and others each sit on two rows), so a
name could never be the key. Do not delete or reorder that column; rows may be
sorted or re-sorted freely, because nothing here depends on sheet order.

A DELETE IS ALL-OR-NOTHING WITH THE REST. If any 0-marked row has a window open
on it, or its ownership of a close bundle is only a guess, the WHOLE sheet is
refused before anything is written. A half-applied sheet is the state nobody can
reason about afterwards.

THE STALE-PLAN GUARD, which is the point of exporting `current_number` at all.
On apply, every row's `current_number` in the sheet must still match what the
row holds on disk. If anything renumbered, enrolled, or closed in between, the
sheet describes a book that no longer exists and applying it would move rows the
plan never meant to touch. That REFUSES; re-export and redo it.

WHY THE MOVE IS TWO PASSES. Rows that swap places have no legal direct order —
whichever moves first lands on a number the other still holds. So every moving
row is first PARKED on a temporary number above everything in use, and only then
placed on its target. No intermediate state ever has two rows on one number.

NO LOCK, SO A CRASH LEAVES A BREADCRUMB. registry_lib documents no blocking lock
and none surviving SIGKILL. If the process dies mid-rearrangement, rows sit on
temporary numbers with nothing to explain why. So the whole plan is written to
`registry.d/.ordinal-plan-in-progress.json` before the first write and removed
after the last. If you ever see that file, it names exactly what was in flight.

THE LEDGER records ONE `renumber` event per row, from its original number to its
final one, written after the row lands. The temporary parking numbers are never
recorded — they are machinery, not history, and logging them would make
`history_for(7)` answer with numbers no project ever really had.

HUMAN-INITIATED ONLY, like every verb in manage-ordinals.py. Never export-and-
apply on your own initiative. Zee fills the column; nothing here invents one.

Python, not TypeScript, by the standing exception: it calls registry_lib and
ordinal_lib writers directly, so the schema gate, the atomic-write pattern and
the append-only ledger discipline all still apply. CSV needs the stdlib only.
XLSX needs openpyxl, which is present on python3 3.14.6 here but NOT on
/usr/bin/python3 3.9.6 — so the format falls back to CSV, loudly, when the
module is missing rather than failing at the last step.

Exit codes: 0 = done, 1 = refused (a stated reason), 2 = could not run.
"""

import argparse
import shutil
import csv
import json
import os
import subprocess
import sys

HERE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE_DIR)
import merge_lib  # noqa: E402
import ordinal_lib  # noqa: E402
import registry_lib  # noqa: E402


def _manage():
    """manage-ordinals.py, imported by path — `import manage-ordinals` is not
    legal Python. Deletes go through ITS verb, never a copy of it: the archive
    step, the freed number, the kept facts and the confirmation route all live
    there, and a second implementation would drift from the first."""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "manage_ordinals", os.path.join(HERE_DIR, "manage-ordinals.py"))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

VIEW = os.path.join(HERE_DIR, "resurrect-view.py")
REG_HOME = os.environ.get("ACOS_REGISTRY_HOME") or None
COLUMNS = ["name", "current_number", "new_number", "status", "tier", "root", "project_uuid"]
DEFAULT_DIR = os.path.join(os.path.expanduser("~"), "Documents", "OKOA")
DEFAULT_STEM = "acos-row-numbers"
IN_PROGRESS = ".ordinal-plan-in-progress.json"


class Refused(Exception):
    pass


def stamp():
    """A filename-safe local timestamp, for the copy kept when a sheet with
    typed-in numbers is replaced."""
    import datetime
    return datetime.datetime.now().strftime("%Y%m%d-%H%M%S")


def in_progress_path(home=None):
    return os.path.join(registry_lib.registry_dir(home), IN_PROGRESS)


def fresh_book(no_cmux=False):
    """The book, computed NOW — the same source of names and numbers the
    resurrect menu prints, so the sheet can never disagree with the book."""
    argv = [sys.executable, VIEW, "--json"]
    if no_cmux or os.environ.get("RESURRECTION_SKIP_CMUX") == "1":
        argv.append("--no-cmux")
    out = subprocess.run(argv, capture_output=True, text=True, timeout=180)
    if out.returncode != 0:
        raise Refused("resurrect-view.py --json rc=%d stderr=%s"
                      % (out.returncode, out.stderr.strip()[:300]))
    try:
        return json.loads(out.stdout)
    except (json.JSONDecodeError, ValueError):
        raise Refused("resurrect-view.py --json returned unparseable JSON: %r"
                      % out.stdout[:200])


def sheet_rows(book):
    """Every NUMBERED row, ARCHIVED included, sorted by the number it holds.

    Archived rows are listed on purpose: they carry numbers now, a number is
    what every verb uses to name a row, and leaving them out would make the
    sheet's numbers look like they had holes that are free to take. They are
    marked by their status column so nothing is disguised."""
    rows = [p for p in book.get("projects", []) if p.get("pick_number") is not None]
    rows.sort(key=lambda p: p["pick_number"])
    return [{"name": p.get("name"), "current_number": p["pick_number"], "new_number": "",
             "status": p.get("status"), "tier": p.get("tier"), "root": p.get("root"),
             "project_uuid": p.get("project_uuid")} for p in rows]


# --------------------------------------------------------------------------
# export
# --------------------------------------------------------------------------

def write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(path, rows):
    """Nicer to edit than raw CSV: a frozen header, sized columns, and the one
    column Zee fills marked so it cannot be confused with the read-only ones."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "row numbers"
    ws.append([c.replace("_", " ") for c in COLUMNS])
    head_fill = PatternFill("solid", fgColor="DDDDDD")
    edit_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
    edit_col = COLUMNS.index("new_number") + 1
    for row in ws.iter_rows(min_row=2, min_col=edit_col, max_col=edit_col):
        for cell in row:
            cell.fill = edit_fill
    for col, width in zip("ABCDEFG", (34, 16, 14, 12, 12, 60, 40)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def carry_forward(rows, path):
    """Copy the new_number column of an EARLIER sheet onto a fresh one.

    Exporting rewrites the whole file, so an export run after Zee has filled the
    column would wipe his typing. Rows are matched by project_uuid, never by
    name — names repeat on this machine, and a row that moved folders keeps its
    uuid. Returns (carried, dropped, blank) where `dropped` names the old rows
    whose project no longer exists, so they are reported rather than lost in
    silence."""
    try:
        old = read_sheet(path)
    except Refused:
        raise
    except Exception as exc:
        raise Refused("could not read the earlier sheet %s: %s" % (path, exc))

    typed = {}
    for i, r in enumerate(old, start=2):
        uuid = str(r.get("project_uuid", "")).strip()
        raw = r.get("new_number", "")
        text = "" if raw is None else str(raw).strip()
        if not uuid or not text:
            continue
        typed[uuid] = (_as_int(raw, "new_number", i), str(r.get("name", "")).strip())

    carried, blank = [], []
    for r in rows:
        hit = typed.pop(str(r.get("project_uuid", "")).strip(), None)
        if hit is None:
            blank.append(r)
        else:
            r["new_number"] = hit[0]
            carried.append(r)
    dropped = [{"project_uuid": u, "name": n, "new_number": v} for u, (v, n) in typed.items()]
    dropped.sort(key=lambda d: (d["new_number"], d["name"]))
    return carried, dropped, blank


def filled_count(path):
    """How many new_number cells an existing sheet already has typed in. Used to
    decide whether overwriting it would destroy work."""
    try:
        rows = read_sheet(path)
    except Exception:
        return 0
    n = 0
    for r in rows:
        v = r.get("new_number", "")
        if v is not None and str(v).strip() != "":
            n += 1
    return n


def cmd_export(args):
    book = fresh_book()
    rows = sheet_rows(book)
    if not rows:
        raise Refused("no numbered rows in the book — nothing to export")

    fmt = args.format
    note = None
    if fmt == "xlsx":
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            fmt = "csv"
            note = ("openpyxl is not importable under %s, so this is a CSV instead. "
                    "Excel opens it the same way; save it as CSV when you are done."
                    % sys.executable)

    path = args.out or os.path.join(DEFAULT_DIR, "%s.%s" % (DEFAULT_STEM, fmt))
    if args.out and not args.out.lower().endswith("." + fmt):
        path = os.path.splitext(args.out)[0] + "." + fmt
    d = os.path.dirname(os.path.abspath(path))
    if not os.path.isdir(d):
        os.makedirs(d, exist_ok=True)

    # Carry Zee's typing forward. An export rewrites the whole file, so without
    # this a re-export after he has filled the column silently blanks his work —
    # which is exactly what happened on 2026-08-25. Default is to carry from the
    # file being replaced; --carry-from names a different one; --blank opts out.
    source = args.carry_from
    if source is None and not args.blank and os.path.exists(path):
        source = path
    carried = dropped = None
    backup = None
    if source and os.path.exists(source):
        # Only when the file we read from is the file we are about to overwrite.
        # Carrying from a different sheet leaves that sheet untouched anyway.
        if os.path.abspath(source) == os.path.abspath(path) and filled_count(source):
            backup = "%s.replaced-%s%s" % (os.path.splitext(source)[0], stamp(),
                                           os.path.splitext(source)[1])
            shutil.copy2(source, backup)
        carried, dropped, _blank = carry_forward(rows, source)

    (write_xlsx if fmt == "xlsx" else write_csv)(path, rows)

    if note:
        print("NOTE — %s" % note)
    print("wrote %s" % path)
    print("  %d numbered rows, sorted by the number they hold now" % len(rows))
    archived = [r for r in rows if r["status"] in ("tombstoned", "completed")]
    if archived:
        print("  %d of them are ARCHIVED (status tombstoned or completed) and are listed "
              "so their numbers are visible, not so they are moved" % len(archived))
    if carried is not None:
        print("")
        print("carried your typed numbers forward from %s" % source)
        print("  %d row%s already have a number typed in — they are still there"
              % (len(carried), "" if len(carried) == 1 else "s"))
        if backup:
            print("  the file it replaced is kept at %s" % backup)
        if dropped:
            print("  %d typed row%s could NOT be carried — no such project any more:"
                  % (len(dropped), "" if len(dropped) == 1 else "s"))
            for d in dropped:
                print("      %-38s you had typed %s" % (d["name"][:38], d["new_number"]))
        dupes = {}
        for r in carried:
            if r["new_number"] != 0:
                dupes.setdefault(r["new_number"], []).append(r["name"])
        shared = {k: v for k, v in dupes.items() if len(v) > 1}
        if shared:
            print("  %d number%s given to more than one row — apply will refuse these "
                  "until you settle each one:" % (len(shared), "" if len(shared) == 1 else "s"))
            for k in sorted(shared):
                print("      %-4s <- %s" % (k, ", ".join(shared[k])))

    print("")
    print("Fill in the `new_number` column:")
    print("  BLANK      leave that row exactly where it is")
    print("  a number   move the row to that number")
    print("  0          DELETE that row — its number is freed, its close bundles are")
    print("             archived, and its knowledge facts are kept")
    print("Do not delete the `project_uuid` column — names repeat on this machine, so the "
          "uuid is the only thing that identifies a row.")
    print("Then: plan-ordinals.py apply --file %s" % path)
    return 0


# --------------------------------------------------------------------------
# apply
# --------------------------------------------------------------------------

def read_sheet(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as fh:
            return [dict(r) for r in csv.DictReader(fh)]
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise Refused("reading %s needs openpyxl, which is not importable under %s. "
                          "Re-export with --format csv, or run this with a python that "
                          "has openpyxl." % (ext, sys.executable))
        ws = load_workbook(path, data_only=True).active
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip().replace(" ", "_") if h is not None else ""
                      for h in next(it)]
        except StopIteration:
            raise Refused("%s is empty" % path)
        out = []
        for values in it:
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            out.append({h: ("" if v is None else v) for h, v in zip(header, values)})
        return out
    raise Refused("unrecognised file type %r — expected .csv or .xlsx" % ext)


def _as_int(value, what, line):
    """Excel hands back 3, 3.0 or '3' for the same typed cell. All three mean 3;
    3.5 does not, and is refused rather than rounded into a different row."""
    if isinstance(value, bool):
        raise Refused("line %d: %s is a boolean, not a number" % (line, what))
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != int(value):
            raise Refused("line %d: %s is %r — whole numbers only" % (line, what, value))
        return int(value)
    text = str(value).strip()
    try:
        return int(text)
    except ValueError:
        raise Refused("line %d: %s is %r — expected a whole number" % (line, what, text))


def build_plan(records, home=None, **kwargs):
    """Validate the sheet against the registry AS IT IS NOW. Returns
    (moves, unchanged, final_layout). Refuses on the first thing that is wrong,
    because a half-checked plan is not safer than an unchecked one."""
    if not records:
        raise Refused("the sheet has no data rows")
    missing = [c for c in ("current_number", "new_number", "project_uuid")
               if c not in records[0]]
    if missing:
        raise Refused("the sheet is missing required column(s): %s. Re-export it."
                      % ", ".join(missing))

    green = kwargs.get("green") or frozenset()
    overrides = kwargs.get("survivors") or {}
    moves, deletes, unchanged, seen_uuids = [], [], [], set()
    typed_same, blank_holders = [], []
    for i, rec in enumerate(records, start=2):  # 2 = first data line in a sheet
        uuid = str(rec.get("project_uuid") or "").strip()
        if not uuid:
            raise Refused("line %d: project_uuid is empty — that column identifies the row "
                          "and cannot be blank" % i)
        if uuid in seen_uuids:
            raise Refused("line %d: project_uuid %s appears twice in the sheet" % (i, uuid))
        seen_uuids.add(uuid)

        try:
            row = registry_lib.load_row(uuid, home)
        except (ValueError, json.JSONDecodeError) as exc:
            raise Refused("line %d: registry row %s is corrupt: %s" % (i, uuid, exc))
        if row is None:
            raise Refused("line %d: no registry row for %s. It may have been deleted since "
                          "the sheet was exported — re-export and redo the plan." % (i, uuid))

        stated = _as_int(rec["current_number"], "current_number", i)
        if row["pick_ordinal"] != stated:
            raise Refused(
                "line %d: %r holds number %s now, but the sheet says %s. The book changed "
                "after this sheet was exported, so applying it would move rows this plan "
                "never meant to touch. Re-export and redo it."
                % (i, row["name"], row["pick_ordinal"], stated))

        raw_new = rec.get("new_number")
        if raw_new is None or str(raw_new).strip() == "":
            unchanged.append((row, stated))
            blank_holders.append((row, stated))
            continue
        want = _as_int(raw_new, "new_number", i)
        if want == 0:
            # Zee, 2026-08-25: 0 in this column means DELETE this row's project.
            deletes.append({"row": row, "uuid": uuid, "name": row["name"],
                            "from": stated, "line": i})
            continue
        if want < 0:
            raise Refused("line %d: new_number is %d. A row cannot hold a negative number. "
                          "Use 0 to delete the row, a positive number to move it, or leave "
                          "the cell blank to leave it alone." % (i, want))
        if want == stated:
            # Typed, and the same as it already holds. On its own that is "leave
            # it alone"; alongside another row typed with the same number it is
            # half of a merge, so it is tracked separately from a blank cell.
            unchanged.append((row, stated))
            typed_same.append((row, stated))
            continue
        moves.append({"row": row, "uuid": uuid, "name": row["name"],
                      "from": stated, "to": want, "line": i})

    if not moves and not deletes:
        raise Refused("no row has a new_number that differs from its current one, and "
                      "none is marked 0 for deletion — nothing to do")

    # A deleted row stops holding its number, so another row may move into it.
    # Checked BEFORE the layout below, because that check asks who holds what.
    freed = {d["from"] for d in deletes}

    # Every number in the FINAL layout, from all three sources.
    # Two or more rows given ONE number is a MERGE instruction, not an error
    # (Zee, 2026-08-25). It says those rows are one project wearing several
    # rows, so they are folded together rather than refused.
    # A merge needs the number TYPED on every row in the group. Zee's words were
    # "I will put in the same number for the duplicates" — a BLANK cell was never
    # typed, so a mover landing on a blank row's number is a collision he did not
    # ask for and still refuses. Live case: he sent #17 to 34 while `zee` sat on
    # 34 with an empty cell. Treating that as a merge would have deleted a row.
    claimants, staying = {}, {}
    for m in moves:
        claimants.setdefault(m["to"], []).append(m)
    for row, n in typed_same:
        claimants.setdefault(n, []).append(
            {"row": row, "uuid": row["project_uuid"], "name": row["name"],
             "from": n, "to": n, "line": None, "staying": True})
    for row, n in blank_holders:
        staying[n] = row

    merges, final = [], {}
    for n, members in sorted(claimants.items()):
        holder = staying.get(n)
        if holder is not None:
            raise Refused("number %d is given to %r, but %r is not moving and still holds "
                          "it. Its new_number cell is EMPTY, so this is a collision rather "
                          "than a merge. Give %r a new_number too, type %d on it as well to "
                          "merge them, mark it 0 to delete it, or pick another number."
                          % (n, members[0]["name"], holder["name"], holder["name"], n))
        if len(members) == 1:
            final[n] = members[0]
            continue
        try:
            winner, losers, reason = merge_lib.pick_survivor(
                members, green=green, overrides={n: overrides[n]} if n in overrides else {},
                home=home)
        except ValueError as exc:
            raise Refused("number %d: %s" % (n, exc))
        merges.append({"to": n, "survivor": winner, "losers": losers, "reason": reason})
        final[n] = winner
        # A loser stops holding its old number the moment it is deleted, so the
        # number it vacates is free for anything else in this same sheet.
        for lose in losers:
            freed.add(lose["from"])
        # A survivor that was only ever "staying put" still counts as unchanged;
        # one that moves is already in `moves` and needs no second entry.

    # Rows on disk the sheet never mentioned — usually a deleted row, or one
    # enrolled after the export. They hold numbers too.
    sheet_uuids = seen_uuids
    for n in sorted(ordinal_lib.held_ordinals(home)):
        if n in freed:
            continue        # a row marked 0 vacates its number in this same pass
        holders = [r for r in ordinal_lib.live_holders(home).get(n, [])
                   if r["project_uuid"] not in sheet_uuids]
        if not holders or n not in final:
            continue
        target = final[n]
        if target.get("line") is None:
            continue
        raise Refused("number %d is given to %r (line %d), but %r holds it and is not in the "
                      "sheet. Re-export so the sheet shows every row."
                      % (n, target["name"], target["line"], holders[0]["name"]))
    return moves, deletes, unchanged, final, merges


def check_deletable(deletes, home=None):
    """Refuse the WHOLE sheet if any 0-marked row cannot be deleted safely.

    Two blockers, both checked BEFORE anything is written:

      LIVE WINDOW — deleting a row with a window open on it leaves that window
        bound to a row the book no longer lists. manage-ordinals refuses this
        per row; here it must refuse the whole sheet, because a sheet that
        applies "most of it" is the state nobody can reason about afterwards.
      UNPROVEN BUNDLE — a close bundle whose ownership is a guess would be
        left behind by the delete and quietly reported. In a bulk run nobody
        reads per-row notes, so it is promoted to a refusal with the fix named.
    """
    if not deletes:
        return
    manage = _manage()
    live, live_err = manage.live_workspaces_by_uuid(False)
    problems = []
    for d in deletes:
        row = d["row"]
        if live_err:
            problems.append("line %d: %r — the open-window check could not run (%s)"
                            % (d["line"], d["name"], live_err))
            continue
        if live.get(row["project_uuid"].lower()):
            problems.append("line %d: %r has a window OPEN on it — close that window first"
                            % (d["line"], d["name"]))
        _proven, guessed = manage.owned_bundles(row, home)
        if guessed:
            problems.append(
                "line %d: %r owns %d close bundle(s) only by a GUESS, so a delete would "
                "leave them behind: %s. Run stamp-bundle-owners.py first."
                % (d["line"], d["name"], len(guessed),
                   ", ".join(os.path.basename(b) for b, _e in guessed)))
    if problems:
        raise Refused("%d row(s) marked 0 cannot be deleted safely; NOTHING was written:\n  - %s"
                      % (len(problems), "\n  - ".join(problems)))


def check_mergeable(merges, home=None):
    """Refuse the WHOLE sheet if any merge cannot be done cleanly.

    Same reasoning as check_deletable: a bulk run that applies "most of it"
    leaves a state nobody can reason about. The one blocker is an UNPROVEN
    bundle on a losing row — ownership by resemblance is not a reason to hand a
    project's history to a different row."""
    if not merges:
        return
    problems = []
    for g in merges:
        members = [g["survivor"]] + list(g["losers"])
        uuids = [m["uuid"] for m in members]
        for lose in g["losers"]:
            _rescued, doubtful = merge_lib.resolvable_guesses(lose["row"], uuids, home)
            if doubtful:
                problems.append(
                    "number %d: %r owns %d close bundle(s) only by a GUESS, and some row "
                    "OUTSIDE this merge could claim them too, so a merge would strand "
                    "them: %s. Run stamp-bundle-owners.py first."
                    % (g["to"], lose["name"], len(doubtful),
                       ", ".join(os.path.basename(b) for b, _e in doubtful)))
    if problems:
        raise Refused("%d merge problem(s); NOTHING was written:\n  - %s"
                      % (len(problems), "\n  - ".join(problems)))


def preview_merges(merges, home=None):
    """What each merge WOULD move. Read before anything is written."""
    out = []
    for g in merges:
        surv = g["survivor"]
        detail = {"to": g["to"], "survivor": surv, "reason": g["reason"], "losers": []}
        for lose in g["losers"]:
            facts, already = merge_lib.merge_knowledge(lose["uuid"], surv["uuid"],
                                                       home, dry=True)
            bundles, guessed = merge_lib.transfer_bundles(
                lose["row"], surv["row"], home, dry=True,
                group_uuids=[m["uuid"] for m in [surv] + list(g["losers"])])
            windows = merge_lib.transfer_windows(lose["uuid"], surv["uuid"], home, dry=True)
            detail["losers"].append({"member": lose, "facts": facts, "already": already,
                                     "bundles": bundles, "guessed": guessed,
                                     "windows": windows})
        out.append(detail)
    return out


def render_merges(previews):
    if not previews:
        return ""
    lines = ["", "MERGE — %d number%s given to more than one row:"
             % (len(previews), "" if len(previews) == 1 else "s")]
    for d in previews:
        lines.append("")
        lines.append("  number %d  KEEP %r  (%s)" % (d["to"], d["survivor"]["name"], d["reason"]))
        for L in d["losers"]:
            m = L["member"]
            lines.append("      absorb %r (was %d)" % (m["name"], m["from"]))
            lines.append("        %d fact(s) move, %d already there" % (L["facts"], L["already"]))
            lines.append("        %d close bundle(s) hand over" % len(L["bundles"]))
            for src_b, dst_b in L["bundles"]:
                if os.path.dirname(src_b) != os.path.dirname(dst_b):
                    lines.append("          %s  ->  %s" % (os.path.basename(src_b), dst_b))
            if L["windows"]:
                lines.append("        %d window claim(s) re-point" % len(L["windows"]))
            lines.append("        then the row is DELETED — number %d freed, remaining "
                         "bundles archived, facts kept" % m["from"])
    return "\n".join(lines)


def apply_merges(merges, home=None):
    """Absorb, then delete. Returns (done, failed).

    ORDER MATTERS. Content moves to the survivor BEFORE the losing row is
    deleted, because delete archives whatever the row still owns — a bundle
    archived first would then be handed over from inside an archive folder,
    where nothing looks for it."""
    manage = _manage()
    done, failed = [], []
    for g in merges:
        surv = g["survivor"]
        for lose in g["losers"]:
            try:
                facts, _already = merge_lib.merge_knowledge(lose["uuid"], surv["uuid"],
                                                            home, dry=False)
                bundles, _guessed = merge_lib.transfer_bundles(
                    lose["row"], surv["row"], home, dry=False,
                    group_uuids=[m["uuid"] for m in [surv] + list(g["losers"])])
                windows = merge_lib.transfer_windows(lose["uuid"], surv["uuid"], home,
                                                     dry=False)
            except Exception as exc:  # noqa: BLE001 — report, never half-die silently
                failed.append((g, lose, "absorb failed: %s: %s" % (type(exc).__name__, exc)))
                continue
            argv = ["delete", str(lose["from"]), "--confirm-name", lose["row"]["name"],
                    "--apply"]
            if home:
                argv += ["--home", home]
            code = manage.main(argv)
            if code != 0:
                failed.append((g, lose, "absorbed, but the delete exited %s" % code))
                continue
            registry_lib.audit_append(
                {"event": "rows-merged", "project_uuid": surv["uuid"],
                 "from_project_uuid": lose["uuid"], "number": g["to"],
                 "reason": g["reason"], "facts_moved": facts,
                 "bundles_moved": len(bundles), "windows_moved": len(windows)},
                home=home)
            done.append((g, lose))
    return done, failed


def render_plan(moves, deletes, unchanged):
    lines = ["PLAN — %d row%s move, %d DELETED, %d stay put:"
             % (len(moves), "" if len(moves) == 1 else "s", len(deletes), len(unchanged)), ""]
    for m in sorted(moves, key=lambda x: x["to"]):
        lines.append("  %4d -> %-4d  %s" % (m["from"], m["to"], m["name"]))
    if deletes:
        lines.append("")
        lines.append("  DELETE (number freed, close bundles archived, knowledge facts kept):")
        for d in sorted(deletes, key=lambda x: x["from"]):
            lines.append("  %4d -> DEL   %s" % (d["from"], d["name"]))
    return "\n".join(lines)


def apply_plan(moves, home=None):
    """Park every mover above everything in use, then place each one.

    Two passes because a swap has no legal direct order: whichever row moves
    first would land on a number the other still holds, and that is exactly the
    collision the single-row verb refuses. Parking guarantees no intermediate
    state ever puts two rows on one number.
    """
    held = ordinal_lib.held_ordinals(home)
    base = max(list(held) + [ordinal_lib.max_ever_issued(home)] +
               [m["to"] for m in moves]) + 1

    plan_record = {"moves": [{"project_uuid": m["uuid"], "name": m["name"],
                              "from": m["from"], "to": m["to"]} for m in moves],
                   "park_base": base}
    marker = in_progress_path(home)
    registry_lib.atomic_write_json(marker, plan_record)
    try:
        for offset, m in enumerate(moves):
            m["park"] = base + offset
            registry_lib.set_pick_ordinal(m["uuid"], m["park"], home)
        for m in moves:
            registry_lib.set_pick_ordinal(m["uuid"], m["to"], home)
            # ONE ledger entry per row, written after it lands, naming the real
            # from and to. The parking number is machinery and is never logged.
            ordinal_lib.append_event("renumber", m["to"], m["uuid"], m["name"], home,
                                     from_ordinal=m["from"], bulk=True)
    finally:
        try:
            os.unlink(marker)
        except OSError:
            pass

    # Re-read. Never trust the writes — report what actually landed.
    problems = []
    for m in moves:
        back = registry_lib.load_row(m["uuid"], home)
        got = back["pick_ordinal"] if back else None
        if got != m["to"]:
            problems.append("%r should hold %d but holds %r" % (m["name"], m["to"], got))
    clashes = {n: rs for n, rs in ordinal_lib.live_holders(home).items() if len(rs) > 1}
    for n, rs in sorted(clashes.items()):
        problems.append("number %d is held by %d rows: %s"
                        % (n, len(rs), ", ".join(r["name"] for r in rs)))
    return problems


def apply_deletes(deletes, home=None):
    """Run manage-ordinals' own delete verb, once per 0-marked row.

    Never a re-implementation: that verb owns the archive step, the freed
    number, the kept facts, the manifest restore reads, and the audit line.
    --confirm-name is the row's OWN name, taken from the registry rather than
    from the sheet — the sheet's `name` column is context, and a stale one must
    not be able to satisfy the confirmation for a row that has been renamed.
    """
    manage = _manage()
    done, failed = [], []
    for d in deletes:
        code = manage.main(["delete", str(d["from"]),
                            "--confirm-name", d["row"]["name"],
                            "--apply", "--home", home] if home else
                           ["delete", str(d["from"]),
                            "--confirm-name", d["row"]["name"], "--apply"])
        (done if code == 0 else failed).append((d, code))
    return done, failed


def cmd_apply(args):
    records = read_sheet(args.file)
    green = merge_lib.sheet_green_uuids(args.file)
    survivors = {}
    for spec in (args.survivor or []):
        if "=" not in spec:
            raise Refused("--survivor wants NUMBER=UUID, got %r" % spec)
        num, uuid = spec.split("=", 1)
        try:
            survivors[int(num.strip())] = uuid.strip()
        except ValueError:
            raise Refused("--survivor %r: %r is not a number" % (spec, num))

    moves, deletes, unchanged, _final, merges = build_plan(
        records, REG_HOME, green=green, survivors=survivors)
    check_deletable(deletes, REG_HOME)
    check_mergeable(merges, REG_HOME)
    previews = preview_merges(merges, REG_HOME)
    print(render_plan(moves, deletes, unchanged))
    print(render_merges(previews))
    print("")
    if green:
        print("%d row%s marked green in the sheet." % (len(green), "" if len(green) == 1 else "s"))
    weak = [d for d in previews if d["reason"].startswith("holds the most")]
    if weak:
        print("READ THIS — %d merge%s picked its survivor by content, because nothing was "
              "marked. Mark the winner green, or pass --survivor NUMBER=UUID, if any of "
              "these is wrong:" % (len(weak), "" if len(weak) == 1 else "s"))
        for d in weak:
            print("    number %-4d keeps %r" % (d["to"], d["survivor"]["name"]))
        print("")
    if not args.apply:
        print("DRY RUN — nothing written. Re-run with --apply to perform it.")
        return 0

    if merges:
        print("=" * 72)
        print("MERGING into %d row%s" % (len(merges), "" if len(merges) == 1 else "s"))
        print("=" * 72)
        mdone, mfailed = apply_merges(merges, REG_HOME)
        print("")
        if mfailed:
            print("PARTIAL — %d merge step(s) failed. Deletes and moves were NOT attempted:"
                  % len(mfailed))
            for g, lose, why in mfailed:
                print("  - number %d, %r: %s" % (g["to"], lose["name"], why))
            print("  Run `manage-ordinals.py status`, fix the cause, then re-export. "
                  "Nothing here retries or guesses.")
            return 1
        print("absorbed and removed %d row%s." % (len(mdone), "" if len(mdone) == 1 else "s"))
        print("")

    # Deletes FIRST: a delete frees a number, and a move may be heading into it.
    if deletes:
        print("=" * 72)
        print("DELETING %d row%s" % (len(deletes), "" if len(deletes) == 1 else "s"))
        print("=" * 72)
        done, failed = apply_deletes(deletes, REG_HOME)
        print("")
        if failed:
            print("PARTIAL — %d of %d deletes failed. The moves were NOT attempted:"
                  % (len(failed), len(deletes)))
            for d, code in failed:
                print("  - %r (line %d) exit %d" % (d["name"], d["line"], code))
            print("  Run `manage-ordinals.py status` to see what landed, fix the cause, "
                  "then re-export the sheet. Nothing here retries or guesses.")
            return 1
        print("deleted %d row%s." % (len(done), "" if len(done) == 1 else "s"))
        print("")

    # Drop movers that no longer exist. A merge REMOVES its losing rows, and a
    # loser was typed with the group's number, so it is sitting in `moves` too.
    # Moving a row that was just deleted raises KeyError halfway through the
    # parking pass — measured 2026-08-25, on the second mover of a real run.
    # The survivor keeps its own move; only the vanished rows are dropped.
    gone = [m for m in moves if registry_lib.load_row(m["uuid"], REG_HOME) is None]
    if gone:
        moves = [m for m in moves if m not in gone]
        print("%d mover%s no longer exist%s — merged away or deleted above, so there is "
              "nothing left to move:" % (len(gone), "" if len(gone) == 1 else "s",
                                         "s" if len(gone) == 1 else ""))
        for m in gone:
            print("    %-34s was %d, would have gone to %d" % (m["name"], m["from"], m["to"]))
        print("")

    if not moves:
        print("DONE — %d row%s deleted, %d merged, no moves left in this sheet."
              % (len(deletes), "" if len(deletes) == 1 else "s", len(merges)))
        return 0

    problems = apply_plan(moves, REG_HOME)
    if problems:
        print("PARTIAL — the rearrangement did not fully land. NOT retried:")
        for p in problems:
            print("  - %s" % p)
        print("  Run `manage-ordinals.py status` to see the current numbers, and "
              "`conflict-scan.py` for ORDINAL-CLASH. Fix by hand; nothing here guesses.")
        return 1
    print("DONE — %d row%s renumbered, verified by re-reading every one.%s"
          % (len(moves), "" if len(moves) == 1 else "s",
             ("  %d number%s merged." % (len(merges), "" if len(merges) == 1 else "s"))
             if merges else ""))
    for m in sorted(moves, key=lambda x: x["to"]):
        print("  %-34s now %d (was %d)" % (m["name"], m["to"], m["from"]))
    return 0


def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="plan-ordinals.py", description=__doc__.splitlines()[0],
        epilog="HUMAN-INITIATED ONLY. Zee fills the new_number column; never invent one.")
    sub = ap.add_subparsers(dest="verb", required=True)

    p = sub.add_parser("export", help="write the sheet of rows and numbers")
    p.add_argument("--out", default=None, help="where to write it (default: %s/%s.xlsx)"
                                               % (DEFAULT_DIR, DEFAULT_STEM))
    p.add_argument("--format", choices=("xlsx", "csv"), default="xlsx")
    p.add_argument("--carry-from", default=None, metavar="PATH",
                   help="copy the new_number column of THIS sheet onto the fresh one "
                        "(default: the file being replaced, so your typing survives)")
    p.add_argument("--blank", action="store_true",
                   help="do not carry anything forward — every new_number starts empty")
    p.set_defaults(fn=cmd_export)

    p = sub.add_parser("apply", help="read the filled-in sheet and renumber")
    p.add_argument("--file", required=True)
    p.add_argument("--apply", action="store_true", help="perform it (default is a dry run)")
    p.add_argument("--survivor", action="append", metavar="NUMBER=UUID", default=[],
                   help="for a merged number, name the row that survives. Beats a green "
                        "fill. Repeatable, one per number.")
    p.set_defaults(fn=cmd_apply)

    args = ap.parse_args(sys.argv[1:] if argv is None else argv)
    try:
        return args.fn(args)
    except Refused as exc:
        print("REFUSED — %s" % exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
