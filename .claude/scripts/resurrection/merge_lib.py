#!/usr/bin/env python3
"""merge_lib.py — fold two or more registry ROWS that are one project into one.

Why this exists (Zee, 2026-08-25). He filled the bulk-renumber sheet and gave
the SAME number to two or more rows in eight places. That is not a mistake to
refuse; it is an instruction: "these rows are one project, join them." His rule
for picking which name survives is a GREEN fill on the winning row's number
cell — "in case they have different names, go with the name for which I mark
the number with green color."

WHAT A MERGE IS HERE
  one SURVIVOR row absorbs every LOSER row, then each loser is deleted.

  knowledge facts   moved into the survivor's store by merge-knowledge.py,
                    which de-duplicates by content hash and carries the
                    struck / superseded edges so the survivor keeps the same
                    view of what is still true.
  close bundles     re-stamped to the survivor. Where the roots differ the
                    bundle DIRECTORY is moved as well, because owned_bundles()
                    only ever looks under a row's own root — a re-stamp alone
                    would leave the history unreachable.
  windows           re-pointed at the survivor, so an open tab keeps working.
  the loser row     deleted through manage-ordinals' own verb: number freed,
                    remaining bundles archived, knowledge facts kept.

WHY ABSORB RATHER THAN JUST DELETE. Delete already keeps a row's facts — they
live in a store addressed by project_uuid, not by the row. But nothing would
ever open that store again, because the row that named it is gone. Absorbing
first is what makes the kept facts reachable. Logo Builder is the live case:
18 facts on one row, 16 on the other, both real.

PICKING THE SURVIVOR, in order:
  1. an explicit override, by uuid — Zee named one for the loan-intake group
  2. a GREEN fill on that row's new_number cell in the sheet
  3. the row holding the most content (facts + 3x bundles), reported as a
     WEAK reason so it is read before it is applied

Python, not TypeScript, by the standing exception: this extends the existing
Python resurrection family and calls knowledge_lib / registry_lib / bundles_lib
writers directly, so their schema gates and append-only discipline still apply.
"""

import importlib.util
import os
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import bundles_lib      # noqa: E402
import knowledge_lib    # noqa: E402
import registry_lib     # noqa: E402
import windows_lib      # noqa: E402

# Green as Excel reports it. openpyxl hands back 8 hex digits (alpha + RGB) or
# 6, and a theme colour comes back as an index instead, which is why the test
# is "does the green channel dominate" rather than one exact string.
GREEN_HINTS = ("FF00B050", "0000B050", "00B050", "FF92D050", "92D050",
               "FFC6EFCE", "C6EFCE", "FF00FF00", "00FF00", "FF008000", "008000")


def _load(stem, filename):
    spec = importlib.util.spec_from_file_location(stem, os.path.join(HERE, filename))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def is_green(rgb):
    """True when a cell fill reads as green. Zee marks the surviving name that
    way. Excel has many greens, so this accepts the common swatches AND any
    colour whose green channel clearly leads — a rule of thumb, not a law."""
    if not rgb or not isinstance(rgb, str):
        return False
    text = rgb.strip().upper()
    if text in GREEN_HINTS:
        return True
    if len(text) == 8:
        text = text[2:]
    if len(text) != 6:
        return False
    try:
        r, g, b = (int(text[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return False
    if r == g == b:
        return False                      # grey, including the sheet's own header
    return g > r + 24 and g > b + 24


def sheet_green_uuids(path):
    """Which rows carry a GREEN new_number cell, by project_uuid.

    Reads the workbook a second time on purpose: read_sheet() asks openpyxl for
    values only, which is faster and is all the numbers need, but that mode
    throws styling away."""
    if not path.lower().endswith((".xlsx", ".xlsm")):
        return set()                      # CSV carries no colour at all
    try:
        from openpyxl import load_workbook
    except ImportError:
        return set()
    try:
        ws = load_workbook(path).active
    except Exception:
        return set()
    header = [str(c.value).strip().replace(" ", "_") if c.value is not None else ""
              for c in ws[1]]
    try:
        ncol = header.index("new_number")
        ucol = header.index("project_uuid")
    except ValueError:
        return set()
    out = set()
    for row in ws.iter_rows(min_row=2):
        if ncol >= len(row) or ucol >= len(row):
            continue
        cell = row[ncol]
        fill = cell.fill
        if not fill or fill.patternType != "solid":
            continue
        if is_green(getattr(fill.start_color, "rgb", None)):
            uuid = row[ucol].value
            if uuid:
                out.add(str(uuid).strip())
    return out


def content_score(project_uuid, root, home=None):
    """(facts, bundles) — how much a row actually holds.

    Used only as the LAST way to pick a survivor, when Zee has marked nothing.
    Bundles weigh 3x a fact in the ranking because a close bundle is a whole
    session's dossier, not one claim."""
    facts = len(knowledge_lib.load_facts(project_uuid, home))
    bundles = 0
    cdir = os.path.join(root or "", "memory", "handoffs", "closed")
    if os.path.isdir(cdir):
        for name in os.listdir(cdir):
            marker = os.path.join(cdir, name, bundles_lib.OWNER_MARKER)
            if not os.path.exists(marker):
                continue
            try:
                with open(marker, encoding="utf-8") as fh:
                    if fh.read().strip() == project_uuid:
                        bundles += 1
            except OSError:
                continue
    return facts, bundles


def pick_survivor(members, green=frozenset(), overrides=None, home=None):
    """Choose which row of a merge group lives. Returns (survivor, losers, reason).

    `members` are dicts carrying at least "uuid", "name" and "row".
    Raises ValueError when Zee's own marks point at two rows at once — guessing
    between two deliberate marks would be worse than stopping."""
    overrides = overrides or {}
    by_uuid = {m["uuid"]: m for m in members}

    named = [u for u in overrides.values() if u in by_uuid]
    if len(named) > 1:
        raise ValueError("more than one survivor named for the same number: %s"
                         % ", ".join(named))
    if named:
        win = by_uuid[named[0]]
        return win, [m for m in members if m["uuid"] != win["uuid"]], "you named it"

    marked = [m for m in members if m["uuid"] in green]
    if len(marked) > 1:
        raise ValueError("%d rows in this group are marked green: %s. Green marks the ONE "
                         "name that survives, so leave green on exactly one."
                         % (len(marked), ", ".join(m["name"] for m in marked)))
    if marked:
        return marked[0], [m for m in members if m["uuid"] != marked[0]["uuid"]], "green"

    ranked = []
    for m in members:
        facts, bundles = content_score(m["uuid"], m["row"]["root"], home)
        ranked.append((facts + bundles * 3, facts, bundles, m))
    ranked.sort(key=lambda t: (t[0], t[1], t[2]), reverse=True)
    top = ranked[0]
    if len(ranked) > 1 and ranked[1][0] == top[0]:
        raise ValueError("no green mark, and these rows hold the same amount, so nothing "
                         "chooses between them: %s. Mark one green, or name it."
                         % ", ".join(m["name"] for _s, _f, _b, m in ranked if _s == top[0]))
    win = top[3]
    return (win, [m for m in members if m["uuid"] != win["uuid"]],
            "holds the most (%d facts, %d bundles) — WEAK, nothing was marked"
            % (top[1], top[2]))


def transfer_bundles(loser_row, survivor_row, home=None, dry=True,
                     group_uuids=None):
    """Hand a loser's PROVEN close bundles to the survivor.

    Re-stamping alone is not enough when the roots differ: owned_bundles() only
    scans a row's own root, so a re-stamped bundle sitting under the loser's
    folder would belong to the survivor on paper and be invisible in practice.
    So the directory moves too. A name already taken at the destination gets a
    numbered suffix rather than overwriting a real dossier."""
    manage = _load("manage_ordinals", "manage-ordinals.py")
    proven, guessed = manage.owned_bundles(loser_row, home)
    if group_uuids:
        # Guesses the merge itself settles travel too, and get stamped on the
        # way. Otherwise the merge would leave behind the very bundle it just
        # made unambiguous.
        rescued, guessed = resolvable_guesses(loser_row, group_uuids, home)
        proven = list(proven) + rescued
    same_root = loser_row["root_casefold"] == survivor_row["root_casefold"]
    dest_dir = os.path.join(survivor_row["root"], "memory", "handoffs", "closed")

    moved = []
    for bundle, _evidence in proven:
        target = bundle
        if not same_root:
            target = os.path.join(dest_dir, os.path.basename(bundle))
            n = 2
            while os.path.exists(target):
                target = os.path.join(dest_dir, "%s--%d" % (os.path.basename(bundle), n))
                n += 1
        if not dry:
            if not same_root:
                os.makedirs(dest_dir, exist_ok=True)
                shutil.move(bundle, target)
            with open(os.path.join(target, bundles_lib.OWNER_MARKER), "w",
                      encoding="utf-8") as fh:
                fh.write(survivor_row["project_uuid"] + "\n")
        moved.append((bundle, target))
    return moved, guessed


def transfer_windows(loser_uuid, survivor_uuid, home=None, dry=True):
    """Move every window claim from the loser to the survivor.

    A window claim is one JSON file naming the cmux workspace (the sidebar entry
    a tab lives in) that is open on a row. Claims are stored in a folder named
    after the project_uuid, so handing them over means moving the files. A tab
    open on the losing row therefore keeps resolving after the merge instead of
    pointing at a row the book no longer lists."""
    claims = windows_lib.all_claims(loser_uuid, home)
    moved = []
    for entry in claims:
        ws = entry.get("workspace_id") or entry.get("ws_id") or ""
        src_path = windows_lib._entry_path(loser_uuid, ws, home)
        dst_path = windows_lib._entry_path(survivor_uuid, ws, home)
        moved.append(ws or "?")
        if dry:
            continue
        os.makedirs(os.path.dirname(dst_path), exist_ok=True)
        entry["project_uuid"] = survivor_uuid
        entry["merged_from_project"] = loser_uuid
        windows_lib._atomic_write(dst_path, entry)
        try:
            os.remove(src_path)
        except OSError:
            pass
    return moved


def merge_knowledge(loser_uuid, survivor_uuid, home=None, dry=True):
    """Move facts through merge-knowledge.py rather than re-implementing it.

    Returns (would_move, already_there). --allow-different-roots is passed
    because a merge group is by definition rows Zee has declared to be one
    project; the root check is that script's guard for a LONE merge, and here
    the declaration is the sheet itself."""
    src = knowledge_lib.load_facts(loser_uuid, home)
    have = {f["id"] for f in knowledge_lib.load_facts(survivor_uuid, home)}
    new = [f for f in src if f["id"] not in have]
    if dry or not new:
        return len(new), len(src) - len(new)

    mk = _load("merge_knowledge", "merge-knowledge.py")
    argv = ["--from", loser_uuid, "--into", survivor_uuid, "--apply",
            "--allow-different-roots"]
    import io as _io
    import contextlib
    buf = _io.StringIO()
    with contextlib.redirect_stdout(buf):
        rc = mk.main(argv)
    if rc not in (0, None):
        raise RuntimeError("merge-knowledge.py refused (%s):\n%s" % (rc, buf.getvalue()))
    return len(new), len(src) - len(new)


def name_claimants(bundle_dir, home=None):
    """Every live row whose NAME could claim this unstamped bundle.

    Deliberately blind to bundles_lib's shared-name refusal. That refusal says
    "I will not guess between these rows", which is the right answer while they
    are separate rows — but the question here is a different one: WHICH rows are
    even in the running? A merge that contains all of them settles it."""
    key = bundles_lib.slug_key(bundles_lib.slug_name(bundle_dir) or "")
    if not key:
        return []
    out = []
    for row in registry_lib._iter_rows(home):
        if bundles_lib.slug_key(row.get("name") or "") == key:
            out.append(row["project_uuid"])
    return out


def unstamped_bundles(row, home=None):
    """Unmarked close bundles under this row's folder that THIS ROW could claim.

    owned_bundles() reports the ones a row can claim outright. This reports the
    ones sitting there unclaimed, which is what a shared name produces: two rows
    both called ACOS 3.0, and bundles_lib refusing to award the bundle to either.

    The name test is not optional. Many rows share one folder — 'FruitSync
    (duplicate)', 'Website Research' and 'Research to Portfolio' are all rooted
    at ACOS 3.0 — so a bare directory scan would have every one of them claiming
    2026-07-18-ACOS-3.0-close, which belongs to none of them. Measured: that bug
    refused three of Zee's eight merges on 2026-08-25."""
    cdir = os.path.join(row["root"], "memory", "handoffs", "closed")
    out = []
    if not os.path.isdir(cdir):
        return out
    for name in sorted(os.listdir(cdir)):
        bundle = os.path.join(cdir, name)
        if not os.path.isdir(bundle):
            continue
        if os.path.exists(os.path.join(bundle, bundles_lib.OWNER_MARKER)):
            continue
        if row["project_uuid"] not in name_claimants(bundle, home):
            continue
        out.append(bundle)
    return out


def resolvable_guesses(loser_row, group_uuids, home=None):
    """(resolvable, still_doubtful) among a losing row's UNPROVEN bundles.

    A merge dissolves its own ambiguity. If every row that could claim a bundle
    is inside this merge group, then whichever of them owned it, the survivor
    owns it now — there is nothing left to be wrong about. Live case: the last
    unresolved bundle on this machine, 2026-07-18-ACOS-3.0-close, is claimed by
    two rows both named ACOS 3.0, and Zee gave both of them number 23.

    A bundle some row OUTSIDE the group could also claim stays doubtful, and
    still refuses the sheet.

    Two kinds of unproven bundle are covered. A GUESSED one is matched by name
    where the name is unique among live rows. A STRANDED one sits under the
    row's folder claimed by nobody, because bundles_lib refuses to award a
    bundle when two live rows share the name. Both become certain once the
    rows sharing that name are one row."""
    manage = _load("manage_ordinals", "manage-ordinals.py")
    _proven, guessed = manage.owned_bundles(loser_row, home)
    group = set(group_uuids)

    candidates = {b: e for b, e in guessed}
    for bundle in unstamped_bundles(loser_row, home):
        candidates.setdefault(bundle, "unclaimed — its name is shared by more than one row")

    resolvable, doubtful = [], []
    for bundle in sorted(candidates):
        who = set(name_claimants(bundle, home))
        if who and who <= group:
            resolvable.append((bundle, candidates[bundle]))
        else:
            doubtful.append((bundle, candidates[bundle]))
    return resolvable, doubtful
