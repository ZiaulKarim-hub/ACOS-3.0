#!/usr/bin/env python3
"""test_resurrection_merge.py — folding several rows of one project into one.

Zee filled the bulk-renumber sheet on 2026-08-25 and gave the SAME number to two
or more rows in eight places. That is an instruction, not a mistake: those rows
are one project wearing several rows, and they are to be joined. His rule for
which NAME survives is a GREEN fill on that row's number cell.

What these tests hold in place:
  - a survivor is chosen by an explicit name, then green, then content, in that
    order, and a content pick is reported as WEAK
  - two green marks on one number STOP the sheet rather than pick one
  - knowledge facts, close bundles and window claims all reach the survivor
  - a bundle whose ownership is only a GUESS refuses the whole sheet
  - the losing row is deleted only AFTER its content has moved

Every test runs against a FIXTURE registry under a throwaway home. Nothing here
reads or writes the real ~/.acos.
"""

import importlib.util
import os
import shutil
import sys
import tempfile
import unittest

_THIS = os.path.dirname(os.path.abspath(__file__))
_RESDIR = os.path.abspath(os.path.join(_THIS, os.pardir, "resurrection"))
sys.path.insert(0, _RESDIR)
import bundles_lib     # noqa: E402
import knowledge_lib   # noqa: E402
import merge_lib       # noqa: E402
import ordinal_lib     # noqa: E402
import registry_lib    # noqa: E402
import windows_lib     # noqa: E402


def _load(stem, filename):
    spec = importlib.util.spec_from_file_location(stem, os.path.join(_RESDIR, filename))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


plan = _load("plan_ordinals", "plan-ordinals.py")


class MergeTestBase(unittest.TestCase):
    def setUp(self):
        self.home = tempfile.mkdtemp(prefix="merge-test-")
        os.makedirs(registry_lib.registry_dir(self.home), exist_ok=True)
        self._prev = plan.REG_HOME
        plan.REG_HOME = self.home
        os.environ["ACOS_REGISTRY_HOME"] = self.home
        os.environ["RESURRECTION_SKIP_CMUX"] = "1"

    def tearDown(self):
        plan.REG_HOME = self._prev
        os.environ.pop("ACOS_REGISTRY_HOME", None)
        shutil.rmtree(self.home, ignore_errors=True)

    def mkrow(self, uuid, name, ordinal, status="parked", root=None):
        root = root or os.path.join(self.home, "roots", name)
        os.makedirs(root, exist_ok=True)
        st = os.stat(root)
        row = {
            "project_uuid": uuid, "root": root,
            "root_casefold": os.path.realpath(root).casefold(),
            "dev_ino": [st.st_dev, st.st_ino], "name": name,
            "workspace_name": name, "status": status,
            "enrolled_at": "2026-01-01T00:00:00+00:00",
            "last_verified_at": "2026-01-01T00:00:00+00:00",
            "last_close": None, "last_session_id_hint": None, "git": None,
            "tombstoned_at": None, "pick_ordinal": ordinal,
        }
        registry_lib.atomic_write_json(registry_lib.row_path(uuid, self.home), row)
        ordinal_lib.append_event("issue", ordinal, uuid, name, self.home)
        return row

    def member(self, uuid, frm=None, to=None):
        row = registry_lib.load_row(uuid, self.home)
        return {"row": row, "uuid": uuid, "name": row["name"],
                "from": frm if frm is not None else row["pick_ordinal"],
                "to": to, "line": None}

    def records(self, *triples):
        """(uuid, current, new) triples -> the dicts read_sheet would produce."""
        out = []
        for uuid, cur, new in triples:
            row = registry_lib.load_row(uuid, self.home)
            out.append({"name": row["name"], "current_number": cur, "new_number": new,
                        "status": row["status"], "tier": "RECENT", "root": row["root"],
                        "project_uuid": uuid})
        return out

    def mkfact(self, uuid, subject, claim):
        return knowledge_lib.append_fact(
            uuid, {"kind": "machine", "subject": subject, "claim": claim,
                   "evidence": {"type": "observation", "value": "fixture"},
                   "checks": [], "entities": [], "tags": [], "single_valued": False},
            home=self.home)

    def mkbundle(self, uuid, slug, owned=True):
        row = registry_lib.load_row(uuid, self.home)
        d = os.path.join(row["root"], "memory", "handoffs", "closed", slug)
        os.makedirs(d, exist_ok=True)
        with open(os.path.join(d, "handoff.yaml"), "w") as fh:
            fh.write("timestamp: x\nstatus: active\nnext_action: y\n")
        if owned:
            with open(os.path.join(d, bundles_lib.OWNER_MARKER), "w") as fh:
                fh.write(uuid + "\n")
        return d


# ------------------------------------------------------- picking the survivor

class SurvivorTests(MergeTestBase):
    def test_a_named_survivor_beats_everything(self):
        """Zee named cae643cb for the loan-intake group, where all three rows
        are empty and nothing else could choose."""
        self.mkrow("u-a", "Backup thing", 1)
        self.mkrow("u-b", "The real thing", 2)
        self.mkfact("u-a", "x", "a claim that would otherwise win")
        win, lose, why = merge_lib.pick_survivor(
            [self.member("u-a"), self.member("u-b")],
            overrides={26: "u-b"}, home=self.home)
        self.assertEqual(win["uuid"], "u-b")
        self.assertEqual([m["uuid"] for m in lose], ["u-a"])
        self.assertEqual(why, "you named it")

    def test_green_beats_content(self):
        """Green marks the surviving NAME. A row with less in it still wins if
        it is the name Zee wants."""
        self.mkrow("u-a", "Loud row", 1)
        self.mkrow("u-b", "Quiet row", 2)
        self.mkfact("u-a", "x", "one")
        self.mkfact("u-a", "y", "two")
        win, lose, why = merge_lib.pick_survivor(
            [self.member("u-a"), self.member("u-b")], green={"u-b"}, home=self.home)
        self.assertEqual(win["uuid"], "u-b")
        self.assertEqual(why, "green")

    def test_two_greens_on_one_number_stop_the_sheet(self):
        """Two deliberate marks contradict each other. Guessing between them
        would be worse than stopping and saying so."""
        self.mkrow("u-a", "Alpha", 1)
        self.mkrow("u-b", "Bravo", 2)
        with self.assertRaises(ValueError) as ctx:
            merge_lib.pick_survivor([self.member("u-a"), self.member("u-b")],
                                    green={"u-a", "u-b"}, home=self.home)
        self.assertIn("marked green", str(ctx.exception))

    def test_content_wins_when_nothing_is_marked_and_says_it_is_weak(self):
        self.mkrow("u-a", "Thin", 1)
        self.mkrow("u-b", "Fat", 2)
        self.mkfact("u-b", "x", "a real claim")
        self.mkbundle("u-b", "2026-01-01-Fat-close")
        win, _lose, why = merge_lib.pick_survivor(
            [self.member("u-a"), self.member("u-b")], home=self.home)
        self.assertEqual(win["uuid"], "u-b")
        self.assertIn("WEAK", why)

    def test_a_dead_heat_refuses_rather_than_flips_a_coin(self):
        """Zee's loan-intake group is exactly this: three empty rows. Silence
        here would delete two real folders' rows on a tie-break."""
        self.mkrow("u-a", "Backup x", 1)
        self.mkrow("u-b", "Clone x", 2)
        with self.assertRaises(ValueError) as ctx:
            merge_lib.pick_survivor([self.member("u-a"), self.member("u-b")],
                                    home=self.home)
        self.assertIn("nothing chooses between them", str(ctx.exception))

    def test_a_bundle_counts_for_more_than_a_single_fact(self):
        """A close bundle is a whole session's dossier, not one claim."""
        self.mkrow("u-a", "Two facts", 1)
        self.mkrow("u-b", "One bundle", 2)
        self.mkfact("u-a", "x", "one")
        self.mkfact("u-a", "y", "two")
        self.mkbundle("u-b", "2026-01-01-One-bundle-close")
        win, _lose, _why = merge_lib.pick_survivor(
            [self.member("u-a"), self.member("u-b")], home=self.home)
        self.assertEqual(win["uuid"], "u-b")


# ------------------------------------------------------------- reading green

class GreenTests(MergeTestBase):
    def test_the_greens_excel_offers_are_recognised(self):
        for rgb in ("FF00B050", "00B050", "FF92D050", "FFC6EFCE", "FF008000"):
            self.assertTrue(merge_lib.is_green(rgb), rgb)

    def test_the_sheets_own_yellow_edit_fill_is_not_green(self):
        """write_xlsx paints every editable cell FFF2CC. If that read as green,
        every row would claim to be the survivor."""
        self.assertFalse(merge_lib.is_green("FFFFF2CC"))
        self.assertFalse(merge_lib.is_green("FFF2CC"))

    def test_grey_and_white_are_not_green(self):
        for rgb in ("FFDDDDDD", "FFFFFFFF", "00000000", "FF808080"):
            self.assertFalse(merge_lib.is_green(rgb), rgb)

    def test_junk_is_not_green(self):
        for rgb in (None, "", "theme4", 7, "ZZZZZZ"):
            self.assertFalse(merge_lib.is_green(rgb), repr(rgb))

    def test_a_green_cell_is_read_back_out_of_a_workbook(self):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        path = os.path.join(self.home, "s.xlsx")
        plan.write_xlsx(path, [
            {"name": "Alpha", "current_number": 1, "new_number": 5, "status": "parked",
             "tier": "RECENT", "root": "/x", "project_uuid": "u-a"},
            {"name": "Bravo", "current_number": 2, "new_number": 5, "status": "parked",
             "tier": "RECENT", "root": "/y", "project_uuid": "u-b"}])
        wb = load_workbook(path)
        ws = wb.active
        ws.cell(row=3, column=3).fill = PatternFill("solid", fgColor="FF00B050")
        wb.save(path)
        self.assertEqual(merge_lib.sheet_green_uuids(path), {"u-b"})

    def test_a_csv_reports_no_green_instead_of_crashing(self):
        """CSV carries no colour. That is a reason to fall back to content, not
        a reason to fail."""
        path = os.path.join(self.home, "s.csv")
        plan.write_csv(path, [
            {"name": "Alpha", "current_number": 1, "new_number": 5, "status": "parked",
             "tier": "RECENT", "root": "/x", "project_uuid": "u-a"}])
        self.assertEqual(merge_lib.sheet_green_uuids(path), set())


# --------------------------------------------------------- moving the content

class AbsorbTests(MergeTestBase):
    def test_facts_reach_the_survivor(self):
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "keep", "already here")
        self.mkfact("u-b", "move", "should arrive")
        moved, already = merge_lib.merge_knowledge("u-b", "u-a", self.home, dry=False)
        self.assertEqual((moved, already), (1, 0))
        claims = {f["claim"] for f in knowledge_lib.load_facts("u-a", self.home)}
        self.assertIn("should arrive", claims)
        self.assertIn("already here", claims)

    def test_a_fact_both_rows_hold_is_not_duplicated(self):
        """A fact id is a hash of its subject and claim, so the same claim
        landing twice collapses to one."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "same", "one claim")
        self.mkfact("u-b", "same", "one claim")
        moved, already = merge_lib.merge_knowledge("u-b", "u-a", self.home, dry=False)
        self.assertEqual((moved, already), (0, 1))
        self.assertEqual(len(knowledge_lib.load_facts("u-a", self.home)), 1)

    def test_a_dry_run_moves_nothing(self):
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-b", "move", "should NOT arrive yet")
        moved, _ = merge_lib.merge_knowledge("u-b", "u-a", self.home, dry=True)
        self.assertEqual(moved, 1)
        self.assertEqual(knowledge_lib.load_facts("u-a", self.home), [])

    def test_a_bundle_at_the_same_root_is_restamped_where_it_lies(self):
        """Both Logo Builder rows point at one folder. Moving the directory
        there would be moving it onto itself."""
        root = os.path.join(self.home, "shared")
        self.mkrow("u-a", "Keeper", 1, root=root)
        self.mkrow("u-b", "Loser", 2, root=root)
        b = self.mkbundle("u-b", "2026-01-01-Loser-close")
        moved, guessed = merge_lib.transfer_bundles(
            registry_lib.load_row("u-b", self.home),
            registry_lib.load_row("u-a", self.home), self.home, dry=False)
        self.assertEqual(guessed, [])
        self.assertEqual(len(moved), 1)
        self.assertTrue(os.path.isdir(b), "the directory stayed put")
        with open(os.path.join(b, bundles_lib.OWNER_MARKER)) as fh:
            self.assertEqual(fh.read().strip(), "u-a")

    def test_a_bundle_at_a_different_root_moves_as_well_as_restamps(self):
        """owned_bundles() only ever looks under a row's own root. A re-stamp
        alone would make the history belong to the survivor on paper and be
        unreachable in practice."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        b = self.mkbundle("u-b", "2026-01-01-Loser-close")
        keeper = registry_lib.load_row("u-a", self.home)
        moved, _ = merge_lib.transfer_bundles(
            registry_lib.load_row("u-b", self.home), keeper, self.home, dry=False)
        self.assertFalse(os.path.isdir(b), "it left the loser's folder")
        dest = os.path.join(keeper["root"], "memory", "handoffs", "closed",
                            "2026-01-01-Loser-close")
        self.assertTrue(os.path.isdir(dest))
        with open(os.path.join(dest, bundles_lib.OWNER_MARKER)) as fh:
            self.assertEqual(fh.read().strip(), "u-a")
        self.assertEqual(moved[0][1], dest)

    def test_a_name_clash_at_the_destination_does_not_overwrite(self):
        """Two rows can both hold a bundle called 2026-01-01-X-close. Losing a
        real dossier to a name collision is not acceptable."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkbundle("u-a", "same-slug")
        self.mkbundle("u-b", "same-slug")
        keeper = registry_lib.load_row("u-a", self.home)
        moved, _ = merge_lib.transfer_bundles(
            registry_lib.load_row("u-b", self.home), keeper, self.home, dry=False)
        self.assertTrue(moved[0][1].endswith("same-slug--2"))
        closed = os.path.join(keeper["root"], "memory", "handoffs", "closed")
        self.assertEqual(sorted(os.listdir(closed)), ["same-slug", "same-slug--2"])

    def test_a_guessed_bundle_is_never_handed_over(self):
        """Ownership by resemblance is not a reason to move a project's history."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        b = self.mkbundle("u-b", "2026-01-01-Loser-close", owned=False)
        moved, guessed = merge_lib.transfer_bundles(
            registry_lib.load_row("u-b", self.home),
            registry_lib.load_row("u-a", self.home), self.home, dry=False)
        self.assertEqual(moved, [])
        self.assertTrue(os.path.isdir(b), "an unproven bundle was left alone")

    def test_a_dry_run_leaves_every_bundle_where_it_is(self):
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        b = self.mkbundle("u-b", "2026-01-01-Loser-close")
        moved, _ = merge_lib.transfer_bundles(
            registry_lib.load_row("u-b", self.home),
            registry_lib.load_row("u-a", self.home), self.home, dry=True)
        self.assertEqual(len(moved), 1)
        self.assertTrue(os.path.isdir(b))

    def test_a_window_claim_follows_the_row_it_was_open_on(self):
        """A tab open on the losing row must keep resolving after the merge."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        windows_lib.claim_window("u-b", "WS-123", label="Loser", home=self.home)
        moved = merge_lib.transfer_windows("u-b", "u-a", self.home, dry=False)
        self.assertEqual(moved, ["WS-123"])
        self.assertEqual([c["workspace_id"] for c in
                          windows_lib.all_claims("u-a", self.home)], ["WS-123"])
        self.assertEqual(windows_lib.all_claims("u-b", self.home), [])


# ------------------------------------------------------ the whole sheet path

class SheetMergeTests(MergeTestBase):
    def test_a_merge_group_carries_its_survivor_and_losers(self):
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "x", "content that decides it")
        _m, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home)
        self.assertEqual(len(merges), 1)
        self.assertEqual(merges[0]["survivor"]["uuid"], "u-a")
        self.assertEqual([L["uuid"] for L in merges[0]["losers"]], ["u-b"])

    def test_a_losing_rows_number_is_freed_for_the_same_sheet(self):
        """The loser is deleted in this pass, so the number it vacates is free
        for another row in the very same sheet."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkrow("u-c", "Third", 3)
        self.mkfact("u-a", "x", "content that decides it")
        _m, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7), ("u-c", 3, 2)), self.home)
        self.assertEqual(len(merges), 1)

    def test_an_unproven_bundle_an_outside_row_could_claim_refuses_the_sheet(self):
        """Same rule as a delete. A bulk run that applies most of itself leaves
        a state nobody can reason about. The doubt has to be REAL though — a
        third row named Loser is not in this merge, so folding the bundle in
        would hand it a history that may belong to that third row."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkrow("u-c", "Loser", 3)          # not in the merge group
        self.mkfact("u-a", "x", "content that decides it")
        self.mkbundle("u-b", "2026-01-01-Loser-close", owned=False)
        _m, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home)
        with self.assertRaises(plan.Refused) as ctx:
            plan.check_mergeable(merges, self.home)
        self.assertIn("stamp-bundle-owners.py", str(ctx.exception))

    def test_an_unproven_bundle_only_the_group_could_claim_is_allowed(self):
        """The mirror of the test above, and the reason it is not simply a
        blanket refusal: with no outside claimant the merge settles it."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "x", "content that decides it")
        self.mkbundle("u-b", "2026-01-01-Loser-close", owned=False)
        _m, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home)
        plan.check_mergeable(merges, self.home)      # must not raise

    def test_the_preview_names_what_would_move(self):
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "x", "content that decides it")
        self.mkfact("u-b", "y", "a fact that would travel")
        self.mkbundle("u-b", "2026-01-01-Loser-close")
        _m, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home,
            survivors={7: "u-a"})
        previews = plan.preview_merges(merges, self.home)
        text = plan.render_merges(previews)
        self.assertIn("Keeper", text)
        self.assertIn("Loser", text)
        self.assertIn("1 fact(s) move", text)
        self.assertIn("1 close bundle(s) hand over", text)
        # and it really was a preview
        self.assertEqual(len(knowledge_lib.load_facts("u-a", self.home)), 1)

    def test_applying_absorbs_then_deletes_and_frees_the_number(self):
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "x", "content that decides it")
        self.mkfact("u-b", "y", "a fact that must survive the delete")
        self.mkbundle("u-b", "2026-01-01-Loser-close")
        _m, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home,
            survivors={7: "u-a"})
        done, failed = plan.apply_merges(merges, self.home)
        self.assertEqual(failed, [])
        self.assertEqual(len(done), 1)

        claims = {f["claim"] for f in knowledge_lib.load_facts("u-a", self.home)}
        self.assertIn("a fact that must survive the delete", claims)
        keeper = registry_lib.load_row("u-a", self.home)
        dest = os.path.join(keeper["root"], "memory", "handoffs", "closed",
                            "2026-01-01-Loser-close")
        self.assertTrue(os.path.isdir(dest), "the bundle reached the survivor")
        self.assertIsNone(registry_lib.load_row("u-b", self.home), "the loser is gone")
        self.assertNotIn(2, ordinal_lib.held_ordinals(self.home), "number 2 is free")

    def test_the_survivor_still_moves_to_the_shared_number(self):
        """Absorbing is only half of it — the survivor must land on the number
        Zee typed."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "x", "content that decides it")
        moves, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home)
        plan.apply_merges(merges, self.home)
        plan.apply_plan([m for m in moves if m["uuid"] == "u-a"], self.home)
        self.assertEqual(registry_lib.load_row("u-a", self.home)["pick_ordinal"], 7)

    def test_a_named_survivor_settles_a_group_nothing_else_could(self):
        """The loan-intake group: three rows, all empty, three real folders."""
        self.mkrow("u-a", "Backup okoa-loan-intake-system", 1)
        self.mkrow("u-b", "Clone-okoa-loan-intake-system", 2)
        self.mkrow("u-c", "okoa-loan-intake-system", 3)
        recs = self.records(("u-a", 1, 26), ("u-b", 2, 26), ("u-c", 3, 26))
        with self.assertRaises(plan.Refused):
            plan.build_plan(recs, self.home)          # a dead heat, on purpose
        _m, _d, _u, _f, merges = plan.build_plan(recs, self.home, survivors={26: "u-c"})
        self.assertEqual(merges[0]["survivor"]["uuid"], "u-c")
        self.assertEqual(merges[0]["reason"], "you named it")
        self.assertEqual({L["uuid"] for L in merges[0]["losers"]}, {"u-a", "u-b"})

    def test_a_blank_cell_never_joins_a_merge_group(self):
        """The guard that saved row 34 on 2026-08-25. `zee` held 34 with an
        empty cell while another row was sent to 34. A merge needs the number
        typed on EVERY row in the group."""
        self.mkrow("u-a", "Mover", 1)
        self.mkrow("u-b", "Sitting quietly", 34)
        with self.assertRaises(plan.Refused) as ctx:
            plan.build_plan(self.records(("u-a", 1, 34), ("u-b", 34, "")), self.home)
        msg = str(ctx.exception)
        self.assertIn("collision rather than a merge", msg)
        self.assertIn("Sitting quietly", msg)

    def test_a_blank_holder_blocks_an_otherwise_valid_merge_group(self):
        """Two rows typed 9, and a third sits on 9 with a blank cell. The two
        typed rows do mean a merge, but the third was never marked for anything
        — so the sheet stops rather than fold in a row Zee did not name."""
        self.mkrow("u-a", "Typed one", 1)
        self.mkrow("u-b", "Typed two", 2)
        self.mkrow("u-c", "Untouched", 9)
        with self.assertRaises(plan.Refused) as ctx:
            plan.build_plan(self.records(("u-a", 1, 9), ("u-b", 2, 9), ("u-c", 9, "")),
                            self.home)
        self.assertIn("Untouched", str(ctx.exception))

class AmbiguousBundleTests(MergeTestBase):
    """A close bundle is "guessed" when more than one row could own it. That
    doubt normally refuses a merge, because handing one project's history to
    another on a resemblance is not acceptable. But a merge dissolves its OWN
    ambiguity: if every row that could claim the bundle is being folded into
    one, then whichever owned it, the survivor owns it now.

    Live case — the last unresolved bundle on this machine is
    2026-07-18-ACOS-3.0-close, claimed by two rows both named ACOS 3.0, and Zee
    gave both of them number 23."""

    def two_rows_one_name(self):
        root = os.path.join(self.home, "shared")
        self.mkrow("u-a", "ACOS 3.0", 1, root=root)
        self.mkrow("u-b", "ACOS 3.0", 2, root=root)
        return self.mkbundle("u-b", "2026-07-18-ACOS-3.0-close", owned=False)

    def test_a_guess_only_the_group_could_claim_is_resolvable(self):
        self.two_rows_one_name()
        rescued, doubtful = merge_lib.resolvable_guesses(
            registry_lib.load_row("u-b", self.home), ["u-a", "u-b"], self.home)
        self.assertEqual(len(rescued), 1)
        self.assertEqual(doubtful, [])

    def test_a_guess_an_outside_row_could_claim_stays_doubtful(self):
        """A third row with the same name is NOT in the merge, so the doubt is
        real and the sheet must still stop."""
        b = self.two_rows_one_name()
        self.mkrow("u-c", "ACOS 3.0", 3, root=os.path.dirname(os.path.dirname(
            os.path.dirname(os.path.dirname(b)))))
        rescued, doubtful = merge_lib.resolvable_guesses(
            registry_lib.load_row("u-b", self.home), ["u-a", "u-b"], self.home)
        self.assertEqual(rescued, [])
        self.assertEqual(len(doubtful), 1)

    def test_a_resolvable_guess_is_stamped_to_the_survivor(self):
        b = self.two_rows_one_name()
        moved, _ = merge_lib.transfer_bundles(
            registry_lib.load_row("u-b", self.home),
            registry_lib.load_row("u-a", self.home), self.home, dry=False,
            group_uuids=["u-a", "u-b"])
        self.assertEqual(len(moved), 1)
        with open(os.path.join(b, bundles_lib.OWNER_MARKER)) as fh:
            self.assertEqual(fh.read().strip(), "u-a")

    def test_without_the_group_the_guess_is_still_left_alone(self):
        """transfer_bundles called with no group is the plain case, and it must
        keep refusing to move anything it only guessed at."""
        b = self.two_rows_one_name()
        moved, _ = merge_lib.transfer_bundles(
            registry_lib.load_row("u-b", self.home),
            registry_lib.load_row("u-a", self.home), self.home, dry=False)
        self.assertEqual(moved, [])
        self.assertFalse(os.path.exists(os.path.join(b, bundles_lib.OWNER_MARKER)))

    def test_the_whole_sheet_accepts_a_merge_that_settles_its_own_doubt(self):
        self.two_rows_one_name()
        self.mkfact("u-a", "x", "content that decides it")
        recs = []
        for uuid, cur in (("u-a", 1), ("u-b", 2)):
            row = registry_lib.load_row(uuid, self.home)
            recs.append({"name": row["name"], "current_number": cur, "new_number": 23,
                         "status": row["status"], "tier": "RECENT", "root": row["root"],
                         "project_uuid": uuid})
        _m, _d, _u, _f, merges = plan.build_plan(recs, self.home)
        plan.check_mergeable(merges, self.home)      # must not raise


    def test_a_shared_root_does_not_make_every_row_claim_every_bundle(self):
        """Many rows share one folder. A bare directory scan had all of them
        claiming 2026-07-18-ACOS-3.0-close, which belongs to none of them —
        that bug refused three of Zee's eight merges on 2026-08-25. A row only
        claims an unmarked bundle its NAME matches."""
        root = os.path.join(self.home, "shared")
        self.mkrow("u-a", "ACOS 3.0", 1, root=root)
        self.mkrow("u-b", "ACOS 3.0", 2, root=root)
        self.mkrow("u-c", "Website Research", 3, root=root)   # same folder, other name
        self.mkbundle("u-b", "2026-07-18-ACOS-3.0-close", owned=False)
        self.assertEqual(merge_lib.unstamped_bundles(
            registry_lib.load_row("u-c", self.home), self.home), [],
            "a row at the same folder must not claim another project's bundle")
        self.assertEqual(len(merge_lib.unstamped_bundles(
            registry_lib.load_row("u-b", self.home), self.home)), 1)

    def test_a_merged_away_row_is_dropped_from_the_move_list(self):
        """The bug that stopped a real run on 2026-08-25, on its second mover.

        A losing row is typed with the group's number, so it sits in `moves`
        alongside every other mover. The merge then deletes it. Feeding the
        original list to apply_plan raised KeyError partway through parking —
        after four deletes and twelve merges had already been written. The
        survivor must keep its own move; only the vanished row is dropped."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "x", "content that decides it")
        moves, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home)
        self.assertEqual(len(moves), 2, "both rows are movers before the merge")

        plan.apply_merges(merges, self.home)
        self.assertIsNone(registry_lib.load_row("u-b", self.home))

        alive = [m for m in moves
                 if registry_lib.load_row(m["uuid"], self.home) is not None]
        self.assertEqual([m["uuid"] for m in alive], ["u-a"])
        problems = plan.apply_plan(alive, self.home)
        self.assertEqual(problems, [])
        self.assertEqual(registry_lib.load_row("u-a", self.home)["pick_ordinal"], 7)

    def test_the_old_move_list_would_have_raised(self):
        """Names the failure directly, so the filter above cannot be dropped
        later without a test going red."""
        self.mkrow("u-a", "Keeper", 1)
        self.mkrow("u-b", "Loser", 2)
        self.mkfact("u-a", "x", "content that decides it")
        moves, _d, _u, _f, merges = plan.build_plan(
            self.records(("u-a", 1, 7), ("u-b", 2, 7)), self.home)
        plan.apply_merges(merges, self.home)
        with self.assertRaises(KeyError):
            plan.apply_plan(moves, self.home)


if __name__ == "__main__":
    unittest.main(verbosity=1)
