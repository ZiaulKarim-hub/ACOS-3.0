#!/usr/bin/env python3
"""xl_update.py — the DETERMINISTIC Excel engine for the acos-xl-update skill.

The skill automates OKOA's weekly XL ("XL Ant") portfolio update. This module owns everything
mechanical and repeatable — file duplication, date roll-forward, writing verified payoff numbers,
and writing the (agent-supplied) progress / blocker bullets — while the SKILL.md protocol owns the
JUDGMENT: fetching each payoff from acos-hypercore-ask (provenance-bound, never guessed) and
drafting each loan's update text from acos-fireflies-ask (diplomatic, carry-forward-aware).

Cardinal rules enforced here by construction:
  * NON-DESTRUCTIVE — only ever writes the NEW duplicated file; the prior weeks are never touched.
  * NO NUMBER IS INVENTED — `apply` writes only the payoff values it is GIVEN (the caller sourced
    them from Hypercore or is deliberately leaving a cell unchanged). This engine computes no
    financial figure itself except the explicit Lux II accrual (current balance + N × per-diem),
    where both inputs are supplied.
  * APPEARANCE-SAFE — formulas and cell formatting are left intact. Rather than recalculating via
    a LibreOffice round-trip (which can alter appearance), it sets `fullCalcOnLoad` so Excel /
    Sheets recalc the date-driven formulas (e.g. Riverdale's payoff) on open.
  * LABEL-ANCHORED — payoff / progress / blocker cells are LOCATED by their row labels per sheet
    (row offsets differ between tabs), never hardcoded.

Subcommands:
  prepare  --folder DIR [--weeks 1] [--date YYYYMMDD]   duplicate latest -> +N weeks, set date
  manifest --file XLSX                                   inspect a file (per-sheet cells + current)
  apply    --file XLSX --spec SPEC.json                  write payoffs + bullets from a spec
  verify   --file XLSX                                   post-write sanity checks

Python 3 stdlib + openpyxl only.
"""

from __future__ import annotations

import argparse
import datetime
import json
import os
import re
import shutil
import sys

import openpyxl


# ---------------------------------------------------------------------------
# Output location
# ---------------------------------------------------------------------------

# The skill READS the latest published workbook from the Dropbox series (passed via --folder)
# but always WRITES the new draft here — never back into the live Dropbox series. This keeps the
# published series clean and enforces "produce a draft for review" at the filesystem level.
# Override per-run with `prepare --out-folder DIR`.
DEFAULT_OUT_FOLDER = "/Users/zee/Documents/OKOA/XL Ant Weekly Update Draft"


# ---------------------------------------------------------------------------
# Filename / date handling
# ---------------------------------------------------------------------------

# "XL Ant Portfolio Update 20260628.xlsx" -> (stem, "20260628", ".xlsx")
_FNAME_RE = re.compile(r"^(?P<stem>.*?)(?P<date>\d{8})(?P<ext>\.xlsx)$", re.IGNORECASE)
_DATE_FMT = "%Y%m%d"


def _is_report_file(name: str) -> bool:
    # a dated report file, excluding Excel lock/temp files ("~$...") and non-report xlsx
    return bool(_FNAME_RE.match(name)) and not name.startswith("~$")


def find_latest(folder: str) -> str:
    """Path of the most recent dated report file in `folder` (by the trailing YYYYMMDD)."""
    best = None  # (date, name)
    for name in os.listdir(folder):
        m = _FNAME_RE.match(name)
        if not m or name.startswith("~$"):
            continue
        try:
            d = datetime.datetime.strptime(m.group("date"), _DATE_FMT).date()
        except ValueError:
            continue
        if best is None or d > best[0]:
            best = (d, name)
    if best is None:
        raise FileNotFoundError(f"no dated 'XL … YYYYMMDD.xlsx' report found in {folder!r}")
    return os.path.join(folder, best[1])


def next_filename(latest_path: str, *, weeks: int = 1, explicit_date: str = None,
                  out_folder: str = None) -> tuple:
    """(new_path, new_date) for the roll-forward. Default = latest date + 7*weeks days.
    The new file is placed in `out_folder` (the draft folder) when given, otherwise next to
    the source file."""
    folder, name = os.path.split(latest_path)
    m = _FNAME_RE.match(name)
    if not m:
        raise ValueError(f"{name!r} does not end with a YYYYMMDD date")
    if explicit_date:
        new_d = datetime.datetime.strptime(explicit_date, _DATE_FMT).date()
    else:
        cur = datetime.datetime.strptime(m.group("date"), _DATE_FMT).date()
        new_d = cur + datetime.timedelta(days=7 * weeks)
    new_name = f"{m.group('stem')}{new_d.strftime(_DATE_FMT)}{m.group('ext')}"
    return os.path.join(out_folder or folder, new_name), new_d


# ---------------------------------------------------------------------------
# Label-anchored cell location (row offsets differ between sheets)
# ---------------------------------------------------------------------------

DATE_INPUT_SHEET = "Riverdale"   # the ONE date input; every other sheet's B3 = =Riverdale!B3
DATE_INPUT_CELL = "B3"
_PAYOFF_LABEL = "current payoff amount:"
_PROGRESS_LABEL = "progress made this week:"
_KEY_ISSUES_MARK = "key issues"          # "Key Issues / Blockers This Week:"
_NEXT_STEPS_MARK = "next steps"          # "NEXT STEPS & ACTION ITEMS" — ends the blocker section
_BULLET = "•"


def _label_row(ws, needle: str, *, exact: bool = True) -> int:
    needle = needle.lower()
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                v = c.value.strip().lower()
                if (v == needle) if exact else (needle in v):
                    return c.row
    return None


def _payoff_cell(ws) -> str:
    r = _label_row(ws, _PAYOFF_LABEL)
    return f"B{r}" if r else None


def _bullet_slots(ws, start_row: int, end_row: int) -> list:
    """B-column cells of the rows carrying a '•' marker in column A, between two label rows."""
    slots = []
    for r in range(start_row + 1, end_row):
        if isinstance(ws[f"A{r}"].value, str) and ws[f"A{r}"].value.strip() == _BULLET:
            slots.append(f"B{r}")
    return slots


def _section_slots(ws) -> dict:
    """{'progress': [B-cells], 'blockers': [B-cells]} located by label, per sheet."""
    prog = _label_row(ws, _PROGRESS_LABEL)
    key = _label_row(ws, _KEY_ISSUES_MARK, exact=False)
    nxt = _label_row(ws, _NEXT_STEPS_MARK, exact=False)
    out = {"progress": [], "blockers": []}
    if prog and key:
        out["progress"] = _bullet_slots(ws, prog, key)
    if key and nxt:
        out["blockers"] = _bullet_slots(ws, key, nxt)
    return out


def _current_bullets(ws, slots: list) -> list:
    return [ws[c].value.strip() for c in slots
            if isinstance(ws[c].value, str) and ws[c].value.strip()]


# ---------------------------------------------------------------------------
# manifest — inspect a file so the caller knows what to fill
# ---------------------------------------------------------------------------

def manifest(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws, wsv = wb[name], wbv[name]
        pc = _payoff_cell(ws)
        payoff_formula = isinstance(ws[pc].value, str) and ws[pc].value.startswith("=") if pc else False
        sec = _section_slots(ws)
        sheets.append({
            "name": name,
            "payoff_cell": pc,
            "payoff_is_formula": bool(payoff_formula),
            "payoff_current": (wsv[pc].value if pc else None),
            "progress_slots": sec["progress"],
            "progress_current": _current_bullets(ws, sec["progress"]),
            "blocker_slots": sec["blockers"],
            "blocker_current": _current_bullets(ws, sec["blockers"]),
        })
    date_val = wbv[DATE_INPUT_SHEET][DATE_INPUT_CELL].value if DATE_INPUT_SHEET in wb.sheetnames else None
    return {"file": path, "date_input": f"{DATE_INPUT_SHEET}!{DATE_INPUT_CELL}",
            "report_date": str(date_val), "sheets": sheets}


# ---------------------------------------------------------------------------
# prepare — duplicate the latest file, roll the date forward
# ---------------------------------------------------------------------------

def prepare(folder: str, *, weeks: int = 1, explicit_date: str = None, overwrite: bool = False,
            out_folder: str = None) -> dict:
    latest = find_latest(folder)
    out_dir = out_folder or DEFAULT_OUT_FOLDER
    os.makedirs(out_dir, exist_ok=True)      # write the draft here, never back into the source series
    new_path, new_date = next_filename(latest, weeks=weeks, explicit_date=explicit_date,
                                       out_folder=out_dir)
    if os.path.exists(new_path) and not overwrite:
        raise FileExistsError(f"target already exists: {new_path} (pass --overwrite to replace)")
    shutil.copy2(latest, new_path)           # copy2 preserves metadata; originals untouched
    wb = openpyxl.load_workbook(new_path, data_only=False)
    if DATE_INPUT_SHEET not in wb.sheetnames:
        raise KeyError(f"expected a {DATE_INPUT_SHEET!r} sheet to hold the report date")
    ws = wb[DATE_INPUT_SHEET]
    # set the single date input as a real date (format already on the cell); every other sheet
    # links to it via =Riverdale!B3, so all report dates roll forward together.
    ws[DATE_INPUT_CELL] = datetime.datetime(new_date.year, new_date.month, new_date.day)
    # force Excel/Sheets to recalc the date-driven formulas (e.g. Riverdale payoff) on open —
    # keeps formulas + formatting intact (no LibreOffice round-trip needed for correctness).
    wb.calculation.fullCalcOnLoad = True
    wb.save(new_path)
    man = manifest(new_path)
    man["copied_from"] = latest
    man["output_folder"] = out_dir
    man["report_date_mmddyyyy"] = new_date.strftime("%m/%d/%Y")
    return man


# ---------------------------------------------------------------------------
# apply — write payoffs + bullets from a spec
# ---------------------------------------------------------------------------

def apply_spec(path: str, spec: dict) -> dict:
    """Write a spec into the file. Spec shape:
        {"sheets": {"<sheet>": {"payoff": <number|null>,
                                 "progress": ["bullet", ...],   # optional; replaces the section
                                 "blockers": ["bullet", ...]}}}  # optional; replaces the section
    `payoff` null/absent => leave the cell as-is (e.g. formula-driven Riverdale, or no-change tabs).
    A progress/blockers list REPLACES that section: existing bullet cells are cleared, then the new
    bullets are written top-down into the available '•' slots (extra bullets beyond the slot count
    are dropped and reported). Returns a per-sheet change report."""
    wb = openpyxl.load_workbook(path, data_only=False)
    report = {"file": path, "sheets": {}, "warnings": []}
    for sheet, s in (spec.get("sheets") or {}).items():
        if sheet not in wb.sheetnames:
            report["warnings"].append(f"sheet {sheet!r} not in workbook — skipped")
            continue
        ws = wb[sheet]
        rec = {}
        # --- payoff ---
        if "payoff" in s and s["payoff"] is not None:
            pc = _payoff_cell(ws)
            if not pc:
                report["warnings"].append(f"{sheet}: no 'Current Payoff Amount:' label found")
            elif isinstance(ws[pc].value, str) and ws[pc].value.startswith("="):
                report["warnings"].append(
                    f"{sheet}: payoff cell {pc} is a FORMULA — refusing to overwrite it with a "
                    f"literal (it recalculates from the report date). Value NOT written.")
            else:
                ws[pc] = float(s["payoff"])
                rec["payoff"] = {"cell": pc, "value": float(s["payoff"])}
        # --- progress / blockers ---
        sec = _section_slots(ws)
        for key in ("progress", "blockers"):
            if key not in s or s[key] is None:
                continue
            slots = sec[key]
            bullets = [b.strip() for b in s[key] if isinstance(b, str) and b.strip()]
            for c in slots:                       # clear existing bullets in this section
                ws[c] = None
            written = bullets[:len(slots)]
            for c, text in zip(slots, written):
                ws[c] = text
            rec[key] = {"slots": len(slots), "written": len(written)}
            if len(bullets) > len(slots):
                report["warnings"].append(
                    f"{sheet}: {len(bullets)} {key} bullets but only {len(slots)} slots — "
                    f"dropped {len(bullets) - len(slots)} (prioritize to the top {len(slots)}).")
        report["sheets"][sheet] = rec
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)
    return report


# ---------------------------------------------------------------------------
# verify — post-write sanity checks
# ---------------------------------------------------------------------------

def verify(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    issues = []
    # date input present and a real date
    dv = wb[DATE_INPUT_SHEET][DATE_INPUT_CELL].value if DATE_INPUT_SHEET in wb.sheetnames else None
    if not isinstance(dv, datetime.datetime):
        issues.append(f"{DATE_INPUT_SHEET}!{DATE_INPUT_CELL} is not a date: {dv!r}")
    # no obviously broken cells
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for c in row:
                if isinstance(c.value, str) and ("#REF!" in c.value or "#NAME?" in c.value):
                    issues.append(f"{name}!{c.coordinate} contains an error token: {c.value!r}")
    return {"file": path, "ok": not issues, "issues": issues,
            "full_calc_on_load": bool(getattr(wb.calculation, "fullCalcOnLoad", False))}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="Deterministic Excel engine for acos-xl-update.")
    sub = p.add_subparsers(dest="cmd", required=True)

    pp = sub.add_parser("prepare", help="duplicate latest report -> +N weeks and set the date")
    pp.add_argument("--folder", required=True, help="SOURCE folder holding the published dated series (read-only)")
    pp.add_argument("--weeks", type=int, default=1)
    pp.add_argument("--date", dest="date", default=None, help="explicit new date YYYYMMDD (overrides +weeks)")
    pp.add_argument("--out-folder", dest="out_folder", default=DEFAULT_OUT_FOLDER,
                    help=f"folder to write the new draft into (default: {DEFAULT_OUT_FOLDER})")
    pp.add_argument("--overwrite", action="store_true")

    pm = sub.add_parser("manifest", help="inspect a report file")
    pm.add_argument("--file", required=True)

    pa = sub.add_parser("apply", help="write payoffs + bullets from a spec json")
    pa.add_argument("--file", required=True)
    pa.add_argument("--spec", required=True, help="path to spec json (or '-' for stdin)")

    pv = sub.add_parser("verify", help="post-write sanity checks")
    pv.add_argument("--file", required=True)

    a = p.parse_args(argv)
    if a.cmd == "prepare":
        print(json.dumps(prepare(a.folder, weeks=a.weeks, explicit_date=a.date,
                                 overwrite=a.overwrite, out_folder=a.out_folder), indent=2, default=str))
    elif a.cmd == "manifest":
        print(json.dumps(manifest(a.file), indent=2, default=str))
    elif a.cmd == "apply":
        spec = json.load(sys.stdin) if a.spec == "-" else json.load(open(a.spec))
        print(json.dumps(apply_spec(a.file, spec), indent=2, default=str))
    elif a.cmd == "verify":
        res = verify(a.file)
        print(json.dumps(res, indent=2, default=str))
        return 0 if res["ok"] else 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
