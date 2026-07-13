#!/usr/bin/env python3
"""
PRISM Institutional Model Builder v1.1.0

Base class for creating Wall Street-compliant Excel financial models.
Enforces color coding, formula integrity, and institutional standards.

Usage:
    from institutional_model import InstitutionalModel
    
    model = InstitutionalModel("Deal Model")
    ws = model.create_sheet("Assumptions")
    model.set_input(ws, "C4", 1000000, model.CURRENCY)
    model.set_formula(ws, "C5", "=C4*1.08", model.CURRENCY)
    model.save("output.xlsx")
"""

from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, field, asdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Protection, 
    NamedStyle, Border, Side
)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule


@dataclass
class CellRecord:
    """Record of cell creation for audit trail."""
    sheet: str
    cell: str
    cell_type: str  # INPUT, FORMULA, LINK, EXTERNAL, HEADER, TOTAL
    value_or_formula: Any
    number_format: Optional[str] = None
    source: Optional[str] = None
    timestamp: Optional[str] = None


class InstitutionalModel:
    """
    Base class for institutional-grade Excel financial models.
    
    Features:
    - Wall Street color coding enforcement
    - Named style management
    - Named range registry
    - Cell creation audit trail (ledger)
    - Input/formula separation
    - Sheet protection management
    
    Color Coding (NON-NEGOTIABLE):
    - INPUT: Blue text (0000FF), yellow background, UNLOCKED
    - FORMULA: Black text (000000), no background, LOCKED
    - LINK: Green text (008000), cross-sheet references, LOCKED
    - EXTERNAL: Red text (FF0000), external workbook refs, LOCKED
    - HEADER: White text on dark blue background
    - TOTAL: Bold black on light blue background
    """
    
    # =========================================================================
    # COLOR CONSTANTS (Wall Street Standard)
    # =========================================================================
    INPUT_FONT = "0000FF"      # Blue
    FORMULA_FONT = "000000"    # Black
    LINK_FONT = "008000"       # Green
    EXTERNAL_FONT = "FF0000"   # Red
    HEADER_FONT = "FFFFFF"     # White
    
    INPUT_BG = "FFF2CC"        # Light yellow
    HEADER_BG = "1F4E79"       # Dark blue
    TOTAL_BG = "D9E1F2"        # Light blue
    CHECK_PASS_BG = "C6EFCE"   # Light green
    CHECK_FAIL_BG = "FFC7CE"   # Light red
    
    # =========================================================================
    # NUMBER FORMAT CONSTANTS
    # =========================================================================
    CURRENCY = '$#,##0;($#,##0);"-"'
    CURRENCY_DECIMAL = '$#,##0.00;($#,##0.00);"-"'
    CURRENCY_MILLIONS = '$#,##0.0,,"mm";($#,##0.0,,"mm");"-"'
    CURRENCY_THOUSANDS = '$#,##0,"k";($#,##0,"k");"-"'
    
    PERCENT = '0.0%;(0.0%);"-"'
    PERCENT_PRECISE = '0.00%;(0.00%);"-"'
    PERCENT_BPS = '0.0000%;(0.0000%);"-"'
    
    MULTIPLE = '0.0"x";(0.0"x");"-"'
    MULTIPLE_PRECISE = '0.00"x";(0.00"x");"-"'
    
    YEAR = '@'
    DATE = 'MM/DD/YYYY'
    DATE_SHORT = 'M/D/YY'
    
    INTEGER = '#,##0;(#,##0);"-"'
    DECIMAL_2 = '#,##0.00;(#,##0.00);"-"'
    
    BASIS_POINTS = '0 "bps";(0 "bps");"-"'
    
    def __init__(self, title: str = "Financial Model"):
        """
        Initialize institutional model.
        
        Args:
            title: Model title for metadata
        """
        self.wb = Workbook()
        self.title = title
        self.named_ranges: Dict[str, str] = {}
        self.cell_ledger: List[CellRecord] = []
        self.created_at = datetime.utcnow().isoformat() + "Z"
        
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        # Setup named styles
        self._setup_styles()
        
    def _setup_styles(self) -> None:
        """Create all named styles for institutional formatting."""
        
        # Common border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # INPUT style (Blue text, yellow background, UNLOCKED)
        input_style = NamedStyle(name="input")
        input_style.font = Font(color=self.INPUT_FONT)
        input_style.fill = PatternFill("solid", fgColor=self.INPUT_BG)
        input_style.protection = Protection(locked=False)
        input_style.border = thin_border
        self.wb.add_named_style(input_style)
        
        # FORMULA style (Black text, no fill, LOCKED)
        formula_style = NamedStyle(name="formula")
        formula_style.font = Font(color=self.FORMULA_FONT)
        formula_style.protection = Protection(locked=True)
        self.wb.add_named_style(formula_style)
        
        # LINK style (Green text, no fill, LOCKED)
        link_style = NamedStyle(name="link")
        link_style.font = Font(color=self.LINK_FONT)
        link_style.protection = Protection(locked=True)
        self.wb.add_named_style(link_style)
        
        # EXTERNAL style (Red text, no fill, LOCKED)
        external_style = NamedStyle(name="external")
        external_style.font = Font(color=self.EXTERNAL_FONT)
        external_style.protection = Protection(locked=True)
        self.wb.add_named_style(external_style)
        
        # HEADER style (White text, dark blue background, bold)
        header_style = NamedStyle(name="header")
        header_style.font = Font(color=self.HEADER_FONT, bold=True)
        header_style.fill = PatternFill("solid", fgColor=self.HEADER_BG)
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = thin_border
        self.wb.add_named_style(header_style)
        
        # TOTAL style (Bold black text, light blue background)
        total_style = NamedStyle(name="total")
        total_style.font = Font(color=self.FORMULA_FONT, bold=True)
        total_style.fill = PatternFill("solid", fgColor=self.TOTAL_BG)
        total_style.border = thin_border
        self.wb.add_named_style(total_style)
        
        # CHECK_PASS style (Green background)
        check_pass = NamedStyle(name="check_pass")
        check_pass.fill = PatternFill("solid", fgColor=self.CHECK_PASS_BG)
        self.wb.add_named_style(check_pass)
        
        # CHECK_FAIL style (Red background, bold red text)
        check_fail = NamedStyle(name="check_fail")
        check_fail.fill = PatternFill("solid", fgColor=self.CHECK_FAIL_BG)
        check_fail.font = Font(color=self.EXTERNAL_FONT, bold=True)
        self.wb.add_named_style(check_fail)
    
    def create_sheet(self, name: str, index: Optional[int] = None) -> Worksheet:
        """
        Create a new worksheet.
        
        Args:
            name: Sheet name
            index: Position (None = append)
            
        Returns:
            Created worksheet
        """
        ws = self.wb.create_sheet(name, index)
        return ws
    
    def set_input(
        self,
        ws: Worksheet,
        cell: str,
        value: Any,
        number_format: Optional[str] = None,
        source: Optional[str] = None
    ) -> None:
        """
        Set an input cell (BLUE text, yellow background, UNLOCKED).
        
        USE FOR: All hardcoded values on Assumptions sheet
        
        Args:
            ws: Worksheet
            cell: Cell reference (e.g., "C4")
            value: Cell value
            number_format: Optional format code
            source: Optional source reference for audit
        """
        cell_obj = ws[cell]
        cell_obj.value = value
        cell_obj.style = "input"
        
        if number_format:
            cell_obj.number_format = number_format
            
        self.cell_ledger.append(CellRecord(
            sheet=ws.title,
            cell=cell,
            cell_type="INPUT",
            value_or_formula=value,
            number_format=number_format,
            source=source,
            timestamp=datetime.utcnow().isoformat() + "Z"
        ))
    
    def set_formula(
        self,
        ws: Worksheet,
        cell: str,
        formula: str,
        number_format: Optional[str] = None
    ) -> None:
        """
        Set a formula cell (BLACK text, LOCKED).
        
        USE FOR: All calculations within the same sheet
        
        Args:
            ws: Worksheet
            cell: Cell reference
            formula: Excel formula (must start with '=')
            number_format: Optional format code
            
        Raises:
            ValueError: If formula doesn't start with '='
        """
        if not formula.startswith('='):
            raise ValueError(f"Formula must start with '=': {formula}")
            
        cell_obj = ws[cell]
        cell_obj.value = formula
        cell_obj.style = "formula"
        
        if number_format:
            cell_obj.number_format = number_format
            
        self.cell_ledger.append(CellRecord(
            sheet=ws.title,
            cell=cell,
            cell_type="FORMULA",
            value_or_formula=formula,
            number_format=number_format,
            timestamp=datetime.utcnow().isoformat() + "Z"
        ))
    
    def set_link(
        self,
        ws: Worksheet,
        cell: str,
        formula: str,
        number_format: Optional[str] = None
    ) -> None:
        """
        Set a cross-sheet link cell (GREEN text, LOCKED).
        
        USE FOR: References to cells on other sheets in same workbook
        
        Args:
            ws: Worksheet
            cell: Cell reference
            formula: Formula with sheet reference (must contain '!')
            number_format: Optional format code
            
        Raises:
            ValueError: If formula doesn't contain '!'
        """
        if '!' not in formula:
            raise ValueError(f"Cross-sheet link must contain '!': {formula}")
            
        cell_obj = ws[cell]
        cell_obj.value = formula
        cell_obj.style = "link"
        
        if number_format:
            cell_obj.number_format = number_format
            
        self.cell_ledger.append(CellRecord(
            sheet=ws.title,
            cell=cell,
            cell_type="LINK",
            value_or_formula=formula,
            number_format=number_format,
            timestamp=datetime.utcnow().isoformat() + "Z"
        ))
    
    def set_external(
        self,
        ws: Worksheet,
        cell: str,
        formula: str,
        number_format: Optional[str] = None
    ) -> None:
        """
        Set an external link cell (RED text, LOCKED).
        
        USE FOR: References to other workbooks
        WARNING: External links should be avoided in institutional models
        
        Args:
            ws: Worksheet
            cell: Cell reference
            formula: Formula with external reference (must contain '[')
            number_format: Optional format code
        """
        if '[' not in formula:
            raise ValueError(f"External link must contain '[': {formula}")
            
        cell_obj = ws[cell]
        cell_obj.value = formula
        cell_obj.style = "external"
        
        if number_format:
            cell_obj.number_format = number_format
            
        self.cell_ledger.append(CellRecord(
            sheet=ws.title,
            cell=cell,
            cell_type="EXTERNAL",
            value_or_formula=formula,
            number_format=number_format,
            timestamp=datetime.utcnow().isoformat() + "Z"
        ))
    
    def set_header(
        self,
        ws: Worksheet,
        cell: str,
        value: str
    ) -> None:
        """
        Set a header cell (white text, dark blue background).
        
        Args:
            ws: Worksheet
            cell: Cell reference
            value: Header text
        """
        cell_obj = ws[cell]
        cell_obj.value = value
        cell_obj.style = "header"
        
        self.cell_ledger.append(CellRecord(
            sheet=ws.title,
            cell=cell,
            cell_type="HEADER",
            value_or_formula=value,
            timestamp=datetime.utcnow().isoformat() + "Z"
        ))
    
    def set_total(
        self,
        ws: Worksheet,
        cell: str,
        formula: str,
        number_format: Optional[str] = None
    ) -> None:
        """
        Set a total/summary cell (bold, light blue background).
        
        Args:
            ws: Worksheet
            cell: Cell reference
            formula: Sum or total formula
            number_format: Optional format code
        """
        cell_obj = ws[cell]
        cell_obj.value = formula
        cell_obj.style = "total"
        
        if number_format:
            cell_obj.number_format = number_format
            
        self.cell_ledger.append(CellRecord(
            sheet=ws.title,
            cell=cell,
            cell_type="TOTAL",
            value_or_formula=formula,
            number_format=number_format,
            timestamp=datetime.utcnow().isoformat() + "Z"
        ))
    
    def set_text(
        self,
        ws: Worksheet,
        cell: str,
        value: str
    ) -> None:
        """
        Set a plain text cell (no special formatting).
        
        Args:
            ws: Worksheet
            cell: Cell reference
            value: Text value
        """
        ws[cell] = value
    
    def create_header_row(
        self,
        ws: Worksheet,
        row: int,
        headers: List[str],
        start_col: int = 1
    ) -> None:
        """
        Create a row of header cells.
        
        Args:
            ws: Worksheet
            row: Row number
            headers: List of header texts
            start_col: Starting column (1-indexed)
        """
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i)
            cell.value = header
            cell.style = "header"
    
    def add_named_range(
        self,
        name: str,
        sheet: str,
        cell: str
    ) -> None:
        """
        Add a named range for a cell.
        
        USE FOR: Key assumptions that need easy reference
        
        Args:
            name: Range name (no spaces)
            sheet: Sheet name
            cell: Cell reference
        """
        # Sanitize name
        safe_name = name.replace(" ", "_").replace("-", "_")
        
        # Create absolute reference
        ref = f"'{sheet}'!${cell}"
        
        # Add to workbook
        defn = DefinedName(name=safe_name, attr_text=ref)
        self.wb.defined_names.add(defn)
        
        # Track in registry
        self.named_ranges[safe_name] = ref
    
    def add_balance_check(
        self,
        ws: Worksheet,
        cell: str,
        formula: str,
        description: str
    ) -> None:
        """
        Add a balance check that must equal zero.
        
        USE FOR: Verification that model balances correctly
        
        Args:
            ws: Worksheet
            cell: Cell for check result
            formula: Formula that should = 0
            description: Description of what's being checked
        """
        # Set the check formula
        check_cell = ws[cell]
        check_cell.value = formula
        check_cell.style = "formula"
        check_cell.number_format = self.CURRENCY
        
        # Add conditional formatting
        # Green if zero (within tolerance)
        ws.conditional_formatting.add(
            cell,
            FormulaRule(
                formula=[f'ABS({cell})<0.01'],
                fill=PatternFill("solid", fgColor=self.CHECK_PASS_BG)
            )
        )
        
        # Red if not zero
        ws.conditional_formatting.add(
            cell,
            FormulaRule(
                formula=[f'ABS({cell})>=0.01'],
                fill=PatternFill("solid", fgColor=self.CHECK_FAIL_BG)
            )
        )
        
        # Add description in adjacent cell
        desc_col = get_column_letter(ws[cell].column + 1)
        desc_row = ws[cell].row
        ws[f"{desc_col}{desc_row}"] = description
        
        self.cell_ledger.append(CellRecord(
            sheet=ws.title,
            cell=cell,
            cell_type="CHECK",
            value_or_formula=formula,
            source=description,
            timestamp=datetime.utcnow().isoformat() + "Z"
        ))
    
    def set_column_widths(
        self,
        ws: Worksheet,
        widths: Dict[str, float]
    ) -> None:
        """
        Set column widths.
        
        Args:
            ws: Worksheet
            widths: Dict of {column_letter: width}
        """
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def set_row_height(
        self,
        ws: Worksheet,
        row: int,
        height: float
    ) -> None:
        """
        Set row height.
        
        Args:
            ws: Worksheet
            row: Row number
            height: Height in points
        """
        ws.row_dimensions[row].height = height
    
    def freeze_panes(
        self,
        ws: Worksheet,
        cell: str
    ) -> None:
        """
        Freeze panes at specified cell.
        
        Args:
            ws: Worksheet
            cell: Cell reference (rows above and columns left will be frozen)
        """
        ws.freeze_panes = cell
    
    def protect_sheet(
        self,
        ws: Worksheet,
        password: Optional[str] = None
    ) -> None:
        """
        Protect worksheet.
        
        Input cells (style="input") remain editable.
        Formula cells (style="formula", "link", etc.) are locked.
        
        Args:
            ws: Worksheet
            password: Optional protection password
        """
        ws.protection.sheet = True
        ws.protection.password = password
        ws.protection.enable()
    
    def get_ledger(self) -> List[Dict]:
        """
        Get cell creation ledger as list of dictionaries.
        
        Returns:
            List of cell records for audit trail
        """
        return [asdict(record) for record in self.cell_ledger]
    
    def save(self, path: str) -> None:
        """
        Save workbook to file.
        
        Args:
            path: Output file path
        """
        self.wb.save(path)
    
    def close(self) -> None:
        """Close workbook and release resources."""
        self.wb.close()


# Convenience function
def create_institutional_model(title: str = "Financial Model") -> InstitutionalModel:
    """
    Create a new institutional model.
    
    Args:
        title: Model title
        
    Returns:
        InstitutionalModel instance
    """
    return InstitutionalModel(title)


if __name__ == "__main__":
    # Example usage
    model = InstitutionalModel("Example Model")
    
    # Create Assumptions sheet
    ws_assumptions = model.create_sheet("Assumptions")
    model.set_header(ws_assumptions, "B2", "Key Assumptions")
    model.set_input(ws_assumptions, "C4", 1000000, model.CURRENCY, "Deal terms")
    model.set_input(ws_assumptions, "C5", 0.08, model.PERCENT, "Market rate")
    model.add_named_range("Principal", "Assumptions", "C4")
    model.add_named_range("Rate", "Assumptions", "C5")
    
    # Create Calculations sheet
    ws_calc = model.create_sheet("Calculations")
    model.set_link(ws_calc, "B4", "=Assumptions!$C$4", model.CURRENCY)
    model.set_formula(ws_calc, "C4", "=B4*1.08", model.CURRENCY)
    
    # Create Checks sheet
    ws_checks = model.create_sheet("Checks")
    model.add_balance_check(
        ws_checks, "C4",
        "=Calculations!C4-Calculations!B4*1.08",
        "Calculation verification"
    )
    
    # Save
    model.save("example_model.xlsx")
    print(f"Created model with {len(model.cell_ledger)} cells")
    model.close()
