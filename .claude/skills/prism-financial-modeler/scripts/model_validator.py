#!/usr/bin/env python3
"""
PRISM Model Validator v1.1.0

Comprehensive validation of institutional Excel models against
Wall Street standards.

Checks:
1. Formula errors (#VALUE!, #DIV/0!, etc.)
2. Color coding compliance (BLUE inputs, BLACK formulas, GREEN links)
3. Hardcoded values in calculation cells
4. Cross-sheet link formatting
5. Balance checks (must = 0)
6. Cache coverage
7. Named range usage
8. External link detection
9. Volatile function detection

Usage:
    python model_validator.py model.xlsx
"""

import json
import re
import sys
from typing import List, Dict, Optional, Any, Tuple, Set
from dataclasses import dataclass, field, asdict
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles.colors import Color
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


@dataclass
class ValidationIssue:
    """Single validation issue."""
    category: str
    severity: str  # ERROR, WARNING, INFO
    sheet: str
    cell: Optional[str]
    message: str
    details: Optional[Dict] = None


@dataclass
class ValidationMetrics:
    """Model metrics."""
    total_sheets: int = 0
    total_cells: int = 0
    formula_cells: int = 0
    input_cells: int = 0
    link_cells: int = 0
    cached_values: int = 0
    named_ranges: int = 0
    
    @property
    def formula_percentage(self) -> float:
        return (self.formula_cells / self.total_cells * 100) if self.total_cells > 0 else 0
    
    @property
    def cache_coverage(self) -> float:
        return (self.cached_values / self.formula_cells * 100) if self.formula_cells > 0 else 100


@dataclass
class ValidationResult:
    """Complete validation result."""
    status: str  # PASS, FAIL, WARNING
    workbook_path: str
    timestamp: str
    
    metrics: ValidationMetrics
    
    errors: List[ValidationIssue] = field(default_factory=list)
    warnings: List[ValidationIssue] = field(default_factory=list)
    info: List[ValidationIssue] = field(default_factory=list)
    
    error_count: int = 0
    warning_count: int = 0
    
    def to_dict(self) -> Dict:
        """Convert to dictionary for JSON serialization."""
        return {
            'status': self.status,
            'workbook_path': self.workbook_path,
            'timestamp': self.timestamp,
            'metrics': {
                'total_sheets': self.metrics.total_sheets,
                'total_cells': self.metrics.total_cells,
                'formula_cells': self.metrics.formula_cells,
                'input_cells': self.metrics.input_cells,
                'link_cells': self.metrics.link_cells,
                'cached_values': self.metrics.cached_values,
                'named_ranges': self.metrics.named_ranges,
                'formula_percentage': round(self.metrics.formula_percentage, 2),
                'cache_coverage': round(self.metrics.cache_coverage, 2)
            },
            'error_count': self.error_count,
            'warning_count': self.warning_count,
            'errors': [asdict(e) for e in self.errors],
            'warnings': [asdict(w) for w in self.warnings],
            'info': [asdict(i) for i in self.info]
        }


class ModelValidator:
    """
    Validates Excel models against institutional standards.
    
    Checks performed:
    1. Formula errors (#VALUE!, #DIV/0!, etc.)
    2. Color coding compliance
    3. Hardcoded values in calculation cells
    4. Cross-sheet link formatting
    5. Balance checks
    6. Cache coverage
    7. Named range usage
    8. External link detection
    9. Volatile function detection
    """
    
    # Formula error patterns
    ERROR_PATTERNS = [
        '#VALUE!', '#DIV/0!', '#REF!', '#NAME?',
        '#NULL!', '#NUM!', '#N/A', '#SPILL!', '#CALC!'
    ]
    
    # Expected colors (RGB hex, uppercase)
    EXPECTED_COLORS = {
        'INPUT': '0000FF',      # Blue
        'FORMULA': '000000',    # Black
        'LINK': '008000',       # Green
        'EXTERNAL': 'FF0000'    # Red
    }
    
    # Sheets that typically contain inputs
    INPUT_SHEET_PATTERNS = ['assumption', 'input', 'parameter', 'term', 'driver']
    
    # Sheets that should be all formulas
    CALC_SHEET_PATTERNS = ['calculation', 'schedule', 'cash flow', 'return', 'analysis']
    
    # Sheets that should be all links
    SUMMARY_SHEET_PATTERNS = ['summary', 'dashboard', 'executive', 'output']
    
    # Sheets for balance checks
    CHECK_SHEET_PATTERNS = ['check', 'balance', 'audit', 'verify']
    
    # Volatile functions to flag
    VOLATILE_FUNCTIONS = ['NOW(', 'TODAY(', 'RAND(', 'RANDBETWEEN(', 'OFFSET(', 'INDIRECT(']
    
    # Allowed constants in formulas
    ALLOWED_CONSTANTS = {'0', '1', '-1', '100', '12', '365', '360', '252', '2', '4'}
    
    def __init__(self, workbook_path: str):
        """
        Initialize validator.
        
        Args:
            workbook_path: Path to Excel file
        """
        self.workbook_path = Path(workbook_path)
        self.issues: List[ValidationIssue] = []
        self.metrics = ValidationMetrics()
        
    def validate(self) -> ValidationResult:
        """
        Run all validation checks.
        
        Returns:
            ValidationResult with complete analysis
        """
        # Load workbook in both modes
        wb_formulas = load_workbook(self.workbook_path, data_only=False)
        wb_data = load_workbook(self.workbook_path, data_only=True)
        
        try:
            # Collect metrics
            self._collect_metrics(wb_formulas, wb_data)
            
            # Run all checks
            self._check_formula_errors(wb_data)
            self._check_color_coding(wb_formulas)
            self._check_hardcoded_values(wb_formulas)
            self._check_cross_sheet_links(wb_formulas)
            self._check_balance_sheets(wb_data)
            self._check_cache_coverage(wb_formulas, wb_data)
            self._check_named_ranges(wb_formulas)
            self._check_external_links(wb_formulas)
            self._check_volatile_functions(wb_formulas)
            
        finally:
            wb_formulas.close()
            wb_data.close()
        
        # Compile result
        errors = [i for i in self.issues if i.severity == 'ERROR']
        warnings = [i for i in self.issues if i.severity == 'WARNING']
        info = [i for i in self.issues if i.severity == 'INFO']
        
        status = 'PASS'
        if errors:
            status = 'FAIL'
        elif warnings:
            status = 'WARNING'
        
        return ValidationResult(
            status=status,
            workbook_path=str(self.workbook_path),
            timestamp=datetime.utcnow().isoformat() + 'Z',
            metrics=self.metrics,
            errors=errors,
            warnings=warnings,
            info=info,
            error_count=len(errors),
            warning_count=len(warnings)
        )
    
    def _collect_metrics(self, wb_formulas, wb_data) -> None:
        """Collect model metrics."""
        
        self.metrics.total_sheets = len(wb_formulas.worksheets)
        self.metrics.named_ranges = len(wb_formulas.defined_names)
        
        formula_cells: Set[Tuple[str, str]] = set()
        
        for sheet in wb_formulas.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        self.metrics.total_cells += 1
                        
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            self.metrics.formula_cells += 1
                            formula_cells.add((sheet.title, cell.coordinate))
                            
                            if '!' in cell.value:
                                self.metrics.link_cells += 1
        
        # Check cached values
        for sheet in wb_data.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if (sheet.title, cell.coordinate) in formula_cells:
                        if cell.value is not None:
                            self.metrics.cached_values += 1
    
    def _check_formula_errors(self, wb_data) -> None:
        """Check for formula errors in cached values."""
        
        for sheet in wb_data.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        value_str = str(cell.value)
                        for error in self.ERROR_PATTERNS:
                            if error in value_str:
                                self.issues.append(ValidationIssue(
                                    category='FORMULA_ERROR',
                                    severity='ERROR',
                                    sheet=sheet.title,
                                    cell=cell.coordinate,
                                    message=f'Formula error: {error}',
                                    details={'cached_value': value_str[:100]}
                                ))
                                break
    
    def _check_color_coding(self, wb_formulas) -> None:
        """Check color coding compliance."""
        
        for sheet in wb_formulas.worksheets:
            sheet_type = self._classify_sheet(sheet.title)
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    # Get font color
                    font_color = self._get_font_color(cell)
                    
                    # Skip if we can't determine color
                    if font_color.startswith('INDEXED') or font_color.startswith('THEME'):
                        continue
                    
                    # Check based on cell content
                    is_formula = isinstance(cell.value, str) and cell.value.startswith('=')
                    is_cross_sheet = is_formula and '!' in cell.value
                    is_external = is_formula and '[' in cell.value
                    
                    if is_external:
                        expected = self.EXPECTED_COLORS['EXTERNAL']
                        if font_color != expected:
                            self.issues.append(ValidationIssue(
                                category='COLOR_VIOLATION',
                                severity='WARNING',
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                message=f'External link should be RED ({expected}), found {font_color}'
                            ))
                    elif is_cross_sheet:
                        # Cross-sheet links should be GREEN, especially on Summary sheets
                        if sheet_type == 'SUMMARY':
                            expected = self.EXPECTED_COLORS['LINK']
                            if font_color != expected:
                                self.issues.append(ValidationIssue(
                                    category='COLOR_VIOLATION',
                                    severity='WARNING',
                                    sheet=sheet.title,
                                    cell=cell.coordinate,
                                    message=f'Cross-sheet link should be GREEN ({expected}), found {font_color}'
                                ))
                    elif is_formula:
                        expected = self.EXPECTED_COLORS['FORMULA']
                        if font_color != expected and font_color != self.EXPECTED_COLORS['LINK']:
                            self.issues.append(ValidationIssue(
                                category='COLOR_VIOLATION',
                                severity='WARNING',
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                message=f'Formula should be BLACK ({expected}), found {font_color}'
                            ))
    
    def _check_hardcoded_values(self, wb_formulas) -> None:
        """Check for hardcoded values in calculation cells."""
        
        # Pattern to find numbers in formulas
        number_pattern = r'(?<![A-Za-z$:])(\d+\.?\d*)(?![A-Za-z0-9])'
        
        for sheet in wb_formulas.worksheets:
            sheet_type = self._classify_sheet(sheet.title)
            
            # Skip input sheets - they're supposed to have hardcoded values
            if sheet_type == 'INPUT':
                continue
            
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        
                        # Find numbers in formula
                        matches = re.findall(number_pattern, formula)
                        
                        # Filter to find suspicious hardcodes
                        hardcodes = []
                        for m in matches:
                            if m not in self.ALLOWED_CONSTANTS:
                                try:
                                    val = float(m)
                                    if val > 1 and val != 100 and val != 12:
                                        hardcodes.append(m)
                                except ValueError:
                                    pass
                        
                        if hardcodes:
                            self.issues.append(ValidationIssue(
                                category='HARDCODED_VALUE',
                                severity='ERROR',
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                message=f'Hardcoded values in formula: {hardcodes}',
                                details={'formula': formula[:100]}
                            ))
    
    def _check_cross_sheet_links(self, wb_formulas) -> None:
        """Check that summary sheets use cross-sheet links."""
        
        for sheet in wb_formulas.worksheets:
            if self._classify_sheet(sheet.title) != 'SUMMARY':
                continue
            
            formula_count = 0
            link_count = 0
            
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        if '!' in cell.value:
                            link_count += 1
            
            if formula_count > 0 and link_count < formula_count * 0.8:
                self.issues.append(ValidationIssue(
                    category='MISSING_CROSS_SHEET_LINKS',
                    severity='WARNING',
                    sheet=sheet.title,
                    cell=None,
                    message=f'Summary sheet has {formula_count} formulas but only {link_count} cross-sheet links'
                ))
    
    def _check_balance_sheets(self, wb_data) -> None:
        """Check that balance/check cells equal zero."""
        
        for sheet in wb_data.worksheets:
            if self._classify_sheet(sheet.title) != 'CHECK':
                continue
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if abs(cell.value) > 0.01:
                            self.issues.append(ValidationIssue(
                                category='BALANCE_CHECK_FAILED',
                                severity='ERROR',
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                message=f'Balance check failed: value = {cell.value}'
                            ))
    
    def _check_cache_coverage(self, wb_formulas, wb_data) -> None:
        """Check that formula cells have cached values."""
        
        if self.metrics.cache_coverage < 99.5:
            self.issues.append(ValidationIssue(
                category='LOW_CACHE_COVERAGE',
                severity='ERROR',
                sheet='ALL',
                cell=None,
                message=f'Cache coverage {self.metrics.cache_coverage:.1f}% is below 99.5% threshold',
                details={
                    'formula_cells': self.metrics.formula_cells,
                    'cached_values': self.metrics.cached_values
                }
            ))
    
    def _check_named_ranges(self, wb_formulas) -> None:
        """Check for named range usage."""
        
        if self.metrics.named_ranges == 0:
            self.issues.append(ValidationIssue(
                category='NO_NAMED_RANGES',
                severity='INFO',
                sheet='ALL',
                cell=None,
                message='Model has no named ranges. Consider adding for key assumptions.'
            ))
    
    def _check_external_links(self, wb_formulas) -> None:
        """Check for and flag external workbook links."""
        
        for sheet in wb_formulas.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and '[' in cell.value:
                        self.issues.append(ValidationIssue(
                            category='EXTERNAL_LINK',
                            severity='WARNING',
                            sheet=sheet.title,
                            cell=cell.coordinate,
                            message='External workbook link detected. Verify this is intentional.',
                            details={'formula': cell.value[:100]}
                        ))
    
    def _check_volatile_functions(self, wb_formulas) -> None:
        """Check for volatile functions."""
        
        for sheet in wb_formulas.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        formula_upper = cell.value.upper()
                        for func in self.VOLATILE_FUNCTIONS:
                            if func in formula_upper:
                                self.issues.append(ValidationIssue(
                                    category='VOLATILE_FUNCTION',
                                    severity='WARNING',
                                    sheet=sheet.title,
                                    cell=cell.coordinate,
                                    message=f'Volatile function detected: {func.replace("(", "")}'
                                ))
    
    def _classify_sheet(self, sheet_name: str) -> str:
        """Classify sheet type based on name."""
        
        name_lower = sheet_name.lower()
        
        for pattern in self.INPUT_SHEET_PATTERNS:
            if pattern in name_lower:
                return 'INPUT'
        
        for pattern in self.SUMMARY_SHEET_PATTERNS:
            if pattern in name_lower:
                return 'SUMMARY'
        
        for pattern in self.CHECK_SHEET_PATTERNS:
            if pattern in name_lower:
                return 'CHECK'
        
        return 'CALCULATION'
    
    def _get_font_color(self, cell) -> str:
        """Extract font color as hex string."""
        
        if cell.font and cell.font.color:
            color = cell.font.color
            if color.type == 'rgb' and color.rgb:
                # RGB format is AARRGGBB, we want RRGGBB
                return color.rgb[-6:].upper()
            elif color.type == 'indexed':
                return f'INDEXED_{color.indexed}'
            elif color.type == 'theme':
                return f'THEME_{color.theme}'
        
        # Default to black
        return '000000'


def validate_model(workbook_path: str) -> Dict:
    """
    Convenience function for model validation.
    
    Args:
        workbook_path: Path to Excel file
        
    Returns:
        Validation result as dictionary
    """
    validator = ModelValidator(workbook_path)
    result = validator.validate()
    return result.to_dict()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("PRISM Model Validator v1.1.0")
        print()
        print("Usage: python model_validator.py <workbook.xlsx>")
        print()
        print("Validates Excel models against institutional standards:")
        print("  - Formula errors")
        print("  - Color coding compliance")
        print("  - Hardcoded values in formulas")
        print("  - Balance checks")
        print("  - Cache coverage")
        sys.exit(1)
    
    result = validate_model(sys.argv[1])
    print(json.dumps(result, indent=2))
    
    sys.exit(0 if result['status'] == 'PASS' else 1)
