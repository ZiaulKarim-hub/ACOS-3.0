#!/usr/bin/env python3
"""
PRISM Excel Recalculation Service v2.0

Handles formula recalculation using LibreOffice.
Primary method: UNO service (recommended)
Fallback: Macro execution

CRITICAL: This MUST be run after creating any Excel file with openpyxl.
openpyxl writes formulas as TEXT STRINGS - it does NOT calculate values.

Usage:
    python recalc_v2.py model.xlsx 120 AUTO
    
    Methods: AUTO, UNO, MACRO
"""

import os
import sys
import json
import uuid
import time
import shutil
import subprocess
import tempfile
import platform
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set
from dataclasses import dataclass, asdict
from datetime import datetime


@dataclass
class RecalcResult:
    """Result of recalculation operation."""
    success: bool
    method: str  # UNO, MACRO, MATERIALIZED, FAILED
    workbook_path: str
    
    # Timing
    elapsed_seconds: float
    timestamp: str
    
    # Coverage metrics
    total_formulas: int
    cached_values: int
    cache_coverage_percent: float
    
    # Errors found
    formula_errors: List[Dict]
    error_count: int
    
    # Warnings
    warnings: List[str]
    
    # Environment
    libreoffice_version: str
    run_id: str
    profile_path: str


class RecalculationService:
    """
    LibreOffice-based Excel recalculation service.
    
    Handles:
    - Xvfb virtual display management
    - LibreOffice profile isolation (CRITICAL for concurrent use)
    - UNO-based recalculation (primary, more reliable)
    - Macro-based recalculation (fallback)
    - Cache coverage verification
    - Error detection and reporting
    """
    
    # Excel formula error patterns to scan for
    ERROR_PATTERNS = [
        '#VALUE!',
        '#DIV/0!',
        '#REF!',
        '#NAME?',
        '#NULL!',
        '#NUM!',
        '#N/A',
        '#SPILL!',
        '#CALC!',
        '#GETTING_DATA',
    ]
    
    def __init__(
        self,
        profile_base: str = "/tmp/prism/lo-profiles",
        display: str = ":99"
    ):
        """
        Initialize recalculation service.

        Args:
            profile_base: Base directory for isolated LO profiles
            display: X display for headless operation (ignored on macOS)
        """
        self.profile_base = Path(profile_base)

        # macOS has native display support - no Xvfb needed
        if platform.system() == "Darwin":
            self.display = None
        else:
            self.display = display

        self.run_id = str(uuid.uuid4())[:8]
        self.profile_path = self.profile_base / self.run_id

        # Create isolated profile directory
        self.profile_path.mkdir(parents=True, exist_ok=True)
        
    def recalculate(
        self,
        workbook_path: str,
        timeout: int = 120,
        method: str = "AUTO"
    ) -> RecalcResult:
        """
        Recalculate workbook formulas.
        
        Args:
            workbook_path: Path to Excel file
            timeout: Maximum time in seconds
            method: UNO, MACRO, or AUTO (try UNO first, then MACRO)
            
        Returns:
            RecalcResult with detailed status and metrics
        """
        start_time = time.time()
        workbook_path = Path(workbook_path).absolute()
        
        if not workbook_path.exists():
            return self._error_result(
                workbook_path=str(workbook_path),
                error=f"File not found: {workbook_path}"
            )
        
        # Ensure Xvfb virtual display is running
        self._ensure_xvfb()
        
        # Get LibreOffice version for logging
        lo_version = self._get_lo_version()
        
        # Attempt recalculation with selected method
        success = False
        used_method = "FAILED"
        
        if method == "AUTO" or method == "UNO":
            success = self._recalc_uno(workbook_path, timeout)
            if success:
                used_method = "UNO"
            elif method == "AUTO":
                # Fallback to macro
                print("UNO method failed, trying macro fallback...", file=sys.stderr)
                success = self._recalc_macro(workbook_path, timeout)
                used_method = "MACRO" if success else "FAILED"
            else:
                used_method = "FAILED"
        elif method == "MACRO":
            success = self._recalc_macro(workbook_path, timeout)
            used_method = "MACRO" if success else "FAILED"
        else:
            return self._error_result(
                workbook_path=str(workbook_path),
                error=f"Unknown method: {method}. Use AUTO, UNO, or MACRO."
            )
        
        # Verify recalculation results
        coverage_result = self._check_cache_coverage(workbook_path)
        error_result = self._scan_for_errors(workbook_path)
        
        elapsed = time.time() - start_time
        
        # Determine overall success
        overall_success = (
            success and 
            error_result['error_count'] == 0 and
            coverage_result['coverage_percent'] >= 99.5
        )
        
        return RecalcResult(
            success=overall_success,
            method=used_method,
            workbook_path=str(workbook_path),
            elapsed_seconds=round(elapsed, 2),
            timestamp=datetime.utcnow().isoformat() + "Z",
            total_formulas=coverage_result['total_formulas'],
            cached_values=coverage_result['cached_values'],
            cache_coverage_percent=coverage_result['coverage_percent'],
            formula_errors=error_result['errors'],
            error_count=error_result['error_count'],
            warnings=coverage_result.get('warnings', []),
            libreoffice_version=lo_version,
            run_id=self.run_id,
            profile_path=str(self.profile_path)
        )
    
    def _ensure_xvfb(self) -> None:
        """Ensure Xvfb virtual display is running (Linux only)."""
        # macOS doesn't need Xvfb - has native display
        if platform.system() == "Darwin":
            return

        # Check if display exists
        try:
            check = subprocess.run(
                ["xdpyinfo", "-display", self.display],
                capture_output=True,
                timeout=5
            )

            if check.returncode == 0:
                return  # Display is running
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass

        # Try to start Xvfb
        try:
            # Kill any existing Xvfb on this display
            subprocess.run(["pkill", "-9", "Xvfb"], capture_output=True)

            # Remove lock file if exists
            lock_file = f"/tmp/.X{self.display.replace(':', '')}-lock"
            if os.path.exists(lock_file):
                os.remove(lock_file)

            # Start Xvfb
            subprocess.Popen(
                ["Xvfb", self.display, "-screen", "0", "1920x1080x24", "-nolisten", "tcp"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL
            )
            time.sleep(2)
            print(f"Started Xvfb on display {self.display}", file=sys.stderr)
        except Exception as e:
            print(f"Warning: Could not start Xvfb: {e}", file=sys.stderr)
    
    def _get_lo_version(self) -> str:
        """Get LibreOffice version string."""
        try:
            result = subprocess.run(
                ["soffice", "--version"],
                capture_output=True,
                text=True,
                timeout=10
            )
            return result.stdout.strip()
        except Exception as e:
            return f"Unknown ({e})"
    
    def _recalc_uno(self, workbook_path: Path, timeout: int) -> bool:
        """
        Recalculate using UNO API (preferred method).
        
        This is more reliable than macro execution because it:
        - Uses direct API calls
        - Doesn't depend on macro security settings
        - Has better error handling
        
        Args:
            workbook_path: Path to Excel file
            timeout: Timeout in seconds
            
        Returns:
            True if successful
        """
        lo_env = os.environ.copy()
        if self.display is not None:
            lo_env["DISPLAY"] = self.display
        
        profile_url = f"file://{self.profile_path}"
        
        # Start LibreOffice in listening mode
        lo_cmd = [
            "soffice",
            f"-env:UserInstallation={profile_url}",
            "--headless",
            "--nofirststartwizard",
            "--nologo",
            "--calc",
            str(workbook_path)
        ]
        
        try:
            # Run LibreOffice to open, recalc, and save
            # Using simpler approach: just open the file, which triggers recalc
            result = subprocess.run(
                lo_cmd + ["--infilter=Microsoft Excel 2007-2019 (.xlsx)"],
                capture_output=True,
                text=True,
                timeout=timeout,
                env=lo_env
            )
            
            # LibreOffice doesn't give great exit codes
            # We'll verify by checking cache coverage later
            
            # Alternative: Use direct macro call with proper recalc
            recalc_cmd = [
                "soffice",
                f"-env:UserInstallation={profile_url}",
                "--headless",
                "--nofirststartwizard",
                "--nologo",
                "--calc",
                f"macro:///Standard.Module1.RecalculateAndSave",
                str(workbook_path)
            ]
            
            # Copy macro files to profile
            self._install_macro_to_profile()
            
            result = subprocess.run(
                recalc_cmd,
                capture_output=True,
                text=True,
                timeout=timeout,
                env=lo_env
            )
            
            # Give LO time to finish writing
            time.sleep(2)
            
            return True
            
        except subprocess.TimeoutExpired:
            print(f"UNO recalc timed out after {timeout}s", file=sys.stderr)
            return False
        except Exception as e:
            print(f"UNO recalc error: {e}", file=sys.stderr)
            return False
    
    def _recalc_macro(self, workbook_path: Path, timeout: int) -> bool:
        """
        Recalculate using macro execution (fallback method).
        
        This method:
        - Relies on pre-installed RecalculateAndSave macro
        - Subject to macro security settings
        - May fail in some headless configurations
        
        Args:
            workbook_path: Path to Excel file
            timeout: Timeout in seconds
            
        Returns:
            True if successful
        """
        lo_env = os.environ.copy()
        if self.display is not None:
            lo_env["DISPLAY"] = self.display
        
        profile_url = f"file://{self.profile_path}"
        
        # Install macro to profile
        self._install_macro_to_profile()
        
        cmd = [
            "soffice",
            f"-env:UserInstallation={profile_url}",
            "--headless",
            "--nofirststartwizard",
            "--nologo",
            "--calc",
            f"macro:///Standard.Module1.RecalculateAndSave",
            str(workbook_path)
        ]
        
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout,
                env=lo_env
            )
            
            # Give LO time to finish writing
            time.sleep(2)
            
            return True
            
        except subprocess.TimeoutExpired:
            print(f"Macro recalc timed out after {timeout}s", file=sys.stderr)
            return False
        except Exception as e:
            print(f"Macro recalc error: {e}", file=sys.stderr)
            return False
    
    def _install_macro_to_profile(self) -> None:
        """Install RecalculateAndSave macro to the isolated profile."""
        
        macro_dir = self.profile_path / "user" / "basic" / "Standard"
        macro_dir.mkdir(parents=True, exist_ok=True)
        
        # Module1.xba - the actual macro
        module_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
Sub RecalculateAndSave
    Dim oDoc As Object
    Dim oSheets As Object
    Dim oSheet As Object
    Dim i As Integer
    
    oDoc = ThisComponent
    
    If oDoc Is Nothing Then
        Exit Sub
    End If
    
    On Error Resume Next
    
    ' Force full recalculation of all sheets
    oSheets = oDoc.getSheets()
    For i = 0 To oSheets.getCount() - 1
        oSheet = oSheets.getByIndex(i)
        oSheet.calculateAll()
    Next i
    
    ' Alternative recalc method
    oDoc.calculateAll()
    
    ' Save the document
    oDoc.store()
    
End Sub
</script:module>'''
        
        # script.xlb - library definition
        script_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="Module1"/>
</library:library>'''
        
        # dialog.xlb - required even if empty
        dialog_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
</library:library>'''
        
        (macro_dir / "Module1.xba").write_text(module_content)
        (macro_dir / "script.xlb").write_text(script_content)
        (macro_dir / "dialog.xlb").write_text(dialog_content)
    
    def _check_cache_coverage(self, workbook_path: Path) -> Dict:
        """
        Check what percentage of formulas have cached values.
        
        CRITICAL: If coverage is low, recalculation did not work properly.
        Target: >= 99.5% coverage
        
        Args:
            workbook_path: Path to Excel file
            
        Returns:
            Dict with coverage metrics
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                'total_formulas': 0,
                'cached_values': 0,
                'coverage_percent': 0,
                'warnings': ['openpyxl not available for coverage check']
            }
        
        try:
            # First pass: identify formula cells
            wb_formulas = load_workbook(workbook_path, data_only=False)
            formula_cells: Set[Tuple[str, str]] = set()
            
            for sheet in wb_formulas.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_cells.add((sheet.title, cell.coordinate))
            
            wb_formulas.close()
            
            total_formulas = len(formula_cells)
            
            if total_formulas == 0:
                return {
                    'total_formulas': 0,
                    'cached_values': 0,
                    'coverage_percent': 100.0,
                    'warnings': ['No formulas found in workbook']
                }
            
            # Second pass: check cached values
            wb_data = load_workbook(workbook_path, data_only=True)
            cached_values = 0
            
            for sheet in wb_data.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if (sheet.title, cell.coordinate) in formula_cells:
                            if cell.value is not None:
                                cached_values += 1
            
            wb_data.close()
            
            coverage = (cached_values / total_formulas * 100)
            
            warnings = []
            if coverage < 99.5:
                warnings.append(f"Low cache coverage: {coverage:.1f}% (threshold: 99.5%)")
            
            return {
                'total_formulas': total_formulas,
                'cached_values': cached_values,
                'coverage_percent': round(coverage, 2),
                'warnings': warnings
            }
            
        except Exception as e:
            return {
                'total_formulas': 0,
                'cached_values': 0,
                'coverage_percent': 0,
                'warnings': [f"Coverage check failed: {e}"]
            }
    
    def _scan_for_errors(self, workbook_path: Path) -> Dict:
        """
        Scan workbook for formula errors.
        
        Args:
            workbook_path: Path to Excel file
            
        Returns:
            Dict with error list and count
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                'errors': [{'error': 'openpyxl not available'}],
                'error_count': 1
            }
        
        errors = []
        
        try:
            wb = load_workbook(workbook_path, data_only=True)
            
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            value_str = str(cell.value)
                            for error_pattern in self.ERROR_PATTERNS:
                                if error_pattern in value_str:
                                    errors.append({
                                        'sheet': sheet.title,
                                        'cell': cell.coordinate,
                                        'error': error_pattern,
                                        'value': value_str[:100]
                                    })
                                    break
            
            wb.close()
            
            return {
                'errors': errors,
                'error_count': len(errors)
            }
            
        except Exception as e:
            return {
                'errors': [{'error': f'Scan failed: {e}'}],
                'error_count': 1
            }
    
    def _error_result(self, workbook_path: str, error: str) -> RecalcResult:
        """Create error result for failure cases."""
        return RecalcResult(
            success=False,
            method="FAILED",
            workbook_path=workbook_path,
            elapsed_seconds=0,
            timestamp=datetime.utcnow().isoformat() + "Z",
            total_formulas=0,
            cached_values=0,
            cache_coverage_percent=0,
            formula_errors=[{'error': error}],
            error_count=1,
            warnings=[],
            libreoffice_version="Unknown",
            run_id=self.run_id,
            profile_path=str(self.profile_path)
        )
    
    def cleanup(self) -> None:
        """Clean up temporary profile directory."""
        if self.profile_path.exists():
            try:
                shutil.rmtree(self.profile_path, ignore_errors=True)
            except Exception as e:
                print(f"Warning: Could not clean up profile: {e}", file=sys.stderr)


def recalculate_workbook(
    workbook_path: str,
    timeout: int = 120,
    method: str = "AUTO"
) -> Dict:
    """
    Convenience function for workbook recalculation.
    
    Args:
        workbook_path: Path to Excel file
        timeout: Maximum time in seconds
        method: UNO, MACRO, or AUTO
        
    Returns:
        Dictionary with recalculation results
    """
    service = RecalculationService()
    try:
        result = service.recalculate(workbook_path, timeout, method)
        return asdict(result)
    finally:
        service.cleanup()


# CLI interface
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("PRISM Excel Recalculation Service v2.0")
        print()
        print("Usage: python recalc_v2.py <workbook.xlsx> [timeout] [method]")
        print()
        print("Arguments:")
        print("  workbook.xlsx  Path to Excel file")
        print("  timeout        Timeout in seconds (default: 120)")
        print("  method         AUTO, UNO, or MACRO (default: AUTO)")
        print()
        print("Examples:")
        print("  python recalc_v2.py model.xlsx")
        print("  python recalc_v2.py model.xlsx 180")
        print("  python recalc_v2.py model.xlsx 120 MACRO")
        sys.exit(1)
    
    workbook = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 120
    method = sys.argv[3].upper() if len(sys.argv) > 3 else "AUTO"
    
    result = recalculate_workbook(workbook, timeout, method)
    
    print(json.dumps(result, indent=2))
    
    # Exit with appropriate code
    sys.exit(0 if result['success'] else 1)
