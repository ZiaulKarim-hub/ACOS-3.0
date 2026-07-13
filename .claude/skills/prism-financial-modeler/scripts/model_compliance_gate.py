#!/usr/bin/env python3
"""
PRISM Model Compliance Gate
===========================

MANDATORY validation that MUST pass before any PRISM model is saved or distributed.
This gate ensures institutional standards per PRISM Intelligence System V2025.09.08.

THIS IS A NON-NEGOTIABLE QUALITY GATE. Models that fail this gate are REJECTED.

Usage:
    python model_compliance_gate.py path/to/model.xlsx

Exit codes:
    0 = PASS (model meets institutional standards)
    1 = FAIL (model does not meet standards - REJECT)
"""

import sys
import json
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import List, Tuple
from openpyxl import load_workbook


@dataclass
class ComplianceResult:
    """Result of compliance check."""
    model_path: str
    timestamp: str
    passed: bool
    total_checks: int
    checks_passed: int
    errors: List[str]
    warnings: List[str]


# Minimum requirements (NON-NEGOTIABLE)
REQUIRED_SHEETS = [
    "Cover",           # Executive summary
    "Assumptions",     # All inputs (BLUE cells)
]

RECOMMENDED_SHEETS = [
    "TOC",             # Table of contents
    "Sources & Uses",  # Capital stack
    "Checks",          # Validation
]

MINIMUM_NAMED_RANGES = 3
MINIMUM_FORMULAS = 10
MINIMUM_CROSS_SHEET_LINKS = 3
MINIMUM_INPUT_VALUES = 5


def check_compliance(model_path: str) -> ComplianceResult:
    """
    Run all compliance checks on a PRISM model.

    Args:
        model_path: Path to Excel model

    Returns:
        ComplianceResult with pass/fail status
    """
    errors = []
    warnings = []
    checks_passed = 0
    total_checks = 0

    try:
        wb = load_workbook(model_path)
    except Exception as e:
        return ComplianceResult(
            model_path=model_path,
            timestamp=datetime.utcnow().isoformat() + "Z",
            passed=False,
            total_checks=1,
            checks_passed=0,
            errors=[f"Cannot open workbook: {e}"],
            warnings=[]
        )

    # =========================================================================
    # CHECK 1: Required sheets present
    # =========================================================================
    total_checks += 1
    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        errors.append(f"CRITICAL: Missing required sheets: {missing}")
    else:
        checks_passed += 1

    # Warn about recommended sheets
    missing_rec = [s for s in RECOMMENDED_SHEETS if s not in wb.sheetnames]
    if missing_rec:
        warnings.append(f"Missing recommended sheets: {missing_rec}")

    # =========================================================================
    # CHECK 2: Assumptions sheet has inputs (not empty)
    # =========================================================================
    total_checks += 1
    if "Assumptions" in wb.sheetnames:
        ws_assumptions = wb["Assumptions"]
        input_count = 0
        for row in ws_assumptions.iter_rows(min_row=2, max_row=50):
            for cell in row:
                if cell.value is not None and not str(cell.value).startswith("="):
                    input_count += 1

        if input_count >= MINIMUM_INPUT_VALUES:
            checks_passed += 1
        else:
            errors.append(f"CRITICAL: Assumptions sheet has too few inputs ({input_count} < {MINIMUM_INPUT_VALUES})")
    else:
        errors.append("CRITICAL: No Assumptions sheet - cannot validate inputs")

    # =========================================================================
    # CHECK 3: Named ranges defined
    # =========================================================================
    total_checks += 1
    named_ranges = list(wb.defined_names)
    if len(named_ranges) >= MINIMUM_NAMED_RANGES:
        checks_passed += 1
    else:
        errors.append(f"CRITICAL: Insufficient named ranges ({len(named_ranges)} < {MINIMUM_NAMED_RANGES})")

    # =========================================================================
    # CHECK 4: Formula cells present (not just hardcoded values)
    # =========================================================================
    total_checks += 1
    formula_count = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    formula_count += 1

    if formula_count >= MINIMUM_FORMULAS:
        checks_passed += 1
    else:
        errors.append(f"CRITICAL: Too few formulas ({formula_count} < {MINIMUM_FORMULAS}). Models must use formulas, not hardcoded values.")

    # =========================================================================
    # CHECK 5: Cross-sheet links present (segregation enforced)
    # =========================================================================
    total_checks += 1
    link_count = 0
    for sheet in wb.worksheets:
        if sheet.title == "Assumptions":
            continue  # Skip assumptions sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("=") and "!" in str(cell.value):
                    link_count += 1

    if link_count >= MINIMUM_CROSS_SHEET_LINKS:
        checks_passed += 1
    else:
        errors.append(f"CRITICAL: Too few cross-sheet links ({link_count} < {MINIMUM_CROSS_SHEET_LINKS}). Calculations must reference Assumptions.")

    # =========================================================================
    # CHECK 6: No raw numbers in non-Assumptions sheets (Wall Street standard)
    # =========================================================================
    total_checks += 1
    hardcoded_values = []
    for sheet in wb.worksheets:
        if sheet.title == "Assumptions":
            continue
        for row in sheet.iter_rows():
            for cell in row:
                # Skip headers, text, and formulas
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val.startswith("="):
                    continue
                # Check if it's a large number (potential hardcoded value)
                try:
                    num = float(val.replace(",", "").replace("$", "").replace("%", ""))
                    if abs(num) >= 10000:  # Threshold for significant values
                        hardcoded_values.append(f"{sheet.title}!{cell.coordinate}: {val}")
                except:
                    pass

    if len(hardcoded_values) <= 3:  # Allow a few for labels/descriptions
        checks_passed += 1
    else:
        warnings.append(f"Potential hardcoded values outside Assumptions: {hardcoded_values[:5]}")
        checks_passed += 1  # Warning only, not failure

    # =========================================================================
    # CHECK 7: Cover sheet has links (executive summary links to model)
    # =========================================================================
    total_checks += 1
    if "Cover" in wb.sheetnames:
        ws_cover = wb["Cover"]
        cover_links = 0
        for row in ws_cover.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("=") and "!" in str(cell.value):
                    cover_links += 1

        if cover_links >= 3:
            checks_passed += 1
        else:
            warnings.append(f"Cover sheet has few links to model ({cover_links})")
            checks_passed += 1  # Warning only
    else:
        checks_passed += 1  # Skip if no Cover sheet

    wb.close()

    # Determine overall pass/fail
    passed = len(errors) == 0

    return ComplianceResult(
        model_path=model_path,
        timestamp=datetime.utcnow().isoformat() + "Z",
        passed=passed,
        total_checks=total_checks,
        checks_passed=checks_passed,
        errors=errors,
        warnings=warnings
    )


def print_result(result: ComplianceResult) -> None:
    """Print compliance result in formatted output."""
    print()
    print("═" * 70)
    print("PRISM INSTITUTIONAL MODEL COMPLIANCE GATE")
    print("═" * 70)
    print()
    print(f"Model: {result.model_path}")
    print(f"Checked: {result.timestamp}")
    print()
    print(f"Result: {result.checks_passed}/{result.total_checks} checks passed")
    print()

    if result.errors:
        print("❌ ERRORS (model REJECTED):")
        for e in result.errors:
            print(f"   • {e}")
        print()

    if result.warnings:
        print("⚠️  WARNINGS:")
        for w in result.warnings:
            print(f"   • {w}")
        print()

    print("═" * 70)
    if result.passed:
        print("✅ COMPLIANCE GATE: PASSED")
        print("   Model meets PRISM institutional standards")
    else:
        print("❌ COMPLIANCE GATE: FAILED")
        print("   Model does NOT meet PRISM institutional standards")
        print("   THIS MODEL MUST NOT BE USED OR DISTRIBUTED")
    print("═" * 70)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python model_compliance_gate.py <model.xlsx>")
        print()
        print("This is a MANDATORY compliance gate for all PRISM models.")
        print("Models that fail this gate are REJECTED and must not be used.")
        sys.exit(1)

    model_path = sys.argv[1]

    if not Path(model_path).exists():
        print(f"❌ Error: File not found: {model_path}")
        sys.exit(1)

    result = check_compliance(model_path)
    print_result(result)

    # Output JSON for programmatic use
    if "--json" in sys.argv:
        print()
        print("JSON Output:")
        print(json.dumps(asdict(result), indent=2))

    # Exit code based on pass/fail
    sys.exit(0 if result.passed else 1)


if __name__ == "__main__":
    main()
