"""Tests for InstitutionalModel — PRISM Financial Modeler."""

import os
import sys
import tempfile
import pytest

# Add scripts directory to path for import
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))

from institutional_model import InstitutionalModel, CellRecord, create_institutional_model


@pytest.fixture
def model():
    """Create a fresh InstitutionalModel for each test."""
    m = InstitutionalModel("Test Model")
    yield m
    m.close()


@pytest.fixture
def sheet(model):
    """Create a model with one worksheet."""
    return model.create_sheet("Assumptions")


class TestInitialization:
    def test_default_title(self):
        m = InstitutionalModel()
        assert m.title == "Financial Model"
        m.close()

    def test_custom_title(self, model):
        assert model.title == "Test Model"

    def test_empty_workbook(self, model):
        assert len(model.wb.sheetnames) == 0

    def test_created_at_is_utc_iso(self, model):
        assert model.created_at.endswith("Z")
        assert "T" in model.created_at

    def test_empty_ledger_and_ranges(self, model):
        assert model.cell_ledger == []
        assert model.named_ranges == {}


class TestSheetCreation:
    def test_create_sheet(self, model):
        ws = model.create_sheet("Summary")
        assert "Summary" in model.wb.sheetnames

    def test_create_multiple_sheets(self, model):
        model.create_sheet("A")
        model.create_sheet("B")
        model.create_sheet("C")
        assert model.wb.sheetnames == ["A", "B", "C"]

    def test_create_sheet_at_index(self, model):
        model.create_sheet("Second")
        model.create_sheet("First", index=0)
        assert model.wb.sheetnames[0] == "First"


class TestSetInput:
    def test_sets_value(self, model, sheet):
        model.set_input(sheet, "C4", 1_000_000)
        assert sheet["C4"].value == 1_000_000

    def test_applies_input_style(self, model, sheet):
        model.set_input(sheet, "C4", 500)
        assert sheet["C4"].style == "input"

    def test_applies_number_format(self, model, sheet):
        model.set_input(sheet, "C4", 1_000_000, model.CURRENCY)
        assert sheet["C4"].number_format == model.CURRENCY

    def test_records_in_ledger(self, model, sheet):
        model.set_input(sheet, "C4", 100, source="Appraisal p.3")
        assert len(model.cell_ledger) == 1
        record = model.cell_ledger[0]
        assert record.cell_type == "INPUT"
        assert record.value_or_formula == 100
        assert record.source == "Appraisal p.3"
        assert record.timestamp.endswith("Z")


class TestSetFormula:
    def test_sets_formula(self, model, sheet):
        model.set_formula(sheet, "C5", "=C4*1.08")
        assert sheet["C5"].value == "=C4*1.08"

    def test_applies_formula_style(self, model, sheet):
        model.set_formula(sheet, "C5", "=SUM(C4:C10)")
        assert sheet["C5"].style == "formula"

    def test_rejects_formula_without_equals(self, model, sheet):
        with pytest.raises(ValueError, match="must start with '='"):
            model.set_formula(sheet, "C5", "C4*1.08")

    def test_records_in_ledger(self, model, sheet):
        model.set_formula(sheet, "C5", "=C4*1.08", model.CURRENCY)
        record = model.cell_ledger[0]
        assert record.cell_type == "FORMULA"
        assert record.number_format == model.CURRENCY


class TestSetLink:
    def test_sets_cross_sheet_link(self, model, sheet):
        model.set_link(sheet, "B2", "=Assumptions!$C$4")
        assert sheet["B2"].value == "=Assumptions!$C$4"

    def test_applies_link_style(self, model, sheet):
        model.set_link(sheet, "B2", "=Sheet2!A1")
        assert sheet["B2"].style == "link"

    def test_rejects_link_without_exclamation(self, model, sheet):
        with pytest.raises(ValueError, match="must contain '!'"):
            model.set_link(sheet, "B2", "=C4*1.08")

    def test_records_in_ledger(self, model, sheet):
        model.set_link(sheet, "B2", "=Data!$A$1")
        assert model.cell_ledger[0].cell_type == "LINK"


class TestSetExternal:
    def test_rejects_external_without_bracket(self, model, sheet):
        with pytest.raises(ValueError, match="must contain"):
            model.set_external(sheet, "D1", "=Sheet1!A1")


class TestNamedRanges:
    def test_add_named_range(self, model, sheet):
        model.add_named_range("Principal", "Assumptions", "C4")
        assert "Principal" in model.named_ranges

    def test_sanitizes_spaces_to_underscores(self, model, sheet):
        model.add_named_range("Loan Amount", "Assumptions", "C4")
        assert "Loan_Amount" in model.named_ranges

    def test_sanitizes_hyphens_to_underscores(self, model, sheet):
        model.add_named_range("Interest-Rate", "Assumptions", "C5")
        assert "Interest_Rate" in model.named_ranges


class TestGetLedger:
    def test_returns_list_of_dicts(self, model, sheet):
        model.set_input(sheet, "C4", 100)
        model.set_formula(sheet, "C5", "=C4*2")
        ledger = model.get_ledger()
        assert len(ledger) == 2
        assert isinstance(ledger[0], dict)
        assert "cell_type" in ledger[0]
        assert "timestamp" in ledger[0]

    def test_empty_ledger(self, model):
        assert model.get_ledger() == []


class TestSaveAndLoad:
    def test_save_creates_valid_xlsx(self, model, sheet):
        model.set_input(sheet, "A1", "Test")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            model.save(path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0
            # Verify it's a valid Excel file by reopening
            from openpyxl import load_workbook
            wb = load_workbook(path)
            assert "Assumptions" in wb.sheetnames
            assert wb["Assumptions"]["A1"].value == "Test"
            wb.close()
        finally:
            os.unlink(path)


class TestConvenienceFunction:
    def test_create_institutional_model(self):
        m = create_institutional_model("Quick Model")
        assert isinstance(m, InstitutionalModel)
        assert m.title == "Quick Model"
        m.close()
