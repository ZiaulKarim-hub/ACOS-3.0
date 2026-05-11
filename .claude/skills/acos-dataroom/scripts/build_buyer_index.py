"""Generate the buyer-facing data room index workbook.

Reads the boss-finalized internal workbook and produces a sanitized version
intended for handing to a counterparty. Strips:
  - All confidence scores
  - Original filenames (only proposed renamed names appear)
  - Why-included reasoning
  - Risks / sensitivity / QA references
  - Files Excluded worksheet
  - Missing / Recommended worksheet

Output: `<DataRoomName>_Index_<YYYY-MM-DD>.xlsx` with 2 worksheets:
  WS 1 — Cover (sanitized: deal description, objective, dataroom guide, basic info only)
  WS 2 — Data Room Index (5 cols: 3 folder + renamed file + brief description)

Run AFTER the boss has reviewed and finalized the internal workbook
(typically as part of `create-room`).
"""

from __future__ import annotations

import argparse
import datetime as _dt
import importlib
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent))
import utils  # noqa: E402


BUYER_COLS = [
    "Primary Subfolder",
    "Sub-folder",
    "Sub-sub-folder",
    "File",
    "Brief Description",
]


FILL_FORWARD_COLS = ("Primary Subfolder", "Sub-folder", "Sub-sub-folder")


def _read_internal_files_included(internal_xlsx: Path) -> list[dict]:
    """Read the Files Included worksheet from the internal workbook.

    Forward-fills the visual-grouping columns (Primary Subfolder / Sub-folder /
    Sub-sub-folder) so blank cells from visual grouping are restored to their
    actual values before sorting.

    Returns rows as dicts with the original column headers as keys.
    """
    openpyxl = importlib.import_module("openpyxl")
    wb = openpyxl.load_workbook(str(internal_xlsx), data_only=True)
    if "Files Included" not in wb.sheetnames:
        raise RuntimeError(f"Worksheet 'Files Included' not found in {internal_xlsx}")
    ws = wb["Files Included"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    last: dict[str, Any] = {}
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rec = {headers[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(headers)}
        # Forward-fill the grouping columns
        for col in FILL_FORWARD_COLS:
            val = rec.get(col, "")
            if val == "" or val is None:
                if last.get(col):
                    rec[col] = last[col]
            else:
                last[col] = val
                # Reset deeper levels when a shallower level changes
                if col == "Primary Subfolder":
                    last["Sub-folder"] = ""
                    last["Sub-sub-folder"] = ""
                elif col == "Sub-folder":
                    last["Sub-sub-folder"] = ""
        out.append(rec)
    return out


def _suppress_repeats(rows: list[list[Any]], group_cols: list[int]) -> list[list[Any]]:
    """Blank repeating values in `group_cols` for visual grouping."""
    out = [list(r) for r in rows]
    last: dict[int, Any] = {}
    for i, row in enumerate(out):
        for col_idx in group_cols:
            val = row[col_idx]
            reset = False
            for shallower in group_cols:
                if shallower >= col_idx:
                    break
                if i > 0 and out[i - 1][shallower] != "" and out[i][shallower] != "":
                    if out[i - 1][shallower] != out[i][shallower]:
                        reset = True
                        break
            if val == "":
                continue
            if last.get(col_idx) == val and not reset:
                row[col_idx] = ""
            else:
                last[col_idx] = val
    return out


def _build_cover(ws, *, data_room_name: str, deal_type: str, objective: str,
                included_count: int, openpyxl) -> None:
    from openpyxl.styles import Alignment, Font

    title_font = Font(name="Calibri", size=20, bold=True, color="1F2937")
    section_font = Font(name="Calibri", size=12, bold=True, color="111827")
    label_font = Font(name="Calibri", size=11, bold=True, color="374151")
    body_font = Font(name="Calibri", size=11, color="111827")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 95

    today = _dt.date.today().isoformat()
    row = 1

    ws.cell(row=row, column=1, value=f"Data Room — {data_room_name}").font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    ws.cell(row=row, column=1, value="Deal Information").font = section_font
    row += 1
    info = [
        ("Date prepared", today),
        ("Transaction type", deal_type.replace("_", " ").title()),
        ("Objective", objective),
        ("Total documents", included_count),
    ]
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value).font = body_font
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="How to Navigate This Data Room").font = section_font
    row += 1
    deal_label = deal_type.replace("_", " ")
    guide_text = (
        f"Documents are organized by {deal_label} phase. The folder index on Worksheet 2 "
        "lists every document by primary subfolder, sub-folder, and sub-sub-folder, with "
        "a one-paragraph description of each. Files are named according to a consistent "
        "convention: [Cat#].[Sub#]_DocType_Borrower-or-Property_Date_Status.ext."
    )
    ws.cell(row=row, column=1, value="Navigation").font = label_font
    c = ws.cell(row=row, column=2, value=guide_text)
    c.font = body_font
    c.alignment = wrap
    ws.row_dimensions[row].height = 90
    row += 2

    ws.cell(row=row, column=1, value="Counterparty Notes").font = section_font
    row += 1
    notes_text = (
        "This data room represents OKOA Capital's good-faith compilation of available "
        "documentation. Documents are provided for diligence purposes. Counterparty is "
        "responsible for independent verification. Nothing in this index constitutes a "
        "warranty or representation as to completeness or accuracy."
    )
    c = ws.cell(row=row, column=1, value=notes_text)
    c.font = body_font
    c.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 70


def _build_data_room_index(ws, *, included_rows: list[dict], openpyxl) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    body_font = Font(name="Calibri", size=11, color="111827")
    wrap = Alignment(wrap_text=True, vertical="top")

    for i, h in enumerate(BUYER_COLS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if h.startswith("Primary") or h.startswith("Sub"):
            ws.column_dimensions[get_column_letter(i)].width = 32
        elif h == "File":
            ws.column_dimensions[get_column_letter(i)].width = 45
        else:
            ws.column_dimensions[get_column_letter(i)].width = 60
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(BUYER_COLS))}1"
    ws.row_dimensions[1].height = 28

    # Sort by folder hierarchy (forward-filled values are populated by the reader)
    sorted_rows = sorted(included_rows, key=lambda r: (
        r.get("Primary Subfolder") or "",
        r.get("Sub-folder") or "",
        r.get("Sub-sub-folder") or "",
        r.get("Proposed Renamed Filename") or r.get("Original Filename") or "",
    ))

    # Build raw rows then suppress repeats for visual grouping
    raw: list[list[Any]] = []
    for r in sorted_rows:
        primary = r.get("Primary Subfolder") or ""
        sub = r.get("Sub-folder") or ""
        subsub = r.get("Sub-sub-folder") or ""
        filename = r.get("Proposed Renamed Filename") or r.get("Original Filename") or ""
        description = r.get("Brief Description") or ""
        raw.append([primary, sub, subsub, filename, description])

    grouped = _suppress_repeats(raw, group_cols=[0, 1, 2])

    data_start = 2
    has_subsub = False
    for row_offset, row_data in enumerate(grouped):
        ws_row = data_start + row_offset
        if row_data[2]:
            has_subsub = True
        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=ws_row, column=col_i, value=val)
            cell.font = body_font
            cell.alignment = wrap
        ws.row_dimensions[ws_row].height = max(28, 18 + 12 * (len(str(row_data[4])) // 80))

    data_end = data_start + len(grouped) - 1

    # Apply outline grouping (collapse-by-primary-subfolder)
    if grouped:
        for r_idx in range(data_start, data_end + 1):
            primary_val = ws.cell(row=r_idx, column=1).value
            if primary_val in (None, "", " "):
                ws.row_dimensions[r_idx].outline_level = 1
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False

    # Hide Sub-sub-folder column if no row populated it
    if not has_subsub:
        ws.column_dimensions[get_column_letter(3)].hidden = True


def build_buyer_workbook(internal_xlsx: Path, *, data_room_name: str, deal_type: str,
                         objective: str, output_dir: Path | None = None) -> Path:
    openpyxl = importlib.import_module("openpyxl")

    included_rows = _read_internal_files_included(internal_xlsx)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_cover = wb.create_sheet("Cover")
    _build_cover(ws_cover, data_room_name=data_room_name, deal_type=deal_type,
                 objective=objective, included_count=len(included_rows), openpyxl=openpyxl)

    ws_index = wb.create_sheet("Data Room Index")
    _build_data_room_index(ws_index, included_rows=included_rows, openpyxl=openpyxl)

    today = _dt.date.today().isoformat()
    out_dir = output_dir or internal_xlsx.parent
    out_path = out_dir / f"{data_room_name}_Index_{today}.xlsx"
    wb.save(str(out_path))
    return out_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build buyer-facing data room index.")
    parser.add_argument("--internal-xlsx", required=True, type=Path,
                        help="Path to the boss-finalized internal workbook.")
    parser.add_argument("--data-room-name", required=True)
    parser.add_argument("--deal-type", required=True)
    parser.add_argument("--objective", required=True)
    parser.add_argument("--output-dir", type=Path, default=None,
                        help="Directory to write the buyer workbook (defaults to same as internal).")
    args = parser.parse_args(argv)

    if not args.internal_xlsx.exists():
        print(f"Internal workbook not found: {args.internal_xlsx}", file=sys.stderr)
        return 2

    out_path = build_buyer_workbook(
        args.internal_xlsx,
        data_room_name=args.data_room_name,
        deal_type=args.deal_type,
        objective=args.objective,
        output_dir=args.output_dir,
    )
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
