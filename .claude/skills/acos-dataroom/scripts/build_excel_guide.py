"""Phase 9 — Build the internal Data Room Guide workbook (4 worksheets).

Produces `<DataRoomName>_Internal_Working_<YYYY-MM-DD>.xlsx` with:

  WS 1 — Cover                       (deal info, navigation w/ per-folder summaries)
  WS 2 — Files Included              (9 cols, visual grouping + collapse)
  WS 3 — Files Excluded              (8 cols, visual grouping + collapse)
  WS 4 — Risks & Missing Documents   (key risks + missing-doc table)

Companion artifacts (separate scripts):
  build_buyer_index.py  — sanitized buyer-facing workbook (post-finalization)
  build_qa_report.py    — standalone markdown QA report

Read references/excel_schema.md for the human-readable column definitions.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import importlib
import re
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent))
import utils  # noqa: E402


# ---------------------------------------------------------------------------
# Pretty-name helpers
# ---------------------------------------------------------------------------

def humanize_folder_segment(seg: str) -> str:
    """`01_Default_and_NOD` → `01 — Default & NOD`."""
    if not seg:
        return ""
    seg = seg.strip().lstrip("/").rstrip("/")
    m = re.match(r"^(\d+(?:\.\d+)?)[_\s-]+(.+)$", seg)
    if m:
        num, name = m.group(1), m.group(2)
        name = name.replace("_", " ").replace(" and ", " & ")
        name = re.sub(r"\s+", " ", name).strip()
        return f"{num} — {name}"
    # Fallback: no leading number found
    name = seg.replace("_", " ").replace(" and ", " & ")
    return re.sub(r"\s+", " ", name).strip()


def split_room_path(path: str) -> tuple[str, str, str]:
    """`12_Payoff/12.07_Credit_Bid` → (humanized primary, humanized sub, humanized sub-sub)."""
    if not path:
        return ("", "", "")
    parts = [p for p in path.replace("\\", "/").split("/") if p.strip()]
    while len(parts) < 3:
        parts.append("")
    return (humanize_folder_segment(parts[0]),
            humanize_folder_segment(parts[1]),
            humanize_folder_segment(parts[2]))


# ---------------------------------------------------------------------------
# Taxonomy parsing (for primary-subfolder names + per-folder summaries)
# ---------------------------------------------------------------------------

def parse_taxonomy(deal_type: str, skill_dir: Path) -> dict[str, dict[str, Any]]:
    """Parse the deal-type taxonomy file.

    Returns a mapping `{ "01": {"name": "01 — Default & NOD", "raw": "01_Default_and_NOD",
                              "subfolders": ["01.01_Default_Declaration_Letter", ...] } }`
    """
    name_map: dict[str, str] = {
        "loan_sale": "loan_sale",
        "loan_participation": "loan_participation",
        "property_sale": "property_sale",
        "foreclosure_auction": "foreclosure",
        "lender_package": "lender_package",
    }
    fname = name_map.get(deal_type, deal_type)
    tax_path = skill_dir / "references" / f"taxonomy_{fname}.md"
    if not tax_path.exists():
        return {}
    text = tax_path.read_text()
    out: dict[str, dict[str, Any]] = {}
    # Top-level folders appear as `### NN_FolderName` or as bullets under "Top-Level Folders"
    for m in re.finditer(r"^###\s+(\d+)_([A-Za-z0-9_]+)\s*$", text, flags=re.MULTILINE):
        cat_num = m.group(1).zfill(2)
        raw = f"{cat_num}_{m.group(2)}"
        out[cat_num] = {"name": humanize_folder_segment(raw), "raw": raw, "subfolders": []}
    # Subfolders appear as bullets `- \`NN.NN_Name/\`` after each ### header
    if not out:
        # Fallback: parse from the top-level tree block
        tree_m = re.search(r"```\s*\n([\s\S]*?)\n```", text)
        if tree_m:
            for line in tree_m.group(1).splitlines():
                lm = re.search(r"(\d+)_([A-Za-z0-9_]+)/", line)
                if lm:
                    cat_num = lm.group(1).zfill(2)
                    if cat_num not in out:
                        raw = f"{cat_num}_{lm.group(2)}"
                        out[cat_num] = {"name": humanize_folder_segment(raw),
                                        "raw": raw, "subfolders": []}
    return out


# ---------------------------------------------------------------------------
# Folder summaries for the cover navigation section
# ---------------------------------------------------------------------------

# Per-deal-type one-liners describing what each top-level folder typically contains.
FOLDER_SUMMARIES: dict[str, dict[str, str]] = {
    "foreclosure_auction": {
        "01": "Default declarations, Notice of Default recordings, cure-period tracking, and mailings/affidavits.",
        "02": "Acceleration letters, payoff demands sent and received, reservation-of-rights correspondence.",
        "03": "Notice of Sale recordings, publication and posting affidavits, mailings to required parties, trustee substitutions.",
        "04": "Junior lienholder list and notices, return receipts, IRS §7425 notices, HOA notices, and standing-to-object filings.",
        "05": "Borrower and guarantor notices, successor-in-interest service, bankruptcy/automatic-stay analysis, SCRA service-member checks.",
        "06": "Court-supervised foreclosure filings: complaint, summons, lis pendens, judgment, order setting sale (judicial states only).",
        "07": "Trustee or sheriff appointment, sale procedures, bidder qualification rules, auction terms, as-is disclosures.",
        "08": "Title commitment dated within 30 days of sale, lien-priority chart, wraparound or senior-lien analysis.",
        "09": "Asset/bidder package: property description, valuation, photos, condition notes, environmental summary, tax and insurance status.",
        "10": "Current occupancy report, tenant leases, prior eviction actions, post-sale eviction plan, PTFA analysis.",
        "11": "Statutory and equitable redemption-rights analysis, redemption escrow setup, redemption certificates.",
        "12": "Current UPB statement, accrued and default interest, costs of sale, attorneys' fees, credit-bid worksheet, opening-bid memo.",
        "13": "Sale date/time/location memo, bidding procedures, deposit requirements, sale-recording plan.",
        "14": "Trustee's deed or sheriff's deed templates, ratification, distribution-of-proceeds analysis, excess-proceeds protocol, 1099 reporting.",
        "15": "State-specific compliance memos, procedural timelines, outside-counsel opinions.",
    },
    "loan_sale": {
        "01": "Promissory note with allonges, loan agreement and amendments, recorded assignment chain.",
        "02": "Recorded deed of trust or mortgage, assignment of rents, UCC filings and continuations, pledge and control agreements.",
        "03": "Title commitment, title policy and endorsements, ALTA survey, recorded documents.",
        "04": "Borrower entity org chart, formation documents, operating agreements, good-standing certificates, borrowing resolutions.",
        "05": "Guaranty documents, guarantor PFS, tax returns, liquidity statements.",
        "06": "Property description, photos, rent roll, operating statements, leases, tenant estoppels and SNDAs.",
        "07": "Appraisal, environmental Phase I/II, property condition report, engineering or seismic reports.",
        "08": "Property and liability insurance, builders' risk, flood, COIs with lender endorsement.",
        "09": "Servicing reports, loan ledger, payment history, escrow analysis, notices and correspondence.",
        "10": "Default notices, forbearance agreements, modifications, workout correspondence, reservation-of-rights letters.",
        "11": "Borrower tax returns, YTD financials, property T-12 / T-3, bank statements, schedule of real estate owned.",
        "12": "Property tax bills, tax-sale notices, special assessments, regulatory filings.",
        "13": "Active and historical litigation, liens and judgments, disclosure schedules.",
        "14": "Closing statement, funding memo, closing checklist, wire confirmations.",
        "15": "Loan sale agreement / PSA, bid letter, disclosure schedules, allonge and assignment to buyer, bill of sale.",
    },
    # Other deal types fall back to a generic phrase if no map entry exists.
}


def folder_summary_for(deal_type: str, cat_num: str) -> str:
    return FOLDER_SUMMARIES.get(deal_type, {}).get(
        cat_num,
        "Documents in this category for the data room.",
    )


# ---------------------------------------------------------------------------
# Worksheet column schemas
# ---------------------------------------------------------------------------

INCLUDED_COLS = [
    "Primary Subfolder",
    "Sub-folder",
    "Sub-sub-folder",
    "Original Filename",
    "Proposed Renamed Filename",
    "Brief Description",
    "Why Included",
    "Confidence — Reasoning",
    "Confidence — Overall Row",
]

EXCLUDED_COLS = [
    "Primary Subfolder",
    "Sub-folder",
    "Sub-sub-folder",
    "Original Filename",
    "Proposed Renamed Filename",
    "Reason for Exclusion",
    "Confidence — Reasoning",
    "Confidence — Overall Row",
]

MISSING_COLS = [
    "Primary Subfolder",
    "Sub-folder",
    "Sub-sub-folder",
    "Document Type Expected",
    "—",
    "Why It Would Strengthen the Data Room",
    "Confidence — Reasoning",
    "Confidence — Overall Row",
]


RISK_COLS = [
    "Severity",
    "Risk Category",
    "Risk Description",
    "Evidence Summary",
    "Recommended Action",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _coerce_float(v: Any) -> float | None:
    if v is None or v == "" or v == "n/a":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _why_text_from_classification(c: dict) -> str:
    cat = (c.get("category") or "").replace("_", " ").strip()
    sub = (c.get("subcategory") or "").replace("_", " ").strip()
    base = f"Belongs in {cat}" if cat else "Belongs in the data room"
    if sub:
        base += f" / {sub}"
    base += " for this deal type."
    return base


def _confidence_fill(value):
    from openpyxl.styles import PatternFill
    v = _coerce_float(value)
    if v is None:
        return None
    if v >= 0.85:
        return PatternFill("solid", fgColor="D1FAE5")  # soft green
    if v >= 0.70:
        return PatternFill("solid", fgColor="FEF3C7")  # soft amber
    return PatternFill("solid", fgColor="FECACA")      # soft red


def _classification_overall_confidence(c: dict) -> float | None:
    comp = [_coerce_float(c.get(k)) for k in
            ("classification_confidence", "extraction_confidence",
             "summary_confidence", "ocr_confidence")]
    valid = [s for s in comp if s is not None]
    return min(valid) if valid else None


# ---------------------------------------------------------------------------
# Visual-grouping helpers
# ---------------------------------------------------------------------------

def _suppress_repeats(rows: list[list[Any]], group_cols: list[int]) -> list[list[Any]]:
    """Blank out repeating values in `group_cols` so each group's value appears only on the first row."""
    out = [list(r) for r in rows]
    last: dict[int, Any] = {}
    for i, row in enumerate(out):
        for col_idx in group_cols:
            val = row[col_idx]
            # Reset deeper levels when an earlier group changes
            reset = False
            for shallower in group_cols:
                if shallower >= col_idx:
                    break
                if i > 0 and out[i - 1][shallower] != "" and out[i][shallower] != "":
                    if out[i - 1][shallower] != out[i][shallower]:
                        reset = True
                        break
            if val == "":
                continue
            if last.get(col_idx) == val and not reset:
                row[col_idx] = ""
            else:
                last[col_idx] = val
    return out


def _apply_outline_grouping(ws, *, data_start_row: int, data_end_row: int,
                            primary_col: int = 1) -> None:
    """Apply Excel row outlining so contiguous rows sharing a primary subfolder
    can be collapsed. Outline level 1 is set on every row EXCEPT the first row
    of each primary-subfolder group.
    """
    if data_end_row < data_start_row:
        return

    # Walk rows; the first row of each group stays at level 0; subsequent rows
    # at level 1.
    last_primary: Any = object()
    for r in range(data_start_row, data_end_row + 1):
        primary = ws.cell(row=r, column=primary_col).value
        if primary not in (None, "", " "):
            last_primary = primary
        else:
            # Empty primary cell means "same group as previous row" (suppressed)
            ws.row_dimensions[r].outline_level = 1
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False


def _hide_empty_subsubfolder(ws, columns: list[str], data_start_row: int,
                             data_end_row: int) -> None:
    """If no row has a Sub-sub-folder value, hide that column."""
    from openpyxl.utils import get_column_letter
    if "Sub-sub-folder" not in columns:
        return
    col_idx = columns.index("Sub-sub-folder") + 1
    has_value = False
    for r in range(data_start_row, data_end_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, "", " "):
            has_value = True
            break
    if not has_value:
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True


# ---------------------------------------------------------------------------
# Header builder (shared across data worksheets)
# ---------------------------------------------------------------------------

def _apply_header(ws, columns: list[str], header_row: int = 1) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for i, h in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if "Confidence" in h:
            ws.column_dimensions[get_column_letter(i)].width = 13
        elif "Severity" == h:
            ws.column_dimensions[get_column_letter(i)].width = 12
        elif h.startswith("Primary") or h.startswith("Sub"):
            ws.column_dimensions[get_column_letter(i)].width = 32
        elif "Filename" in h or "Document Type" in h:
            ws.column_dimensions[get_column_letter(i)].width = 35
        elif h == "Risk Category":
            ws.column_dimensions[get_column_letter(i)].width = 28
        elif "Description" in h or "Why" in h or "Reason" in h or "Summary" in h or "Action" in h:
            ws.column_dimensions[get_column_letter(i)].width = 55
        else:
            ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[header_row].height = 28


# ---------------------------------------------------------------------------
# Cover worksheet
# ---------------------------------------------------------------------------

def _build_cover(ws, *, data_room_name: str, deal_type: str, objective: str,
                source: str, run_id: str, included_count: int, excluded_count: int,
                missing_count: int, source_total: int, primary_folders_used: list[tuple[str, str, str]]) -> None:
    """Cover with title, run info, navigation (with per-folder summaries), counts.

    `primary_folders_used` is a list of (cat_num, humanized_name, summary_text) tuples
    sorted by cat_num — drawn from the Files Included worksheet's actual folder usage.
    """
    from openpyxl.styles import Alignment, Font

    title_font = Font(name="Calibri", size=20, bold=True, color="1F2937")
    section_font = Font(name="Calibri", size=12, bold=True, color="111827")
    label_font = Font(name="Calibri", size=11, bold=True, color="374151")
    body_font = Font(name="Calibri", size=11, color="111827")
    metric_font = Font(name="Calibri", size=11, color="0F172A")
    folder_font = Font(name="Calibri", size=11, bold=True, color="1F2937")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100

    today = _dt.date.today().isoformat()

    row = 1

    # Title
    ws.cell(row=row, column=1, value=f"Data Room — {data_room_name}").font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws.cell(row=row, column=1, value="Internal Working Workbook").font = Font(
        name="Calibri", size=12, italic=True, color="6B7280"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    # Run metadata
    ws.cell(row=row, column=1, value="Run Information").font = section_font
    row += 1
    metadata = [
        ("Date prepared", today),
        ("Run ID", run_id),
        ("Source folder", source),
        ("Deal type", deal_type.replace("_", " ").title()),
        ("Objective", objective),
    ]
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = label_font
        c = ws.cell(row=row, column=2, value=value)
        c.font = body_font
        c.alignment = wrap
        row += 1
    row += 1

    # Counts
    ws.cell(row=row, column=1, value="Cover Information").font = section_font
    row += 1
    counts = [
        ("Files in source folder", source_total),
        ("Files included in data room", included_count),
        ("Files excluded (PII / privileged / superseded)", excluded_count),
        ("Documents missing or recommended", missing_count),
    ]
    for label, value in counts:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value).font = metric_font
        row += 1
    row += 1

    # Navigation guide — DETAILED, with per-folder summaries
    ws.cell(row=row, column=1, value="How to Navigate This Data Room").font = section_font
    row += 1
    deal_label = deal_type.replace("_", " ")
    intro = (
        f"This data room is organized by {deal_label} phase. The folder index on Worksheet 2 "
        "(Files Included) lists every document by Primary Subfolder, Sub-folder, and (when "
        "applicable) Sub-sub-folder, with a brief description of each file and the skill's "
        "reasoning for why each file was placed where it was. The folders that appear in "
        "this data room are summarized below."
    )
    c = ws.cell(row=row, column=1, value="Overview").font = label_font
    c2 = ws.cell(row=row, column=2, value=intro)
    c2.font = body_font
    c2.alignment = wrap
    ws.row_dimensions[row].height = 80
    row += 2

    # Per-primary-folder summary list
    if primary_folders_used:
        ws.cell(row=row, column=1, value="Folders in this data room").font = label_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        for cat_num, folder_name, summary in primary_folders_used:
            ws.cell(row=row, column=1, value=folder_name).font = folder_font
            c = ws.cell(row=row, column=2, value=summary)
            c.font = body_font
            c.alignment = wrap
            ws.row_dimensions[row].height = max(28, 18 + 12 * (len(summary) // 90))
            row += 1
    row += 1

    # Worksheet legend
    ws.cell(row=row, column=1, value="Workbook Legend").font = section_font
    row += 1
    legend = [
        ("Worksheet 1", "Cover — this page. Run information, navigation, folder summaries."),
        ("Worksheet 2", "Files Included — every file going into the data room, grouped by primary subfolder. Click the +/− gutter at the left to collapse a primary group."),
        ("Worksheet 3", "Files Excluded — files we recommend withholding, with reasoning. Driver's licenses, privileged communications, superseded versions."),
        ("Worksheet 4", "Risks & Missing Documents — top deal-type-specific risks at the top, missing-document recommendations below."),
    ]
    for label, value in legend:
        ws.cell(row=row, column=1, value=label).font = label_font
        c = ws.cell(row=row, column=2, value=value)
        c.font = body_font
        c.alignment = wrap
        ws.row_dimensions[row].height = max(24, 18 + 12 * (len(value) // 100))
        row += 1
    row += 1

    # Footer note
    ws.cell(row=row, column=1, value="This workbook is for internal use only.").font = Font(
        name="Calibri", size=10, italic=True, color="9CA3AF"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)


# ---------------------------------------------------------------------------
# Files Included worksheet
# ---------------------------------------------------------------------------

def _build_files_included(ws, *, classifications: list[dict]) -> set[str]:
    """Build Files Included worksheet. Returns the set of category numbers actually used."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    _apply_header(ws, INCLUDED_COLS)

    body_font = Font(name="Calibri", size=11, color="111827")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Filter to included only
    included = []
    for c in classifications:
        rec = (c.get("external_room_recommendation") or "").strip().lower()
        if rec.startswith("internal_only_") or rec in ("superseded", "absent"):
            continue
        included.append(c)

    # Build rows
    rows: list[list[Any]] = []
    cats_used: set[str] = set()
    for c in included:
        primary, sub, subsub = split_room_path(c.get("proposed_room_path", ""))
        m = re.match(r"^(\d+)\b", primary)
        if m:
            cats_used.add(m.group(1).zfill(2))
        original = c.get("file_name", "") or ""
        renamed = c.get("proposed_renamed_filename", "") or ""
        description = c.get("brief_summary", "") or ""
        why = c.get("qa_notes") or _why_text_from_classification(c)
        conf_reasoning = _coerce_float(c.get("classification_confidence"))
        if conf_reasoning is not None:
            conf_reasoning = round(conf_reasoning, 2)
        conf_overall = _classification_overall_confidence(c)
        if conf_overall is not None:
            conf_overall = round(conf_overall, 2)
        rows.append([primary, sub, subsub, original, renamed, description, why,
                     conf_reasoning if conf_reasoning is not None else "",
                     conf_overall if conf_overall is not None else ""])

    # Sort by primary, then sub, then original filename
    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2]), str(r[3])))

    # Suppress repeated values in folder columns (visual grouping)
    rows = _suppress_repeats(rows, group_cols=[0, 1, 2])

    # Write
    data_start = 2
    for row_offset, row_data in enumerate(rows):
        ws_row = data_start + row_offset
        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=ws_row, column=col_i, value=val)
            cell.font = body_font
            cell.alignment = wrap
            if INCLUDED_COLS[col_i - 1].startswith("Confidence") and isinstance(val, (int, float)):
                cell.number_format = "0.00"
                fill = _confidence_fill(val)
                if fill:
                    cell.fill = fill
        ws.row_dimensions[ws_row].height = max(30, 18 + 12 * (max(
            len(str(row_data[5])) // 80, len(str(row_data[6])) // 80, 1
        )))

    data_end = data_start + len(rows) - 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(INCLUDED_COLS))}{max(data_end, 1)}"
    if rows:
        _hide_empty_subsubfolder(ws, INCLUDED_COLS, data_start, data_end)
        _apply_outline_grouping(ws, data_start_row=data_start, data_end_row=data_end)

    return cats_used


def _build_files_excluded(ws, *, classifications: list[dict],
                          internal_only: list[dict]) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    _apply_header(ws, EXCLUDED_COLS)

    body_font = Font(name="Calibri", size=11, color="111827")
    wrap = Alignment(wrap_text=True, vertical="top")

    internal_lookup = {r.get("file_id"): r for r in internal_only if r.get("file_id")}

    excluded = []
    for c in classifications:
        rec = (c.get("external_room_recommendation") or "").strip().lower()
        if rec.startswith("internal_only_") or rec in ("superseded", "absent"):
            excluded.append(c)

    rows: list[list[Any]] = []
    for c in excluded:
        primary, sub, subsub = split_room_path(c.get("proposed_room_path", ""))
        original = c.get("file_name", "") or ""
        renamed = c.get("proposed_renamed_filename", "") or ""
        rec = (c.get("external_room_recommendation") or "").lower()
        rec_label = {
            "internal_only_pii": "Contains PII",
            "internal_only_privileged": "Privileged / attorney–client material",
            "internal_only_strategic": "Strategic — boss-only",
            "superseded": "Superseded by a later version",
        }.get(rec, rec or "Excluded")
        reasoning_source = internal_lookup.get(c.get("file_id"), {}).get("reasoning", "")
        reason_full = (f"{rec_label}. {reasoning_source}" if reasoning_source
                       else f"{rec_label}. {c.get('qa_notes', '') or ''}")
        conf_reasoning = _coerce_float(c.get("classification_confidence"))
        if conf_reasoning is not None:
            conf_reasoning = round(conf_reasoning, 2)
        conf_overall = _classification_overall_confidence(c)
        if conf_overall is not None:
            conf_overall = round(conf_overall, 2)
        rows.append([primary, sub, subsub, original, renamed, reason_full,
                     conf_reasoning if conf_reasoning is not None else "",
                     conf_overall if conf_overall is not None else ""])

    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2]), str(r[3])))
    rows = _suppress_repeats(rows, group_cols=[0, 1, 2])

    data_start = 2
    for row_offset, row_data in enumerate(rows):
        ws_row = data_start + row_offset
        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=ws_row, column=col_i, value=val)
            cell.font = body_font
            cell.alignment = wrap
            if EXCLUDED_COLS[col_i - 1].startswith("Confidence") and isinstance(val, (int, float)):
                cell.number_format = "0.00"
                fill = _confidence_fill(val)
                if fill:
                    cell.fill = fill
        reason_len = len(str(row_data[5]))
        ws.row_dimensions[ws_row].height = max(30, 18 + 12 * (reason_len // 80))

    data_end = data_start + len(rows) - 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCLUDED_COLS))}{max(data_end, 1)}"
    if rows:
        _hide_empty_subsubfolder(ws, EXCLUDED_COLS, data_start, data_end)
        _apply_outline_grouping(ws, data_start_row=data_start, data_end_row=data_end)


# ---------------------------------------------------------------------------
# Risks & Missing Documents worksheet (combined WS 4)
# ---------------------------------------------------------------------------

def _build_risks_and_missing(ws, *, risks: list[dict], gaps: list[dict],
                             taxonomy: dict[str, dict[str, Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    body_font = Font(name="Calibri", size=11, color="111827")
    section_font = Font(name="Calibri", size=14, bold=True, color="1F2937")
    label_font = Font(name="Calibri", size=11, bold=True, color="374151")
    wrap = Alignment(wrap_text=True, vertical="top")
    severity_fills = {
        "critical": PatternFill("solid", fgColor="FECACA"),
        "high":     PatternFill("solid", fgColor="FED7AA"),
        "medium":   PatternFill("solid", fgColor="FEF3C7"),
        "low":      PatternFill("solid", fgColor="D1FAE5"),
    }
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}

    row = 1

    # ---- Section A: Key Risks ----
    ws.cell(row=row, column=1, value="Key Risks").font = section_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(RISK_COLS))
    row += 2

    # Risk header
    risk_header_row = row
    _apply_header(ws, RISK_COLS, header_row=risk_header_row)
    row += 1
    risk_data_start = row

    sorted_risks = sorted(risks, key=lambda r: sev_order.get(
        (r.get("severity") or "medium").lower(), 99
    ))
    for r in sorted_risks:
        sev = (r.get("severity") or "medium").lower()
        cat = r.get("category", "") or ""
        desc = r.get("risk_description", "") or ""
        evidence = r.get("evidence_summary", "") or ""
        action = r.get("recommended_action", "") or ""
        cells = [
            ws.cell(row=row, column=1, value=sev.upper()),
            ws.cell(row=row, column=2, value=cat),
            ws.cell(row=row, column=3, value=desc),
            ws.cell(row=row, column=4, value=evidence),
            ws.cell(row=row, column=5, value=action),
        ]
        for c in cells:
            c.font = body_font
            c.alignment = wrap
        if sev in severity_fills:
            cells[0].fill = severity_fills[sev]
            cells[0].alignment = Alignment(horizontal="center", vertical="center")
            cells[0].font = label_font
        ws.row_dimensions[row].height = max(28, 18 + 12 * (max(
            len(desc) // 80, len(evidence) // 80, len(action) // 80, 1
        )))
        row += 1
    risk_data_end = row - 1

    row += 2  # spacer

    # ---- Section B: Missing or Recommended Documents ----
    ws.cell(row=row, column=1, value="Missing or Recommended Documents").font = section_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(MISSING_COLS))
    row += 2

    missing_header_row = row
    _apply_header(ws, MISSING_COLS, header_row=missing_header_row)
    row += 1
    missing_data_start = row

    # Build missing rows with proper folder names (taxonomy lookup)
    sev_to_conf = {"critical": 0.95, "high": 0.85, "medium": 0.70, "low": 0.55}
    sorted_gaps = sorted(gaps, key=lambda g: (
        sev_order.get((g.get("severity") or "medium").lower(), 99),
        g.get("linked_scope_item_id") or g.get("item_id") or "",
    ))

    rows: list[list[Any]] = []
    for g in sorted_gaps:
        item_id = g.get("linked_scope_item_id") or g.get("item_id") or ""
        cat_num = item_id.split(".")[0].zfill(2) if "." in item_id else ""
        # Look up the human folder name from the taxonomy; fall back to "Category NN"
        if cat_num and cat_num in taxonomy:
            primary = taxonomy[cat_num]["name"]
        elif cat_num:
            primary = f"Category {cat_num}"
        else:
            primary = ""
        # Sub-folder: use the item_id (e.g. "01.04") with a snippet of the description
        sub = item_id if item_id else ""
        # If we have an item description with a meaningful first phrase, attach it
        item_desc = g.get("item_description", "") or ""
        if item_desc and item_id:
            short_title = item_desc.split(" — ")[0][:60]
            sub = f"{item_id} — {short_title}"
        subsub = ""
        doc_type = item_desc[:120]
        why = g.get("why_it_matters") or g.get("where_to_look", "") or ""
        severity = (g.get("severity") or "medium").lower()
        conf_reasoning = round(sev_to_conf.get(severity, 0.65), 2)
        conf_overall = 0.60
        rows.append([primary, sub, subsub, doc_type, "—", why,
                     conf_reasoning, conf_overall])

    # Sort by primary then sub
    rows.sort(key=lambda r: (str(r[0]), str(r[1])))
    rows = _suppress_repeats(rows, group_cols=[0, 1, 2])

    for row_offset, row_data in enumerate(rows):
        ws_row = missing_data_start + row_offset
        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=ws_row, column=col_i, value=val)
            cell.font = body_font
            cell.alignment = wrap
            if MISSING_COLS[col_i - 1].startswith("Confidence") and isinstance(val, (int, float)):
                cell.number_format = "0.00"
                fill = _confidence_fill(val)
                if fill:
                    cell.fill = fill
        ws.row_dimensions[ws_row].height = max(28, 18 + 12 * (len(str(row_data[5])) // 80))

    missing_data_end = missing_data_start + len(rows) - 1

    if rows:
        _hide_empty_subsubfolder(ws, MISSING_COLS, missing_data_start, missing_data_end)
        # Outline grouping on the missing-doc section only
        _apply_outline_grouping(ws, data_start_row=missing_data_start,
                                data_end_row=missing_data_end)

    # Freeze on the risk header for now
    ws.freeze_panes = ws.cell(row=risk_header_row + 1, column=1)


# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------

def build_internal_workbook(run_dir: Path, *, data_room_name: str, deal_type: str,
                            objective: str, source: str | None = None,
                            run_id: str | None = None,
                            skill_dir: Path | None = None) -> Path:
    openpyxl = importlib.import_module("openpyxl")
    skill_dir = skill_dir or Path(__file__).resolve().parents[1]

    inter = run_dir / "intermediate"
    classifications = utils.read_json(inter / "classifications.json") if (inter / "classifications.json").exists() else []
    risks = utils.read_json(inter / "risk_dashboard.json") if (inter / "risk_dashboard.json").exists() else []
    gaps = utils.read_json(inter / "gaps.json") if (inter / "gaps.json").exists() else []
    internal_only = utils.read_json(inter / "internal_only.json") if (inter / "internal_only.json").exists() else []
    manifest = utils.read_json(inter / "file_manifest.json") if (inter / "file_manifest.json").exists() else {}

    source_total = manifest.get("file_count", 0)

    included = [c for c in classifications if not (
        (c.get("external_room_recommendation") or "").lower().startswith("internal_only_")
        or (c.get("external_room_recommendation") or "").lower() in ("superseded", "absent")
    )]
    excluded = [c for c in classifications if c not in included]

    # Parse taxonomy for folder summaries + missing-doc category names
    taxonomy = parse_taxonomy(deal_type, skill_dir)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Build data sheets first so we know which categories are used
    ws_included = wb.create_sheet("Files Included")
    cats_used = _build_files_included(ws_included, classifications=classifications)

    ws_excluded = wb.create_sheet("Files Excluded")
    _build_files_excluded(ws_excluded, classifications=classifications,
                          internal_only=internal_only)

    ws_risks = wb.create_sheet("Risks & Missing Documents")
    _build_risks_and_missing(ws_risks, risks=risks, gaps=gaps, taxonomy=taxonomy)

    # Compose the per-folder summaries for the cover, sorted by category number,
    # for only the folders actually used in Files Included.
    primary_folders_used: list[tuple[str, str, str]] = []
    for cat_num in sorted(cats_used):
        if cat_num in taxonomy:
            folder_name = taxonomy[cat_num]["name"]
        else:
            folder_name = f"Category {cat_num}"
        summary = folder_summary_for(deal_type, cat_num)
        primary_folders_used.append((cat_num, folder_name, summary))

    # Now build cover with the populated folder list, then move to position 1
    ws_cover = wb.create_sheet("Cover")
    _build_cover(
        ws_cover,
        data_room_name=data_room_name,
        deal_type=deal_type,
        objective=objective,
        source=source or "",
        run_id=run_id or "",
        included_count=len(included),
        excluded_count=len(excluded),
        missing_count=len(gaps),
        source_total=source_total,
        primary_folders_used=primary_folders_used,
    )

    # Move Cover to first position
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws_cover)))

    today = _dt.date.today().isoformat()
    out_path = run_dir / f"{data_room_name}_Internal_Working_{today}.xlsx"
    wb.save(str(out_path))
    return out_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build the internal Data Room Guide workbook.")
    parser.add_argument("--run-dir", required=True, type=Path)
    parser.add_argument("--data-room-name", required=True)
    parser.add_argument("--deal-type", required=True)
    parser.add_argument("--objective", required=True)
    parser.add_argument("--source", default="")
    parser.add_argument("--run-id", default="")
    args = parser.parse_args(argv)

    layout = {"base": args.run_dir, "logs": args.run_dir / "logs",
              "extraction": args.run_dir / "extraction",
              "evidence": args.run_dir / "evidence",
              "intermediate": args.run_dir / "intermediate"}
    utils.ensure_run_dirs(layout)
    logger = utils.make_logger(layout)

    out_path = build_internal_workbook(
        args.run_dir,
        data_room_name=args.data_room_name,
        deal_type=args.deal_type,
        objective=args.objective,
        source=args.source,
        run_id=args.run_id,
    )
    logger.info("Internal workbook written: %s", out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
