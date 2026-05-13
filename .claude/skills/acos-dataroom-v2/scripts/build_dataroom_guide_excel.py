"""Build the final 2-worksheet buyer-facing Excel for acos-dataroom-v2.

Worksheet 1: Data Room Guide (rendered from phase6/ws1_guide.md)
Worksheet 2: File Tree (rendered from phase6/descriptions_final.json + taxonomy + dataroom listing)

CLI usage:
  python3 build_dataroom_guide_excel.py \
    --guide-md <PATH> \
    --descriptions-json <PATH> \
    --taxonomy-json <PATH> \
    --dataroom-dir <PATH> \
    --output <PATH>
"""
from __future__ import annotations

import argparse
import datetime as _dt
import importlib
import json
import re
import sys
from pathlib import Path
from typing import Any


HEADER_FILL_HEX = "1F2937"  # Carbon-style dark grey
HEADER_FONT_HEX = "FFFFFF"
BODY_FONT_HEX = "111827"
SUBFOLDER_LABEL_HEX = "374151"


def _parse_guide_md(md_path: Path) -> dict[str, str]:
    """Parse phase6/ws1_guide.md into section -> body text."""
    text = md_path.read_text(encoding="utf-8")
    # Extract h1 title (e.g., "Data Room — Ascent Park City")
    title_match = re.search(r"^#\s+(.+)$", text, re.MULTILINE)
    title = title_match.group(1).strip() if title_match else "Data Room"

    sections = {}
    # Match "## Section Title\n<body...>" blocks
    pattern = re.compile(r"^##\s+(.+?)\s*\n([\s\S]*?)(?=^##\s+|\Z)", re.MULTILINE)
    for m in pattern.finditer(text):
        section_name = m.group(1).strip()
        body = m.group(2).strip()
        sections[section_name] = body
    return {"_title": title, **sections}


def _build_cover_sheet(ws, guide_data: dict[str, str], file_total: int, folder_count: int, openpyxl) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    title_font = Font(name="Calibri", size=22, bold=True, color=BODY_FONT_HEX)
    section_font = Font(name="Calibri", size=13, bold=True, color=BODY_FONT_HEX)
    body_font = Font(name="Calibri", size=11, color=BODY_FONT_HEX)
    label_font = Font(name="Calibri", size=11, bold=True, color=SUBFOLDER_LABEL_HEX)
    wrap = Alignment(wrap_text=True, vertical="top")
    today = _dt.date.today().isoformat()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95

    row = 1
    # Title
    ws.cell(row=row, column=1, value=guide_data.get("_title", "Data Room")).font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    # Metadata block
    ws.cell(row=row, column=1, value="Deal Information").font = section_font
    row += 1
    info = [
        ("Date prepared", today),
        ("Folder count", folder_count),
        ("Total files", file_total),
    ]
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value).font = body_font
        row += 1
    row += 1

    # Sections from guide
    section_order = [
        "About this data room",
        "What's in this data room",
        "How to navigate",
        "A note on file naming",
        "Counterparty notes",
    ]
    for section in section_order:
        body = guide_data.get(section, "")
        if not body:
            continue
        ws.cell(row=row, column=1, value=section).font = section_font
        row += 1
        # Body cell spans both columns, wrap text
        ws.cell(row=row, column=1, value=body).font = body_font
        ws.cell(row=row, column=1).alignment = wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        line_count = body.count("\n") + max(1, len(body) // 90)
        ws.row_dimensions[row].height = max(60, 15 * line_count)
        row += 2


def _gather_file_tree(
    dataroom_dir: Path,
    taxonomy: list[dict[str, Any]],
    descriptions: dict[str, str],
) -> list[dict[str, str]]:
    """Walk the dataroom and build a (sub_folder, file, description) row list."""
    rows = []

    # taxonomy folders in order
    folder_lookup = {}
    for f in taxonomy:
        folder_num = f.get("num")
        folder_name = f.get("name", "")
        prefix = f"{int(folder_num):02d}_"
        folder_lookup[prefix] = (folder_num, folder_name)

    # Walk the dataroom
    for sub in sorted(dataroom_dir.iterdir()):
        if not sub.is_dir():
            continue
        # Match folder by number prefix
        m = re.match(r"^(\d{2})_(.+)$", sub.name)
        if m:
            folder_num = int(m.group(1))
            folder_label = f"{folder_num:02d} — {m.group(2).replace('_', ' ')}"
        else:
            folder_label = sub.name  # e.g., 00_Pending_Classification
        # List files in folder
        for f in sorted(sub.iterdir()):
            if f.is_file() and not f.name.startswith(".") and not f.name.startswith("~$"):
                # description: lookup by filename or by file_id (caller decides key)
                desc = descriptions.get(f.name, descriptions.get(str(f), ""))
                rows.append({
                    "sub_folder": folder_label,
                    "file": f.name,
                    "description": desc,
                })
    return rows


def _suppress_repeats(rows: list[dict[str, str]], col: str) -> list[dict[str, str]]:
    """Blank repeating values in `col` for visual grouping."""
    out = [dict(r) for r in rows]
    last = None
    for i, r in enumerate(out):
        cur = r.get(col, "")
        if cur and cur == last:
            r[col] = ""
        elif cur:
            last = cur
    return out


def _build_file_tree_sheet(ws, rows: list[dict[str, str]], openpyxl) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_HEX)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    body_font = Font(name="Calibri", size=11, color=BODY_FONT_HEX)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["Sub-folder", "File", "Description"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 75
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C1"

    grouped = _suppress_repeats(rows, "sub_folder")
    for offset, row in enumerate(grouped):
        r = 2 + offset
        ws.cell(row=r, column=1, value=row.get("sub_folder", "")).font = body_font
        ws.cell(row=r, column=2, value=row.get("file", "")).font = body_font
        c3 = ws.cell(row=r, column=3, value=row.get("description", ""))
        c3.font = body_font
        c3.alignment = wrap
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).alignment = wrap
        line_count = (row.get("description", "") or "").count("\n") + max(
            1, len(row.get("description", "")) // 80
        )
        ws.row_dimensions[r].height = max(20, 16 * line_count)

    # Apply outline level 1 to rows with blank sub-folder (the visual grouping)
    if grouped:
        for i, row in enumerate(grouped, start=2):
            if not row.get("sub_folder"):
                ws.row_dimensions[i].outline_level = 1
        ws.sheet_properties.outlinePr.summaryBelow = False


def build_workbook(
    guide_md: Path,
    descriptions_json: Path,
    taxonomy_json: Path,
    dataroom_dir: Path,
    output: Path,
) -> Path:
    openpyxl = importlib.import_module("openpyxl")
    guide_data = _parse_guide_md(guide_md)
    with descriptions_json.open("r", encoding="utf-8") as f:
        descriptions_raw = json.load(f)
    # Normalize descriptions to a {filename: description_text} dict
    if isinstance(descriptions_raw, dict):
        descriptions = descriptions_raw
    else:
        descriptions = {
            row.get("file") or row.get("filename"): row.get("description", "")
            for row in descriptions_raw
        }
    with taxonomy_json.open("r", encoding="utf-8") as f:
        taxonomy_data = json.load(f)
    taxonomy = taxonomy_data.get("folders", taxonomy_data) if isinstance(taxonomy_data, dict) else taxonomy_data

    rows = _gather_file_tree(dataroom_dir, taxonomy, descriptions)
    file_total = len(rows)
    folder_count = len({r["sub_folder"] for r in rows})

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_cover = wb.create_sheet("Data Room Guide")
    _build_cover_sheet(ws_cover, guide_data, file_total, folder_count, openpyxl)

    ws_tree = wb.create_sheet("File Tree")
    _build_file_tree_sheet(ws_tree, rows, openpyxl)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return output


def cli_main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build acos-dataroom-v2 Excel guide.")
    parser.add_argument("--guide-md", required=True, type=Path)
    parser.add_argument("--descriptions-json", required=True, type=Path)
    parser.add_argument("--taxonomy-json", required=True, type=Path)
    parser.add_argument("--dataroom-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args(argv)

    for p in (args.guide_md, args.descriptions_json, args.taxonomy_json, args.dataroom_dir):
        if not p.exists():
            print(f"ERROR: input not found: {p}", file=sys.stderr)
            return 2

    out = build_workbook(
        args.guide_md, args.descriptions_json, args.taxonomy_json, args.dataroom_dir, args.output
    )
    print(str(out))
    return 0


if __name__ == "__main__":
    sys.exit(cli_main())
