"""
Thurston Lux 2 — Tax Tracker XLSX Builder

Three-column financial layout: Past Due / 2025 Tax / Total Due.
Past Due = full delinquent amount including penalty + interest accrued.
2025 Tax = principal of the 2025 tax year bill (informational baseline).
Total Due = total to pay today to clear all obligations (= Past Due + any new current-year bill).
"""
import csv
import datetime as dt
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

ROOT = Path(__file__).parent
PARCELS = ROOT / "parcels.csv"
TODAY = dt.date.today().isoformat()
OUT = ROOT / f"lux2_tax_tracker_{TODAY}.xlsx"


# ── Verified data per parcel ──
# Each entry: past_due, tax_2025, total_due, status, source
DATA = {
    # Wasatch — user-supplied from emprep portal lookup 2026-04-14
    "LUX2-01": {  # Hot Pots Land
        "past_due": 107.20, "tax_2025": 95.61, "total_due": 202.81,
        "status": "Past Due",
        "source": "Wasatch portal (user lookup 2026-04-14)",
    },
    "LUX2-10": {  # Charleston Ranch — TBD; needs same Wasatch lookup
        "past_due": "", "tax_2025": "", "total_due": "",
        "status": "Not yet pulled",
        "source": "Wasatch portal — needs lookup (same flow as Hot Pots)",
    },
    # Kauai — live Playwright scrape 2026-04-14 (kauairpt.ehawaii.gov)
    "LUX2-02": {  # Anini Big House
        "past_due": 163705.92, "tax_2025": 141736.72, "total_due": 163705.92,
        "status": "Past Due",
        "source": "Kauai ehawaii.gov live 2026-04-14",
    },
    "LUX2-04": {  # Santana
        "past_due": 105831.39, "tax_2025": 91628.90, "total_due": 105831.39,
        "status": "Past Due",
        "source": "Kauai ehawaii.gov live 2026-04-14",
    },
    "LUX2-05": {  # Silver Farms
        "past_due": 119298.84, "tax_2025": 103289.04, "total_due": 119298.84,
        "status": "Past Due",
        "source": "Kauai ehawaii.gov live 2026-04-14",
    },
    # Miami-Dade — TaxSys live 2026-04-14 (folio 02-3227-045-1120)
    "LUX2-06": {  # Penthouse
        "past_due": 70281.25, "tax_2025": 70281.25, "total_due": 70281.25,
        "status": "Past Due",
        "source": "Miami-Dade TaxSys live 2026-04-14 (2025 Annual Bill unpaid)",
    },
    # Utah County — utahcounty.gov/landrecords live 2026-04-14
    "LUX2-08": {  # Sundance Big Cabin
        "past_due": 36413.97, "tax_2025": 34589.54, "total_due": 36413.97,
        "status": "Past Due",
        "source": "Utah County Treasurer payoff 2026-04-14",
    },
    "LUX2-09": {  # Sundance Small Cabin
        "past_due": 12986.07, "tax_2025": 12335.43, "total_due": 12986.07,
        "status": "Past Due",
        "source": "Utah County Treasurer payoff 2026-04-14",
    },
    "LUX2-11": {  # Sundance Sperry
        "past_due": 13677.36, "tax_2025": 12992.09, "total_due": 13677.36,
        "status": "Past Due",
        "source": "Utah County Treasurer payoff 2026-04-14",
    },
    # PR — SSN-gated
    "LUX2-03": {  # Security House
        "past_due": "", "tax_2025": "", "total_due": "",
        "status": "SSN-gated",
        "source": "CRIM requires SSN — order Debt Cert from First American PR",
    },
    "LUX2-07": {  # Primary Home
        "past_due": "", "tax_2025": "", "total_due": "",
        "status": "SSN-gated",
        "source": "CRIM requires SSN — order Debt Cert from First American PR",
    },
}


# ── URLs ──
def verify_url(juris, parcel_id, address):
    if juris == "Wasatch County":
        parts = [p.strip() for p in parcel_id.split("/")]
        serial = parts[1] if len(parts) > 1 else parcel_id
        return f"https://emprep.wasatch.utah.gov/Property-Tax-Information-Lookup/Current-Year-Property-Tax-Lookup?ser={quote(serial)}"
    if juris == "Utah County":
        nodash = parcel_id.replace(":", "")
        return f"https://www.utahcounty.gov/landrecords/Property.asp?av_serial={nodash}003"
    if juris == "Kauai County":
        return "https://kauairpt.ehawaii.gov/propertytax/"
    if juris == "Miami-Dade County":
        if "TODO" in parcel_id.upper():
            return f"https://miamidade.county-taxes.com/public/search/property_tax?pattern={quote(address)}"
        folio = parcel_id.replace("-", "")
        return f"https://miamidade.county-taxes.com/public/real_estate/parcels/{folio}"
    if juris.startswith("CRIM"):
        return "https://www.crimpr.net/"
    return ""


def manual_url(juris, parcel_id, address, entity):
    if juris == "Wasatch County":
        parts = [p.strip() for p in parcel_id.split("/")]
        parcel_num = parts[0] if parts else parcel_id
        serial = parts[1] if len(parts) > 1 else parcel_id
        subj = quote(f"Payoff Request - Parcel {parcel_num} - {address}")
        body = quote(f"Hello,\n\nPlease provide current tax payoff figures for the following parcel:\n\n"
                     f"  Parcel: {parcel_num}\n  Serial: {serial}\n  Property: {address}\n  Owner: {entity}\n\n"
                     f"This is for lender servicing/diligence purposes.\n\nThank you,\nOkoa Capital LLC\n")
        return f"mailto:taxpayoff@wasatch.utah.gov?subject={subj}&body={body}"
    if juris == "Utah County":
        today_us = dt.date.today().strftime("%m/%d/%Y")
        return f"https://www.utahcounty.gov/landrecords/TaxPayoff.asp?av_serial={parcel_id}&av_date={quote(today_us, safe='')}"
    if juris == "Kauai County":
        return "https://qpublic.schneidercorp.com/Application.aspx?App=KauaiCountyHI&PageType=Search"
    if juris == "Miami-Dade County":
        if "TODO" in parcel_id.upper():
            return "https://www.miamidadepa.gov/pa/property_search.asp"
        folio = parcel_id.replace("-", "")
        return f"https://www.miamidadepa.gov/pa/property_search.asp?folio={folio}"
    if juris.startswith("CRIM"):
        return "https://app.regrid.com/us/pr/dorado"
    return ""


def manual_label(juris):
    return {
        "Wasatch County": "Email treasurer →",
        "Utah County": "Live payoff page →",
        "Kauai County": "qPublic →",
        "Miami-Dade County": "Property Appraiser →",
    }.get(juris, "Regrid PR →" if juris.startswith("CRIM") else "Alternate →")


# ── helpers ──
def load_parcels():
    with PARCELS.open(newline="") as f:
        return list(csv.DictReader(f))


def fmt_empty(val, ph="—"):
    return ph if val is None or val == "" else val


# ── styling ──
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY = '"$"#,##0.00;[Red]("$"#,##0.00)'
RED = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")


COLUMNS = [
    ("#", 4),
    ("Property", 24),
    ("Address", 30),
    ("Parcel / Tax ID", 22),
    ("Past Due", 14),
    ("2025 Tax", 14),
    ("Total Due", 14),
    ("Status", 14),
    ("Verify Here", 16),
    ("Manual Lookup", 22),
    ("Notes", 56),
]


def main():
    parcels = load_parcels()
    wb = Workbook()
    ws = wb.active
    ws.title = "Lux 2 Tax Status"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.cell(row=1, column=1, value="Thurston Lux 2 — Property Tax Status").font = Font(bold=True, size=16, color="1F4E79")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    ws.cell(row=2, column=1, value=f"As of {TODAY}.  Past Due = currently delinquent (incl. penalty + interest).  2025 Tax = original 2025 bill principal.  Total Due = pay-today figure.").font = Font(italic=True, color="595959", size=10)
    ws.row_dimensions[2].height = 16

    # Header
    header_row = 4
    for i, (name, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=header_row, column=i, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 28

    data_start = header_row + 1
    for idx, p in enumerate(parcels, start=1):
        pid = p["property_id"]
        r = data_start + idx - 1
        juris = p["county_or_jurisdiction"]
        d = DATA.get(pid, {})

        # Build flag for Notes col
        flag = ""
        if pid == "LUX2-01":
            flag = " · GREENBELT (5-yr rollback risk if use changes)"
        elif pid == "LUX2-06":
            flag = " · FL tax-certificate sale JUN 1, 2026"
        elif pid == "LUX2-07":
            flag = " · Individually titled — verify Act 60 exemption"
        elif pid in ("LUX2-08", "LUX2-09", "LUX2-11"):
            flag = " · 2025 unpaid since Nov 30, 2025"
        elif pid in ("LUX2-02", "LUX2-04", "LUX2-05"):
            flag = " · Missed Feb 20, 2026 installment"

        source = d.get("source", "")
        notes = (source + flag) if source else flag.lstrip(" · ")

        v_url = verify_url(juris, p["parcel_id"], p["address"])
        m_url = manual_url(juris, p["parcel_id"], p["address"], p["entity_owner"])

        row = [
            idx,
            p["property_name"],
            f"{p['address']}, {p['city']} {p['state']}",
            p["parcel_id"],
            fmt_empty(d.get("past_due", "")),
            fmt_empty(d.get("tax_2025", "")),
            fmt_empty(d.get("total_due", "")),
            d.get("status", "Not yet pulled"),
            v_url,
            m_url,
            notes,
        ]

        for col_i, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_i, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_i in (5, 6, 7) and isinstance(value, (int, float)):
                cell.number_format = CURRENCY
                cell.alignment = Alignment(vertical="top", horizontal="right")
            if col_i == 8:
                cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
                cell.font = Font(bold=True)
            if col_i == 9 and value:
                cell.hyperlink = value
                cell.value = "Open portal →"
                cell.font = Font(color="0563C1", underline="single", size=10)
            if col_i == 10 and value:
                cell.hyperlink = value
                cell.value = manual_label(juris)
                cell.font = Font(color="0563C1", underline="single", size=10)
        ws.row_dimensions[r].height = 38

    last_row = data_start + len(parcels) - 1

    # Conditional formatting on Status (col 8)
    status_col = get_column_letter(8)
    rng = f"{status_col}{data_start}:{status_col}{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Past Due"'], fill=RED, font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Paid"'], fill=GREEN, font=Font(bold=True, color="006100")))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("Not yet pulled",{status_col}{data_start}))'], fill=YELLOW, font=Font(color="9C5700")))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("SSN",{status_col}{data_start}))'], fill=YELLOW, font=Font(color="9C5700")))

    # Totals row
    total_row = last_row + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value="All 11 parcels").font = Font(italic=True, color="595959")
    for col_i in (5, 6, 7):
        col_letter = get_column_letter(col_i)
        cell = ws.cell(row=total_row, column=col_i, value=f"=SUM({col_letter}{data_start}:{col_letter}{last_row})")
        cell.number_format = CURRENCY
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = BORDER

    ws.freeze_panes = f"A{data_start}"
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
